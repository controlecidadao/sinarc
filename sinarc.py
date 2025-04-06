
#########################################################
#                     S I N A R C
#    Sistema Integrado de Análise de Redes Complexas
#########################################################
 

#######################
# DESCRIÇÃO DO PROGRAMA
#######################
# O SINARC - Sistema Integrado de Análise de Redes Complexas é um programa open source experimental, em fase de desenvolvimento (versão Alfa), 
# destinado à análise computacional exploratória de redes complexas por meio de grafos.
# O objeto de análise do SINARC são as relações existentes entre pessoas físicas, pessoas jurídicas, telefones, endereços e e-mails constantes da base de dados pública de CNPJ da Receita Federal.
# Para sua criação, adotou-se uma metodologia de codificação simplificada, porém funcional, concentrando todo o código Python e Javascript em um único arquivo.
# O objetivo é demonstrar a possibilidade de criação de ferramentas tecnológicas open source que utilizam dados abertos, bem como incentivar o desenvolvimento de projetos semelhantes pela sociedade.
# O SINARC foi desenvolvido a partir do código fonte do projeto Rede CNPJ (https://github.com/rictom/rede-cnpj/).
# O programa realiza o processamento de dados públicos da Recceita Federal previamente tratados e disponibilizados à sociedade para consulta pelo projeto Rede CNPJ e
# gera uma interface gráfica interativa (página web) com recursos para:
#   a) exploração visual da rede por meio de um ambiente que combina princípios da Física (gravitação universal) e da Psicologia (percepção visual segundo a Gestalt); e
#   b) exploração automática usando algoritmos computacionais que identificam informações ocultas (nós centrais, caminhos mais curtos entre nós etc.) de acordo com a metodologia de análise definida pelo usuário.
# Ao executar o SINARC, o projeto Rede CNPJ também é carregado em outra aba do navegador, permitindo a utilização simultânea dos dois sistemas pelo usuário.
# As ferramentas de análise do SINARC são complementares às disponibilizadas pelo Rede CNPJ.


#####################################
# DESCRIÇÃO RESUMIDA DO FUNCIONAMENTO
#####################################
# A comunicação entre o front-end (Javascript) e o back-end (Python) se dá por meio da função cópia (CTRL + C) do sistema operacional.
# O programa realiza a consulta de parâmetros (CNPJ, nome do sócio, rezão social, nome fantasia etc.) à API, criada pelo Rede CNPJ, usando a biblioteca Requests do Python.
# A partir da resposta JSON do servidor (Flask), o SINARC cria dois dataframes (df_no e df_ligacao) com a biblioteca Pandas do Python.
# Esses dataframes são usados para gerar o grafo com a biblioteca Networkx do Python e aplicar os algoritmos nela disponíveis.
# O grafo no formato Networkx é utlizado pela biblioteca Pyvis do Python (criada sobre a biblioteca vis.js do Javascript) para gerar o arquivo HTML interativo do front-end.
# O arquivo HTML é então alterado via Regex com a inserção de scripts da biblioteca vis.js do Javascript.
# Por fim, o arquivo HTML é aberto com a módulo Webbrowser do Python e exibido em nova aba do navegador.
# O grafo gerado exibe as ligações existentes entre pessoas físicas, pessoas jurídicas, telefones, endereços e e-mails contantes da bases de dados pública de CNPJ da Receita Federal.


##############################################################################
# PROCEDIMENTO PARA GERAR O ARQUÍVO EXECUTÁVEL ÚNICO DO SINARC COM PYINSTALLER
##############################################################################
# 1) Abrir o prompt de comando e navegar até a pasta onde se encontram os arquivos 'sinarc.py' e 'logo.ico'
# 2) Ativar o ambiente virtual criado, onde se encontram o interpretador Python e os módulos instalados usados no programa (conda activate [nome_do_ambiente_virtual])
# 3) Instalar o módulo pyinstaller (https://pyinstaller.org/en/stable/#) no ambiente virtual (pip install pyinstaller)
# 4) Digitar no prompt de comando: pyinstaller -F --icon=logo.ico sinarc.py  <<<<<<<<<<<<<<<<<<<<<<<< CÓDIGO PARA INSERIR NO TERMINAL
# É possível substituir logo.ico e sinarc.py pelos endereços completos dos arquivos
# Para gerar o arquivo 'requirements.txt': entrar no ambiente virtual e digitar 'pip freeze > requirements.txt'. O arquivo será criado no diretório local.
# O arquivo 'requirements.txt' contém todos os módulos Python usados pelo SINARC


#################################################################################################################
# PROCEDIMENTO PARA GERAR O ARQUÍVO EXECUTÁVEL (PASTA DE ARQUIVOS) DO SINARC COM INTERFACE GRÁFICA AUTO-PY-TO-EXE
#################################################################################################################
# 1) Abrir o prompt de comando e navegar até a pasta onde se encontram os arquivos 'sinarc.py' e 'logo.ico'
# 2) Ativar o ambiente virtual criado, onde se encontram o interpretador Python e os módulos instalados usados no programa (conda activate [nome_do_ambiente_virtual])
# 3) Instalar o módulo auto-py-to-exe (https://pypi.org/project/auto-py-to-exe/) no ambiente virtual (pip install auto-py-to-exe)
# 4) Digitar no prompt de comando: auto-py-to-exe <<<<<<<<<<<<<<<<<<<<<<<< CÓDIGO PARA INSERIR NO TERMINAL
# 5) Configurar parâmetros de criação usando a interface gráfica
# Obs.: Incluir como arquivos: "sinarc.py" "help.html" "README.txt" "requirements.txt" "arquivo_network_do_pyvis_com_alteracoes.py" "print_screen.png" "exemplo_resposta_pequeno.json" "exemplo_resposta_grande.json" "logo.png" "logo.ico" 
# Se a inclusão dos arquivos não funcionar. copiar e colar manualmente na pasta 'output\sinarc'


#############################################################
# SCRIPT PARA GERAR O HASH DO ARQUIVO ZIP A SER COMPARTILHADO
#############################################################
# Criado com o Chat-GPT e NÃO TESTADO!
#import hashlib

#def zip_hash_sha3(file_path):
#    sha3 = hashlib.sha3_256()
#    with open(file_path, 'rb') as f:
#        while True:
#            data = f.read(1024)
#            if not data:
#                break
#            sha3.update(data)
#    return sha3.hexdigest()
#print(zip_hash_sha3("example.zip"))

# SITE PARA VERIFICAÇÃO DO HASH
# https://emn178.github.io/online-tools/sha256_checksum.html


##########################################################################################
# ALTERAÇÃO NECESSÁRIA NO CÓDIGO DO ARQUIVO network.py DO MÓDULO PYVIS NO AMBIENTE VIRTUAL
##########################################################################################
# Para poder gerar o arquivo executável com o módulo pyinstaller, é necessário alterar
# o arquivo 'network.py', localizado nas pastas 'site-package\pyvis' do ambiente virtual criado (Ex.: C:\Users\<nome_usuario>\.conda\envs\sinarc\Lib\site-packages\pyvis).
# O texto a ser alterado no arquivo 'network.py' tem a função de copiar todo o conteúdo da pasta pyvis que acompanha o programa (extraída do módulo pyvis) e
# colar na pasta temporária criada pelo programa e gerenciada pelo Windows.
# Foram inseridos no arquivo network.py 2 comandos 'print()', que imprimem o local da pasta onde se encontra o arquivo original network.py (source) e a pasta destino (destination).

########### INICIO PARTE ALTERADA DO ARQUIVO network.py ##########

"""
# with tempfile.mkdtemp() as tempdir: # JÁ ESTAVA COMENTADO ANTES DA ALTERAÇÃO

# LINHAS COMENTADAS PELO AUTOR:
#if os.path.exists(f"{tempdir}/lib"):
#    shutil.rmtree(f"{tempdir}/lib")
#shutil.copytree(f"{os.path.dirname(__file__)}/templates/lib", f"{tempdir}/lib")

# LINHAS INCLUÍDAS PELO AUTOR:
source = ".\\pyvis"
print('source:', source)

destination = f"{os.path.dirname(__file__)}"
print('destination:', destination)

if not os.path.exists(destination):
    shutil.copytree(source, destination)

with open(f"{tempdir}/{name}", "w+") as out:
    out.write(self.html)
    #webbrowser.open(f"{tempdir}/{name}") # <<<<<<<<< COMENTADO PORQUE ESTAVA ABRINDO O NAVEGADOR INTERNET EXPLORER E NAO O CHROME
"""

########### FINAL PARTE ALTERADA ##########


#################################################
# ALTERAÇÕES REALIZADAS NOS ARQUIVOS DO REDE CNPJ
#################################################

# ARQUIVO: 'SINARC\rede-cnpj-master\rede\rede.ini'

# parametros para o flask-limiter (COMENTADO POR SINARC - PADRÃO)
#limiter_padrao =2/second;20/minute;200/hour;400/day
#limiter_dados =10/second;600/minute
#limiter_arquivos =2/minute;30/hour;100/day

# parametros para o flask-limiter (INCLUÍDO POR SINARC - x1000)
"""
limiter_padrao =2000/second;20000/minute;200000/hour;400000/day
limiter_dados =10000/second;600000/minute
limiter_arquivos =2000/minute;30000/hour;100000/day
"""

# ARQUIVO: SINARC\rede-cnpj-master\rede\rede.py

# app.run(host='0.0.0.0',debug=True, use_reloader=False, port=porta) # Parâmetro 'use_realoader' alterado para 'False'. Em 'True' faz com que uma segunda aba do navegador seja aberta.


# ===============================================

# PARA FUNCIONAR A GEOLOCALIZAÇÃO NO REDE CNPJ
# No arquivo mapa.py, substituir "r = requests.get(url+'&format=json')" por "r = requests.get(url+'&format=json', verify=False)" para evitar erro de SSL






###################
# INFORMAÇÕES ÚTEIS
###################

# PLAYGROUND JS:
# https://jsfiddle.net/tsee1ap0/1/
# https://plnkr.co/edit/zECQ2zQiCcD71oeReEj6?p=preview&preview
# https://stackoverflow.com/questions/43886139/how-to-use-modifiers-with-click-event-in-vis-js

# CONSULTA DE CNPJS POR CIDADES PARA TESTE
# https://cadastroempresa.com.br/

# API CNPJRECEITA FEDERAL - CIDADÃO NÃO PODE CONSULTAR (WHY?)
# https://www.gov.br/conecta/catalogo/apis/consulta-cnpj

# PGFN
# https://www.regularize.pgfn.gov.br/

# SITE REDE CNPJ
# https://www.redecnpj.com.br/rede/
# https://dadosabertos.social/t/dados-relacionados-a-base-do-cnpj-liberados-pela-receita-federal-do-brasil/23

# REQUISIÇÃO FETCH API NO DEV TOOLS - MÉTODO GET
# https://stackoverflow.com/questions/14248296/making-http-requests-using-chrome-developer-tools
# MÉTODO GET:
# fetch('https://viacep.com.br/ws/01001000/json/')
#  .then(res => res.json())
#  .then(console.log)
# OU
# var a = ''
# fetch('https://viacep.com.br/ws/01001000/json/')
#    .then(resposta => {
#        return resposta.json()
#    })
#    .then(corpo => {
#        console.log(corpo)
#        a = corpo
#    });
# MÉTODO POST:
# 



#############
# PARA TESTAR
#############

# TESTAR: Atribuir à aresta um tamanho que seja proporcional ao grau do seu nó de origem/destino/ambos (sem precisar alterar as massas)



###############
# PARA CORRIGIR
###############

# Quando se seleciona um nó e clica e arrasta outro nó, na sequência, o nó anteriormente selecionado permanece com texto vermelho (convertido em funcionalidade: mudança da cor do rótulo para vermelho)



#########################################
# LISTA DE COMANDOS DO PROMPT DO ANACONDA
#########################################

# COMO ABRIR O PROMPT DO ANACONDA: NA BARRA DE PESQUISA DO WINDOWS, DIGITE:         anaconda prompt
# COMO CRIAR UM ABIENTE VIRTUAL COM O ANACONDA PROMPT:                              conda create --name myenv
# COMO ATIVAR O AMBIENTE VIRTUAL:                                                   conda activate <nome do ambiente virtual>
# COMO DESATIVAR O AMBIENTE VIRTUAL:                                                conda deactivate <nome do ambiente virtual>
# COMO INSTALAR O PIP NO AMBIENTE VIRTUAL CRIADO COM O ANACONDA PROMPT:             conda install pip
# COMO VERIFICAR OS MÓDULOS INSTALADOS EM UM AMBIENTE VIRTUAL:                      conda list
# COMO LISTAR TODOS OS AMBIENTES VIRTUAIS:                                          conda env list
# COMO DELETAR UM AMBIENTE VIRTUAL:                                                 conda env remove --name myenv

# DOWNOLOAD PORTABLE VSCODE:                                                        https://code.visualstudio.com/Download



#################
# IMPORTA MÓDULOS
#################

# MÓDULOS NATIVOS DO PYTHON
from copy import copy              # https://docs.python.org/3/library/copy.html
import json                        # https://docs.python.org/3/library/json.html
from pprint import pprint          # https://docs.python.org/3/library/pprint.html
import time                        # https://docs.python.org/3/library/time.html
import winsound                    # https://docs.python.org/pt-br/3.7/library/winsound.html
import re                          # https://docs.python.org/3/library/re.html
import os                          # https://docs.python.org/3/library/os.html
import glob                        # https://docs.python.org/3/library/glob.html
from itertools import combinations # https://docs.python.org/3/library/itertools.html
import webbrowser                  # https://docs.python.org/3/library/webbrowser.html
from io import BytesIO             # https://docs.python.org/3/library/io.html
from datetime import datetime      # https://docs.python.org/3/library/datetime.html
from traceback import format_exc   # https://docs.python.org/dev/library/traceback.
import shutil                      # https://docs.python.org/3/library/shutil.html
from zipfile import ZipFile        # https://docs.python.org/3/library/zipfile.html
            

# MÓDULOS EXTERNOS INSTALADOS NO AMBIENTE VIRTUAL
import pyperclip                   # https://pyperclip.readthedocs.io/en/latest/
import requests                    # https://requests.readthedocs.io/en/latest/
import pandas as pd                # https://pandas.pydata.org/
import numpy as np                 # https://numpy.org/
import networkx as nx              # https://networkx.org/
from pyvis.network import Network  # https://networkx.org/documentation/stable/reference/index.html
from PIL import Image              # https://python-pillow.org/
from openpyxl import styles        # https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter



###############################
# IMPRIME PREÂMBULO NO TERMINAL
###############################

print()
print('#' * 56)
print('             S I S T E M A   S I N A R C             ')
print('#' * 56)
print()



########################
# CONFIGURAÇÕES INICIAIS
########################

# Define título da janela do programa (janela do terminal)
os.system("title S I N A R C - Sistema Integrado de Análise de Redes Complexas")

# Extrai altura e largura da tela do monitor para ajustar a janela do navegador
# O ajuste da altura da área do grafo é realizado apertando-se a TECLA b
cmd = "wmic path Win32_VideoController get CurrentVerticalResolution,CurrentHorizontalResolution"
screen_width, screen_height = tuple(map(int,os.popen(cmd).read().split()[-2::]))
print('Dimensões da tela do monitor:', screen_width, 'x', screen_height)

# Desabilita a mensagem de erro 'InsecureRequestWarning' do módulo requests. Não há necessidade de validação da autenticidade da conexão com o site consultado.
requests.packages.urllib3.disable_warnings()

# Remove os arquivos Excel criados pela execução anterior do programa
files = glob.glob('arquivo_excel*.xlsx')
for file in files:
    try:
        # Remove arquivo que não esteja aberto
        os.remove(file)
    except:
        pass


# Extrai configuração inicial do arquivo 'config_sinarc.txt'
# Contém as variáveis 'cnpj_inicial' e 'num_inicial_camadas'
configs = {}
with open('config_sinarc.txt') as file:
    for line in file:
        key, value = line.strip().split('=')
        key = key.strip()
        value = value.strip()
        configs[key] = value





###################
# VARIÁVEIS GLOBAIS
###################

# Variável global criada para tentar resolver erro intermitente ainda não localizado
# A variável 'b' está apresentando o seguinte erro:
#    b = '--> ' + node2[:3] + df_no.loc[df_no['id'] == node2, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
#    UnboundLocalError: local variable 'situacao' referenced before assignment
situacao = ''

# Cria variável global para ser usada nos nomes dos arquivos Excel
num_camadas = None

# Variável global criada para poder ser usada nos nomes dos arquivos Excel
lista_de_nos = None

# Cira variável global para amazenar lista de nós da requisição imediatamente anterior (usado para destacar nós acrescentados pela requisição seguinte)
lista_nos_anterior = None

# Contador de arquivos excel abertos
#c_excel = len(glob.glob('arquivo_excel*.xlsx')) + 1

# Nome do arquivo Excel inicial
nome_do_arquivo = 'arquivo_excel.xlsx'

# 
#inserir_borda_vermelha = None


def captura_cnpj():

    """
    Utiliza a função CTRL + C do computador para:
        a) copiar números de CNPJ nos formatos 00.000.000/0000-00 ou 00000000000000 contiddos no texto selecionado;
        b) copiar nome de pessoas físicas e de pessoas jurídicas;
        c) copiar lista de 'id' dos nós selecionados diretamente da página HTML do grafo.
    Esta função é usada para receber dados da página do grafo. Ela também invoca as duas outras funções do programa (cria_dataframes, gera_grafo).
    Funciona de forma permanente no modo de captura, aguardando algum texto para ser copiado.
    Retorna None

    """



    ###################################################################################
    #                               MÓDULO PYPERCLIP
    ###################################################################################

    global num_camadas
    global lista_nos_anterior
    #global inserir_borda_vermelha

    # Variável de controle do beep no início do script
    play_beep_inicial = True

    # Submete CNPJ da Chocolates Garoto Ltda com 1 camadas (CNPJ para testes) para permitir a exploração da rede camada por camada.
    #cnpj_consulta_inicial = '28053619000183***1***' # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< Inicia com 1 camadas
    cnpj_consulta_inicial = configs['cnpj_inicial'] + '***' + configs['num_inicial_camadas'] + '***'

    # Variável global que armazena o algorítmo utilizado para destacar nós centrais
    # O usuário pressiona a TECLA z e seleciona o algorítmo desejado que é copiado ('===nome_do_algoritmo===') por script Javascript (front end) e capturado por script Python (back end).
    # O back end salva o nome do algorítmo nesta variável global para uso na montagem das redes seguintes
    # Para saber mais sobre algorítmos de centralidade, vide: What is Network Centrality? (https://www.youtube.com/watch?v=teRuQnQ3v7o)
    #algoritmo = 'betweenness'

    # Variável que armazena parâmetros da consulta anterior para uso cumulativo com a próxima consulta à base de dados
    #par_anterior = ''

    # Variável que controla ativação e desativação do modo de captura
    ativa_modo_espera = True

    cont_excel = 1

    # Inicia loop infinito do modo de captura
    while True:

        #inserir_borda_vermelha = True

        start = time.time()

        if ativa_modo_espera == True:
            print('\n' * 10)
            print('Selecione e copie o texto que contém números de CNPJs nos formatos 00.000.000/0000-00 ou 00000000000000...')
            print()
            print()


        ###############  FASE DE CAPTURA DO TEXTO  ###############
        
        # Copia '' para o clipboard (deixa o clipboard vazio)
        if ativa_modo_espera == True:
            pyperclip.copy('')

        # Beep grave curto tocado para indicar que o programa iniciou funcionamento no modo de captura
        if play_beep_inicial == True and ativa_modo_espera == True:
            winsound.Beep(600, 300) # (frequência, tempo em ms)

        # Copia para o clipboard o CNPJ da Chocolates Garoto Ltda apenas na primeira execução do script
        if len(cnpj_consulta_inicial) > 0:
            pyperclip.copy(cnpj_consulta_inicial)
            cnpj_consulta_inicial = ''
            
        # Monitora o clipboard aguardando até que seja copiado qualquer texto (entra no modo de captura)
        # O script para aqui para aguardar um texto ser copiado
        text = pyperclip.waitForPaste() # <<<<<<<<<<<<<<<<<<<<<<< PONTO DE ESPERA


        # Controle do destaque das arestas coloridas nos caminhos mais curtos
        destacar_ligacoes = True
        if '***true' in text:
            destacar_ligacoes = True
            text = text.replace('***true', '***')
        elif '***false' in text:
            destacar_ligacoes = False
            text = text.replace('***false', '***')
        #print('destacar_ligações:', destacar_ligacoes)


        # Controle de ativação do modo de captura
        if 'AAAdesativaAAA' in text and ativa_modo_espera == True:
            ativa_modo_espera = False
            pyperclip.copy(text[14:]) # Copia id do nó selecionado para a área de transferência
            print('Modo de captura desativado')
            print()

        elif 'AAAativaAAA' in text:
            ativa_modo_espera = True
            print('Modo de captura ativado')
            print()
            continue

        if ativa_modo_espera == False:
            time.sleep(1) # Intervalo entre cada loop quando o modo de captura está desativado
            continue
        elif ativa_modo_espera == True:
            pass  


        # ABRE ARQUIVOS EXCEL DE NÓS E ARESTAS (O ARQUIVO DE ARESTAS PODE NÃO EXISTIR NO CASO DE NÓS ÚNICOS)
        if text == '***ABRE-EXCEL***':

            try:
                # criando uma cópia do arquivo original (sempre existe)
                shutil.copy("arquivo_excel.xlsx", f"arquivo_excel_{num_camadas}_{len(lista_de_nos)}.xlsx")

                #try:
                    # criando uma cópia do arquivo original de arestas (pode não existir)
                #    shutil.copy("edges.csv", f"edges_{num_camadas}_{len(lista_de_nos)}.csv")

                    # abrindo a cópia do arquivo de arestas (pode não existir)
                #    os.system(f"start EXCEL.EXE edges_{num_camadas}_{len(lista_de_nos)}.csv")
                #except:
                #    pass

                # abrindo a cópia do arquivo original (sempre existe)
                os.system(f"start EXCEL.EXE arquivo_excel_{num_camadas}_{len(lista_de_nos)}.xlsx")
                continue

            # Se o arquivo estiver aberto, cria outro arquivo com um número de índice ao final (cont_excel)
            except:
                cont_excel += 1
                # criando uma cópia do arquivo original de nós
                shutil.copy("arquivo_excel.xlsx", f"arquivo_excel_{num_camadas}_{len(lista_de_nos)}_{cont_excel}.xlsx")

                #try:
                    # criando uma cópia do arquivo original de arestas (pode não existir)
                #    shutil.copy("edges.csv", f"edges_{num_camadas}_{len(lista_de_nos)}_{cont_excel}.csv")

                    # abrindo a cópia do arquivo de arestas (pode não existir)
                #    os.system(f"start EXCEL.EXE edges_{num_camadas}_{len(lista_de_nos)}_{cont_excel}.csv")
                #except:
                #    pass

                # abrindo a cópia do arquivo de nós
                os.system(f"start EXCEL.EXE arquivo_excel_{num_camadas}_{len(lista_de_nos)}_{cont_excel}.xlsx")
                continue


        # ABRE ÚLTIMO ARQUIVO HTML SALVO
        if text.lower().strip().replace('"', '') == 'abrir':
        
            # Exibe grafo no navegador padrão
            webbrowser.open('grafo_final.html', new=1, autoraise=True)
            continue

        
        # Indica início da fase de captura de texto
        print('#' * 56)
        print('# FASE DE CAPTURA DO TEXTO')
        print('#' * 56)
        print()

        # Exibe os primeiros 500 caracteres do texto copiado
        print('Texto copiado (500 primeiros caracteres):', text[:500])
        print()

        # Beep agudo longo avisa que um texto foi copiado
        winsound.Beep(1000, 300)

        #print(text)
        ## Hipóteses de erros ocorridos durante testes para PJ_08334818000152***NULL***, PJ_00000000000000
        #if ('500 Internal Server Error' in text) or ('O FOI LOCALIZADO NA BASE' in text) or ('URL was not found' in text):
        #    winsound.Beep(600, 1000)
        #    winsound.Beep(600, 1000)
        #    continue


        ############################################################### INCLUÍDO POR ÚLTIMO

        # Verifica se o texto contém nomes de pessoas físicas ou jurídicas e inclui caracteres delimitadores para requisição à API
        temp = text.split(' ')
        
        lista_4 = []
        for i in temp:

            # Número máximo de strings do nome da empresa ou da pessoa física
            if len(temp) >= 10:
                lista_4 = []
                break

            i = i.replace('\n', ' ').replace('\r', ' ').replace('#_#', '')
            i = i.strip()

            # Elimina caracteres especiais
            #i = ''.join(e for e in i if e.isalnum())

            if len(i) >= 1:

                # Se houver CNPJ na string
                if len(i) >= 3:
                    if re.search(r'\d\d\.\d\d\d\.\d\d\d/\d\d\d\d-\d\d', i) != None or re.search(r'\d{14}', i) != None or i[:3] in ['PF_', 'PJ_', 'PE_', 'EN_', 'EM_', 'TE_']:
                        lista_4 = []
                        break

                # Filtra itens e insere em lista_4
                #if (i[0].isupper() and i != '' and i != ' ') or (i in ['e', 'da', 'das', 'do', 'dos', 'de', 'dos', "d'", "D'"]):
                #if (i != '' and i != ' ') or (i in ['e', 'da', 'das', 'do', 'dos', 'de', 'dos', "d'", "D'"]):
                #    lista_4.append(i)
                lista_4.append(i)

        if lista_4 != []:
            temp = ' '.join(lista_4)
            nomes_pf_pj = "#_#" + temp + "#_#"
            print(nomes_pf_pj)


        ##############################################################################     

        # PARA ABERTURA DO MÓDULO D-TALE (NÃO IMPLEMENTADO)
        #if text == '***D-TALE***':
        #if os.path.exists('nodes.csv'):
        #    df = pd.read_csv('nodes.csv', sep=';')
        #    d = dtale.show(df)
        #    d.open_browser()
            #continue
        #else:
        #    print('Arquivo "nodes.csv" não localizado no diretório local.')
        #    print()


        # # PARA SELEÇÃO DE ALGORITMO - TECLA z
        # # Extrai padrão da string recebida da função cópia acionada a partir da página do grafo (TECLA z)
        # # Vide script Javascript que é acionado pela TECLA z
        # temp1 = re.search(r'===(.+)===', text)

        # # Verifica se foi passado algum nome de algorítmo e se o texto é uma das opções disponíveis
        # if temp1 != None: # Quando o padrão regex não é localizado, o método re.search retorna None
        #     algoritmo = temp1.group(1)
        #     if algoritmo in ['betweenness', 'closeness', 'eigenvector']:
        #         if algoritmo == 'betweenness':
        #             algoritmo = 'betweenness'
        #         elif algoritmo == 'closeness':
        #             algoritmo = 'closeness'
        #         elif algoritmo == 'eigenvector':
        #             algoritmo = 'eigenvector'
            
        #         print('Algorítmo selecionado:', algoritmo)
        #         print()

        #         # Recomeça loop
        #         continue


        ###############  FASE DE TRATAMENTO DO TEXTO COPIADO (INDEPENDENTEMENTE DO TAMANHO)  ###############
        
        # Gera uma lista de strings a partir do texto copiado, usando como caractere separador ' ' (espaço)

        # Deve oferecer tratamentos diferentes para os três tipos de textos copiados
        # 1) CNPJ copiados do texto:
        #    00.000.000/0000-00
        #    00000000000000
        # 2) IDs dos nós copiados diretamente da página do grafo (começam com prefixo): 
        #    TE_11 33343232
        #    PJ_12345678901234
        #    PF_***000000**
        #    EN_..., EM_..., PE_...
        # 3) Consulta livre por meio da digitação dos parâmetros de busca
        
        # Verifica se o usuário solicitou cumulação de pesquisa
        #cumular_pesquisa = False
        #par_ = re.search(r'AAA(\d)AAA', text)
        #if par_ != None:
        #    if int(par_.group(1)) == 1:
        #        cumular_pesquisa = True
        #    text = re.sub('AAA(\d)AAA', '', text)

        # Extrai (recorta) o numero de camadas passado pelo texto copiado diretamente da página do grafo (incluído pelo script Javascript), caso existente.
        # O script em Javascript da página do grafo copia o texto e acrescenta ao final o padrão '***[número de camadas]***'
        # A função cópia serve como meio de comunicação entre o script Javascript e o script Python
        num_camadas = re.search(r'\*\*\*(\d+)\*\*\*', text)

        if num_camadas != None:
            num_camadas = int(num_camadas.group(1))
            text = re.sub('\*\*\*(\d+)\*\*\*', '', text)
            print('Número de camadas recebido no formato ***[número]*** (requisição inicial ou página do grafo):', num_camadas)
            print()

        # Substitui 0 por 1
        if num_camadas == 0:
            num_camadas = 1
        
        # Se não for fornecido número de camadas
        if re.search(r'\*\*\*\*\*\*', text):
            num_camadas = 1

        # Ocorreu um erro em que o parâmetro de busca passado pela TECLA 'o' apresentou a string 'NULL' (em maiúsculas) como número de camadas
        # Nessa hipótese, o número de camadas deve ser 1
        # Se digitar qualquer coisa diferente de número, faz num_camada == 1
        if num_camadas == 'NULL' or num_camadas == 'null' or num_camadas == None:
            num_camadas = 1
        
        # Elimina sequência de asteriscos e parâmetros passados pela consulta realizada na página do grafo
        text = re.sub('\*\*\*\*\*\*', '', text)
        text = text.replace('******', '')
        
        # Substitui um ou mais '\n' por ' ' (espaço)
        # Esses critérios  de formatação foram definidos a partir de testes práticos
        text = re.sub(r'\\n+', ' ', text)
        
        # Substitui um ou mais '\t' por ' '
        text = re.sub(r'\\t+', ' ', text)

        # Substitui um ou mais '\s' por um único ' '
        text = re.sub(r'\s+', ' ', text) # deve ficar por último

        # Cria primeira lista de strings usando como separador ' ' (espaço)
        # A definição do caractere separador é muito importante, pois é ele quem define os limites da string a ser pesquisada no texto
        lista_1 = text.split(' ')

        # Elimina itens vazios ou com apenas 1 caractere
        lista_1 = [x for x in lista_1 if len(x) > 1]


        ###############  FASE DE CRIAÇÃO DO CONJUNTO DE PALAVRAS (PARÂMETROS DE BUSCA)  ###############

        # Cria conjunto (set) vazio para armazenar strings únicas
        conjunto_parametros = set()

        # PARA CNPJ NOS FORMATOS 00.000.000/0000-00 OU 00000000000000
        # Loop sobre a primeira lista de strings do texto copiado
        for i in lista_1:


            ###############  EXTRAI NÚMEROS DE CNPJ DO TEXTO COPIADO  ###############

            # Substitui todos os demais caracteres por '' (nada)
            i = re.sub(r'[^\d\./-]+', '', i)

            # Substitui caracteres quando iniciam a string
            i = re.sub(r'^[\./-]+', '', i)

            # Substitui caracteres quando terminam a string
            i = re.sub(r'[\./-]+$', '', i)


            #################################################

            # Extrai CNPJ com o padrão 00.000.000/0000-00
            cnpj = re.search(r'\d\d\.\d\d\d\.\d\d\d/\d\d\d\d-\d\d', i)

            # Adiciona CNPJ ao conjunto de strings únicas
            if cnpj != None:

                # Elimina caracteres que não sejam números
                cnpj = cnpj.group(0).replace('.', '').replace('/', '').replace('-', '')
                
                # Adiciona prefixo 'PJ_' e inclui CNPJ ao conjunto de strings únicas (set)
                # O prefixo PJ_ é exigido pela API do site Rede CNPJ
                cnpj = 'PJ_' + str(cnpj) # Novo formato de texto submetido ao site Rede CNPJ (o anterior não incluía o prefixo)
                conjunto_parametros.add(cnpj)

                #inserir_borda_vermelha = False


            ##################################################

            # Verifica se o número possui 14 caracteres (se tiver mais, o regex captura os 14 primeiros - não localizando na base de dados)
            if len(i) == 14:

                # Extrai CNPJ com o padrão 00000000000000
                cnpj = re.search(r'\d{14}', i) 

                # Adiciona prefixo 'PJ_' e inclui CNPJ ao conjunto de strings únicas (set)
                if cnpj != None:
                    cnpj = cnpj.group(0)
                    cnpj = 'PJ_' + str(cnpj)
                    conjunto_parametros.add(cnpj)

                    #inserir_borda_vermelha = False


        ######################################################

        # PARA NÓS SELECIONADOS NA PÁGINA DO GRAFO - TECLA o
        # Cria segunda lista de strings a partir dos IDs copiados usando como separador ',' (vírgula)
        lista_2 = text.split(',')

        cont = 0
        #destacar_ligacoes = True

        # Loop sobre itens da segunda lista de strings
        for i in lista_2:

            # Elimina espaço inicial nos itens copiados da página do grafo
            i = i.strip()

            # Verifica se a string foi copiada da página do grafo (padrões únicos identificados pelos 3 primeiros caracteres)
            # O número de camadas foi extraído no passo anterior
            # Exemplo: "PJ_09251692000115, PJ_08283040000108" (#_#12345678;12334567#_#***2*** < PODE SER CONSULTADO VIA TECLA s)
            if i[:3] == 'PF_' or i[:3] == 'PJ_' or i[:3] == 'PE_' or i[:3] == 'EN_' or i[:3] == 'EM_' or i[:3] == 'TE_':
                
                # Elimina as aspas
                i = i.replace('"', '')

                # Adiciona à lista de strings únicas (set). Como já está no formato correto para consulta, não precisa formatar.
                conjunto_parametros.add(i)
                
                # Destaca ligações (arestas) entre nós copiados da página do grafo somente se apenas dois nós forem copiados
                # Quando os CNPJ são copiados do texto, essa variável é sempre 'True'
                
                # TRECHO COMENTADO PARA DESTACAR AS LIGAÇÕES DE QUALQUER QUANTIDADE DE NÓS SELECIONADOS NO GRAFO. *** EM TESTE ***
                #cont += 1
                #if cont == 2:
                #    destacar_ligacoes = True
                #else:
                #    destacar_ligacoes = False
                #continue
       

        ######################################################

        # PARA CONSULTA LIVRE - TECLA s
        # Cria terceira lista de strings a partir do parâmetro de busca digitado pelo usuário na página do grafo
        lista_3 = re.search(r"#_#(.+)#_#", text)
        
        if lista_3 != None:
            lista_3 = lista_3.group(1)
            lista_3 = lista_3.split(';')

            # Loop sobre itens da terceira lista
            for i in lista_3:

                # Elimina espaço inicial nos itens copiados da página do grafo
                # ATENÇÃO! A ausência de tratamento do parâmetro recebido pode gerar erro na pesquisa
                i = i.strip()

                # Adiciona à lista de strings únicos (set)
                conjunto_parametros.add(i)
                continue
            

        ####################################################### INCLUÍDO POR ÚLTIMO

        if lista_4 != []:

            lista_3 = re.search(r"#_#(.+)#_#", nomes_pf_pj)
                    
            if lista_3 != None:
                lista_3 = lista_3.group(1)
                lista_3 = lista_3.split(';')

                # Loop sobre itens da terceira lista
                for i in lista_3:

                    # Elimina espaço inicial nos itens copiados da página do grafo
                    # ATENÇÃO! A ausência de tratamento do parâmetro recebido pode gerar erro na pesquisa
                    i = i.strip()

                    # Adiciona à lista de strings únicos (set)
                    conjunto_parametros.add(i)
                    continue
                    

        #######################################################

        # PARA OBTENÇÃO DE CNPJ ALEATÓRIO NA CONSULTA LIVRE (TESTE)
        if text.lower() == 'teste':
            conjunto_parametros = {'TESTE'}
            num_camadas = 1

        # Exibe parâmetros da consulta e número de camadas a ser solicitado à API
        print('Conjunto de parâmetros para consulta:', conjunto_parametros)
        print('Nº de camadas:', num_camadas)
        

        ###############  PREPARA DADOS PARA REQUISIÇÃO À API DO SITE REDE CNPJ  ###############

        
        # Verifica se foi identificado algum CNPJ
        if len(conjunto_parametros) > 0:

            # Reproduz beep grave longo para indica a existência de CNPJ no texto selecionado
            winsound.Beep(600, 1000)

            # imprime lista de CNPJ no terminal
            print()
            print('CNPJs e Ids identificados:', len(conjunto_parametros))
            if len(conjunto_parametros) > 0:
                print(conjunto_parametros)
            print()

            # Gera string com CNPJs no formato correto para envio ao site Rede CNPJ (separador ';')
            parametro = ';'.join(conjunto_parametros)

            # Define número de camadas caso não seja fornecido pelo usuário
            if num_camadas == None:
                num_camadas = 1         
            
            # Verifica se o usuário solicitou consulta cumulativa via tecla o
            #if cumular_pesquisa == True and par_anterior != '':
            #    parametro = parametro + ';' + par_anterior

            #par_anterior = copy(parametro)


            print('Parametros de requisição:', parametro)
            print('Nº de camadas:', num_camadas)
            print()


            ###############  REQUISITA DADOS À API DO SITE REDE CNPJ E CRIA DATAFRAMES  ###############

            # Indica início da fase de captura de texto
            print('#' * 56)
            print('# FASE DE REQUISIÇÃO E DE CRIAÇÃO DOS DATAFRAMES')
            print('#' * 56)
            print()


            ##############  ATUALIZA HISTÓRICO DE PARÂMETROS DE CONSULTAS  ###############

            with open('historico_de_consultas.txt', 'a') as f:
                now = datetime.now()
                agora = now.strftime(r"%A %d/%m/%Y %H:%M:%S")
                temp = agora + ' - ' + parametro + ' - ' + str(num_camadas) + '\n'
                f.write(temp)


            # Invoca função que requisita dados ao site Rede CNPJ e cria dataframes
            # recebendo como parâmetros as variáveis 'parametro' (parâmetros de busca separados por ';') e 'num_camadas' (número de camadas)
            df_no, df_ligacao = cria_dataframes(parametro, num_camada=num_camadas)

            # A não localização do parâmetro de busca retorna dataframes vazios. Reinicia loop infinito.
            if df_no.empty and df_ligacao.empty:
                print('Parâmetro de pesquisa não localizado na base de dados.')
                winsound.Beep(600, 300)
                print()
                continue

            # Parâmetro localizado!
            else:
                ###############  GERA GRAFO  ###############

                # Indica início da fase de geração do grafo
                print('#' * 56)
                print('# FASE DE GERAÇÃO DO GRAFO')
                print('#' * 56)
                print()


                # Invoca função que gera grafo
                gera_grafo(parametro, num_camadas, lista_3, destacar_ligacoes=destacar_ligacoes, df_no=df_no, df_ligacao=df_ligacao)

                # Ativa beep, caso tenha sido desativado por não ter sido identificado CNPJ na consulta anterior (vide cláusula else abaixo)
                play_beep_inicial = True

        
        # Caso o texto copiado não contenha número de CNPJ ou itens capturados da página do grafo
        else:

            # Imprime texto copiado e número de camadas no caso do texto não conter CNPJ nos formatos corretos
            print()
            print('Texto copiado (CNPJ não identificados):', text[:500])
            print('Nº de camadas (CNPJ não identificados):', num_camadas)

            # Indica a inexistência de CNPJ no texto selecionado por meio de 2 beeps curtos
            winsound.Beep(600, 300)
            winsound.Beep(600, 300)

            # Desativa o beep inicial para não ter um terceiro beep (o retorno sonoro é padronizado)
            play_beep_inicial = False

        print()
        stop = time.time() - start
        print('Tempo de execução:', round(stop, 2))


# Variável global usada dentro da função 'cria_dataframe' (em teste)
mensagem = ''

def cria_dataframes(parametro, num_camada=''):

    """
    Requisita parâmetros de busca e número de camadas à API do site Rede CNPJ (www.redecnpj.com.br).
    Recebe como parâmetros os itens a serem pesquisados, separados por ';', e o número de camadas.
    Retorna 2 dataframes: um dataframe com os nós (df_no) e outro com as arestas (df_ligacao).
    """

    #global c_excel
    global nome_do_arquivo

    ###################################################################################
    #                           MÓDULOS REQUESTS E PANDAS
    ###################################################################################

    if len(parametro.split(';')) > 1:

        # Gera lista 'final' com parâmetros para submeter à API do site Rede CNPJ
        lista_cnpj = parametro.split(';')
        final = [x.replace('\n', '') for x in lista_cnpj]

    else:
        final = [parametro]


    # BAIXA DADOS DO SITE REDE CNPJ E CRIA DATAFRAMES DE NÓS (df_no) E DE ARESTAS (df_ligacao)

    # Cria dataframes vazios
    df_no = pd.DataFrame()
    df_ligacao = pd.DataFrame()

    contador = 1
    cnpj_nao_localizados = []
    for i in final:
        
        # Verifica se o item é e-mail para deixar em letras minúsculas (exigido pela API)
        if i[:3].upper() == 'EM_':
            temp = i[3:].lower()
            i = 'EM_' + temp

        # Caso contrário, converte para maiúsculas (exigidos pela API)
        else:
            i = i.upper()



        #########################################
        # HEADERS PARA VERSÃO ONLINE DA REDE CNPJ
        #########################################
        
        # Request code / headers da consulta à API do site Rede CNPJ obtidos por meio dos seguintes passos:
        # Chrome Browser -> inspect -> Network -> Fetch/XHR -> Select desired file -> Copy all as cURL bash -> Colar em https://curlconverter.com/
        # headers = {
        # 'Accept': '*/*',
        # 'Accept-Language': 'en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7',
        # 'Connection': 'keep-alive',
        # # Already added when you pass json=
        # # 'Content-type': 'application/json',
        # 'Origin': 'https://www.redecnpj.com.br',
        # 'Referer': 'https://www.redecnpj.com.br/rede/',
        # 'Sec-Fetch-Dest': 'empty',
        # 'Sec-Fetch-Mode': 'cors',
        # 'Sec-Fetch-Site': 'same-origin',
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
        # 'sec-ch-ua': '"Google Chrome";v="107", "Chromium";v="107", "Not=A?Brand";v="24"',
        # 'sec-ch-ua-mobile': '?0',
        # 'sec-ch-ua-platform': '"Windows"',
        # }



        ############################################
        # HEADERS PARA INSTALAÇÃO LOCAL DO REDE CNPJ
        ############################################

        headers = {
            'Accept': '*/*',
            'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            # Already added when you pass json=
            # 'Content-type': 'application/json',
            'Origin': 'http://127.0.0.1:5000',
            'Referer': 'http://127.0.0.1:5000/rede/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
        }



        json_data = [
            f'{i}',
        ]

        # Substitui espaço pelos caracteres utilizados na formatação do endereço (URL)
        i = i.replace(' ', '%20')


        # Permite a consulta por e-mail (não funciona)
        # https://www.w3schools.com/tags/ref_urlencode.ASP
        #if '@' in i:
        #    i = i.lower()



        # LINK PARA CONSULTA ONLINE À REDE CNPJ
        #link = fr'https://www.redecnpj.com.br/rede/grafojson/cnpj/{num_camada}/{i}'

        # LINK PARA CONSULTA LOCAL AO REDE CNPJ
        link = fr'http://127.0.0.1:5000/rede/grafojson/cnpj/{num_camada}/{i}'



        print(f'{contador}) Requisitando item:', i)
        print(link)
        #contador += 1 # Usado apenas para relacionar e imprimir itens na tela

        # Submete consulta ao site Rede CNPJ pelo método POST (antes pela pelo método GET)
        response = requests.post(link, headers=headers, json=json_data, verify=False)

        print('Response code:', response)
        print('Response text (primeiros 500 caracteres):', response.text[:500])
        print()

        # Caso a API retorne mensagens de erro, retorna dataframes vazios
        # Hipóteses de erros ocorridos durante testes para PJ_08334818000152***NULL***, PJ_00000000000000
        #if '500 Internal Server Error' in response.text or 'O FOI LOCALIZADO NA BASE' in response.text: # MUDEI RECENTEMENTE, QUANDO TESTEI O NÚMERO MÁXIMO DE REQUISIÇÕES POR INTERVALO DE TEMPO
        if '500 Internal Server Error' in response.text:
            
            # Beep que junto com o outro formam o duplo beep de negativa
            winsound.Beep(600, 300)
            print('Mensagem de erro: 500 Internal Server Error')

            # Retorna dataframes vazios
            return df_no, df_ligacao


        # Hipóteses de erros ocorridos durante testes para PJ_08334818000152***NULL***, PJ_00000000000000
        if 'O FOI LOCALIZADO NA BASE' in response.text or 'URL was not found' in response.text:

            cnpj_nao_localizados.append(i)
            
            continue
            
            winsound.Beep(600, 300)
            #print('Mensagem de erro: ')

            # Retorna dataframes vazios
            return df_no, df_ligacao     


        # Trecho da mensagem de erro retornada pelo servidor quando se excede o limite de requisições por período
        if 'Too Many Requests' in response.text:

            # Beep que junto com o outro formam o duplo beep de negativa
            winsound.Beep(600, 300)
            print('Mensagem de erro: Excedido o limite de 200 requisições por hora. Aguarde alguns minutos antes de realizar novas requisições.')

            # Retorna dataframes vazios
            return df_no, df_ligacao

        
        # Converte resposta (response object) para JSON
        grafojson = json.loads(response.text)

        # A estrutura do arquivo JSON recebido (grafojson) possui 3 chaves:
        # 'ligacao': [],
        # 'mensagem': '',
        # 'no': []

        #pprint(grafojson) # <<<<<<<<<<<<<<<<<<<<<< É aqui que vefificamos se a resposta do site está correta
        
        if grafojson['no'] == []:
            print('Site retornou arquivo JSON vazio')
            winsound.Beep(600, 300) # Este beep é seguido pelo beep único do início do script, totalizando 2 beeps
            
            # Retorna 2 dataframes vazios (usados pela função gera_grafo) e sai da função (não executa restante do código)
            return pd.DataFrame(), pd.DataFrame()

        # Cria arquivo JSON com a resposta e salva no diretório local
        # JSON OLINE: https://jsoneditoronline.org/#
        json_object = json.dumps(grafojson, indent = 2)
        with open("resposta.json", "w") as outfile:
            json.dump(json_object, outfile)

        # Cria dataframe de nós
        # pd.json_normalize: https://pandas.pydata.org/docs/reference/api/pandas.json_normalize.html
        df_temp = pd.json_normalize(data=grafojson, record_path='no', meta='mensagem')
        
        # Concatena ao dataframe já criado 'df_no'
        df_no = pd.concat([df_no, df_temp])

        # Cria dataframe de ligações
        df_temp = pd.json_normalize(data=grafojson, record_path='ligacao')

        # Concatena ao dataframe já criado 'df_ligacao'
        df_ligacao = pd.concat([df_ligacao, df_temp])

        # Tempo de espera para realizar nova requisição ao servidor até 20 consultas (para evitar erro na versão online)
        #if len(final) > 1:
        #    print('Aguardando...')
        #    print()

            # Incluído para uso com a consulta online do site Rede NPJ
        #    time.sleep(0.5) # <<<<<<<<<<<<<<<<<<<<<<<<<<<< AJUSTAR TEMPO ENTRE REQUISIÇÕES
        
        # LIMITES DE REQUISIÇÕES DO SITE REDE CNPJ
        # 2 por segundo
        # 20 por minuto (3 por segundo)
        # 200 por hora (3,3 por minuto)

        contador += 1

    
    # Imprime CNPJs não localizados (erro de digitação, númerco com 14 dítigos capturados etc.)
    print()
    print('CNPJs não localizados:')
    print(cnpj_nao_localizados)
    print()

 
    # Cria variável global contendo mensagem de erro DANDO ERRO QUANDO PESQUISADO UM ÚNICO CNPJ QUE NÃO EXISTE 39.409.211/0001-55
    # global mensagem
    # mensagem = str(set(df_no['mensagem'].to_list()))[3:-3]
    # print('Mensagem de erro:', mensagem)
    # print()


    global mensagem
    if df_no.empty:
        pass
    else:
        mensagem = str(set(df_no['mensagem'].to_list()))[3:-3]
    print('Mensagem de erro:', mensagem)
    print()





    # ALTERA DATAFRAME df_ligacao PARA GERAR O ARQUIVO EXCEL
    # Verifica se a rede possui mais de uma ligação (não é de nó único)
    if len(df_ligacao) > 0:

        # Inclui as colunas 'id' e 'descricao' do dataframe 'df_no' no dataframe 'df_ligacao', casando 'id' e destino'
        df_ligacao = df_ligacao.merge(df_no[['id', 'descricao']], left_on='destino', right_on='id') # <<<<<<<<<<<<<<<<<< ESTÁ DANDO ERRO AO ABRIR O CNPJ 29.569.694/0001-64 COM CTRL+C

        # renomear a coluna 'descricao' para 'destino_descricao'
        df_ligacao = df_ligacao.rename(columns={'descricao': 'destino_descricao'})

        # remover a coluna 'id' trazida pela função 'merge'
        df_ligacao = df_ligacao.drop('id', axis=1)


        # Inclui as colunas 'id' e 'descricao' do dataframe 'df_no' no dataframe 'df_ligacao', casando 'id' e 'origem'
        df_ligacao = df_ligacao.merge(df_no[['id', 'descricao']], left_on='origem', right_on='id')

        # renomear a coluna 'descricao' para 'destino_descricao'
        df_ligacao = df_ligacao.rename(columns={'descricao': 'origem_descricao'})

        # remover a coluna 'id' trazida pela função 'merge'
        df_ligacao = df_ligacao.drop('id', axis=1)

        # reordenar colunas
        df_ligacao = df_ligacao[['camada', 'cor', 'origem', 'origem_descricao', 'label', 'destino_descricao', 'destino', 'tipoDescricao']]

   




    # Imprime informações do dataframe de nós
    print('=' * 56)
    print('DF_NO INFO:')
    print('=' * 56)
    print(df_no.info())
    print()

    # Imprime dataframe de nós
    print('=' * 56)
    print('DF_NO DATAFRAME:')
    print('=' * 56)
    print(df_no.head(50))
    print()




    ##############################
    # CRIA E FORMATA ARQUIVO EXCEL
    ##############################
    
    # verifica se o arquivo existe para excluí-lo
    #if os.path.exists(nome_do_arquivo):
    #    try:
    #        os.remove(nome_do_arquivo)

        # se o arquivo estiver aberto
    #    except IOError:

            # criar a cópia do arquivo
    #        nome_do_arquivo = f'arquivo_excel_{c_excel}.xlsx'
    #        c_excel += 1

    # cria arquivo Excel a partir dos dataframes
    # O arquivo cirado nunca é aberto diretamente. O que é aberto é uma cópia atualizada dele.
    with pd.ExcelWriter('arquivo_excel.xlsx') as writer:
        df_no.to_excel(writer, sheet_name='nodes')
        df_ligacao.to_excel(writer, sheet_name='edges')

    # Carrega arquivo Excel
    workbook = load_workbook('arquivo_excel.xlsx')

    # cria lista contendo nomes das planilhas
    worksheet_names = workbook.sheetnames

    # realiza loop sobre lista de nomes das planilhas
    for name in worksheet_names:

        # seleciona a planilha pelo nome
        worksheet = workbook[name]

        # deleta a primeira coluna por ser desnecessária
        worksheet.delete_cols(1)

        # alinha os títulos das colunas à esquerda e substitui cor da fonte
        for coluna in worksheet.columns:
            coluna[0].alignment = styles.Alignment(horizontal='left')
            coluna[0].font = styles.Font(color='FFFFFFFF')

        # ajusta automaticamente a largura das colunas, incluindo o texto do título
        for coluna in worksheet.columns:
            max_largura = 0
            for celula in coluna:
                if celula.value is not None:
                    largura_celula = len(str(celula.value))
                    if largura_celula > max_largura:
                        max_largura = largura_celula
            largura_titulo = len(str(coluna[0].value))
            ajuste = max(max_largura, largura_titulo) + 6
            worksheet.column_dimensions[coluna[0].column_letter].width = ajuste

        # oculta a grade
        worksheet.sheet_view.showGridLines = False

        # fixa a primeira linha
        worksheet.freeze_panes = 'A2'

        # extrai o número de linhas e colunas
        num_linhas = worksheet.max_row
        num_colunas = worksheet.max_column

        # gera a string de formatação para a tabela
        ref = f"A1:{get_column_letter(num_colunas)}{num_linhas}"

        tab = Table(displayName=name, ref=ref)

        # cria um estilo a partir de um padrão interno
        style = TableStyleInfo(name="TableStyleMedium2",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            )

        # aplica estilo à tabela
        tab.tableStyleInfo = style

        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unique through out defined names and all other table name. 
        '''

        # adiciona tabela à planilha
        worksheet.add_table(tab)

        # salva arquivo Excel
        workbook.save('arquivo_excel.xlsx')





    # Imprime informações do dataframe de ligações
    print('=' * 56)
    print('DF_LIGACAO INFO:')
    print('=' * 56)
    print(df_ligacao.info())
    print()

    # Imprime dataframe de ligações
    print('=' * 56)
    print('DF_LIGACAO DATAFRAME:')
    print('=' * 56)
    print(df_ligacao.head(50))
    print()
    
    # Retorna dataframes
    return df_no, df_ligacao


def gera_grafo(parametro, num_camadas, lista_3, destacar_ligacoes, df_no, df_ligacao):

    """
    Recebe string com parâmetros separados por ';' (parametro), número de camadas (num_camadas), lista de strings
    recebidas da consulta livre (lista_3), variável booleana para colorir aretas (destacar_ligacoes) e os dois
    dataframes (df_no e df_ligacao).
    Cria página HTML com grafo e abre no navegador padrão.
    Inclui no arquivo HTML salvo no diretório local scripts (Javascript) que permitem interagir com o grafo.
    Os algorítmos de análise são gerados pelo módulo Networkx (Python) e a exibição é feita por meio dos módulos Pyvis (Python)
    e Vis JS (Javascript).
    
    """

    global situacao
    global lista_de_nos


    ###################################################################################
    #                                MÓDULO NETWORKX
    ###################################################################################

    # Instancia gráfico direcional (arestas com seta)
    G = nx.DiGraph()
    
    # Adiciona ligações e outros valores ao grafo G
    # Os nós que não possuem ligações são adicionados ao grafo mais adiante
    # Verifica se o dataframe de ligações não está vazio
    if not df_ligacao.empty:

        print('Gerando grafo a partir dos dataframes (inserindo atributos aos nós e às arestas)...')
        print()

        # Loop sobre linhas das colunas selecionadas do dataframe => ['PF_***135248**-MIGUEL ANGEL DIAZ VARGAS', 'PJ_60409075000152', 'rep-sócio-Administrador']
        for n, i in enumerate(df_ligacao[['origem', 'destino', 'label']].values.tolist()):

            # Em cada linha (i), extrai valores das respectivas colunas (descompacta lista 'i')
            origem, destino, label = i

            # TELEFONE, EMAIL E ENDEREÇO
            # Atribui valor '' ao atributo label quando for 'TE_', 'EN_' ou 'EM_' for origem da seta
            # Salvo melhor juízo, esses itens não podem ser origem, mas apenas destino. Por isso não há necessidade de inclur texto aos labels das arestas
            if origem[:3] == 'TE_' or origem[:3] == 'EM_' or origem[:3] == 'EN_':
                label = ''

            # Atribui valor '' ao atributo label quando for 'TE_', 'EN_' ou 'EM_'
            if destino[:3] == 'TE_' or destino[:3] == 'EM_' or destino[:3] == 'EN_':
                label = ''

            # Adiciona ao grafo G pares de nós e sua aresta de ligação (nós isolados, sem ligação, serão incluídos adiante)
            if label == '':
                G.add_edge(origem, destino, arrowStrikethrough=False, label=label) # Não insere title nos casos de TE_, EM_ e EN_
            else:
                G.add_edge(origem, destino, arrowStrikethrough=False, label=label, title=label) # insere title para permitir a função hover
  
    # Calcula número de nós
    num_nos = nx.number_of_nodes(G)

    # Calcula número de arestas
    num_arestas = nx.number_of_edges(G)

    # Cria variável 'conectado' para indicar se todos os nós do grafo estão interconectados
    if num_nos > 0:
        conectado = nx.is_connected(G.to_undirected()) # Função 'is_connected' só se aplcia a grafo não direcional
        if conectado:
            conectado = 'Sim'
        else:
            conectado = 'Não'
    else:
        conectado = 'Nó único'
        num_nos = 1

    # Cria conjunto (set) com valores únicos de todos os nós existentes nas colunas 'origem' e 'destino' de df_ligacao
    set_lista = set()
    if not df_ligacao.empty:
        temp = np.unique(df_ligacao[['destino', 'origem']].values).tolist()
        for i in temp:
            set_lista.add(i)
    
    # Verifica se existem nós isolados (sem conexão com outros nós) e adiciona ao conjunto (set)
    for i in df_no['id'].tolist():
        if i not in set_lista:
            set_lista.add(i)
            conectado = 'Não'



    # Atualiza variável 'lista_nos_anterior' com o valor da variável 'lista_de_nos'
    # Essa nova lista será usada para destacar apenas os nós acrescentados ao grafo com a requisição anterior
    lista_nos_anterior = copy(lista_de_nos)
    if lista_nos_anterior != None:
        print('lista_nos_anterior:', len(lista_nos_anterior))


    # Converte conjunto (set) em lista (list)
    lista_de_nos = list(set_lista)


    if lista_de_nos != None:
        print('lista_de_nos:', len(lista_de_nos))

    if lista_nos_anterior != None and lista_de_nos != None:
        print('diferença:', len(lista_de_nos) - len(lista_nos_anterior))
    print()



    # Gera grafo sem setas (undirected graph) para aplicação de algorítmos e cálculo de alguns parâmetros
    G2 = G.to_undirected()

    # Cria lista de tuplas contendo o 'id' e o respectivo número de arestas (grau ou degree)
    # Exemplo de lista criada: [('PF_***135248**-MIGUEL ANGEL DIAZ VARGAS', 1), ('PJ_60409075000152', 19)...]
    degrees = nx.degree(G)


    # BAIXA IMAGENS DA INTERNET
    # Verifica se o arquivo PNG existe. Se não existir, baixa ele da internet.
    # O objetivo é salvar os arquivos no diretório local para não precisar baixar novamente e, assim, reduzir o tempo de execução do programa.
    # No entanto, essa redução de tempo não foi verificada na prática.
    # Como a consulta à API é realizada via internet, sempre será necessário ter acesso à internet para realizar novas consultas, mesmo que as imagens já estejam salvas no computador.
    # QUANDO TIVERMOS NOSSA PRÓPRIA API, TEREMOS QUE DISPONIBILIZAR AS IMAGENS TAMBÉM, PRINCIPALMENTE AQUELAS QUE DECORREM DE COMBINAÇÃO DE IMAGENS

    # Endereço do logo do SINARC: https://cdn-icons-png.flaticon.com/512/4803/4803070.png

    if not os.path.isfile('pf_homem.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/3220/3220358.png', verify=False)
        pf_homem = Image.open(BytesIO(response.content))
        pf_homem.save('pf_homem.png')

    # Abre arquivo PNG e atribui ele à variável que será usada mais adiante para armazenar a sobreposição de imagens. Essa sobreposição utiliza método 'inplace'
    pf_homem_com_bandeira = Image.open('pf_homem.png')

    if not os.path.isfile('pf_mulher.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/3220/3220315.png', verify=False)
        pf_mulher = Image.open(BytesIO(response.content))
        pf_mulher.save('pf_mulher.png')
    pf_mulher_com_bandeira = Image.open('pf_mulher.png')

    if not os.path.isfile('pj_exterior_ativa.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/4402/4402353.png', verify=False)
        pj_exterior_ativa = Image.open(BytesIO(response.content))
        pj_exterior_ativa.save('pj_exterior_ativa.png')
    pj_exterior_ativa_com_bandeira = Image.open('pj_exterior_ativa.png')

    if not os.path.isfile('pj_exterior_inativa.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/4400/4400008.png', verify=False)
        pj_exterior_inativa = Image.open(BytesIO(response.content))
        pj_exterior_inativa.save('pj_exterior_inativa.png')
    pj_exterior_inativa_com_bandeira = Image.open('pj_exterior_inativa.png')

    if not os.path.isfile('pj_brasil_ativa.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/1684/1684121.png', verify=False)
        pj_brasil_ativa = Image.open(BytesIO(response.content))
        pj_brasil_ativa.save('pj_brasil_ativa.png')
    pj_brasil_ativa_com_bandeira = Image.open('pj_brasil_ativa.png')

    if not os.path.isfile('pj_brasil_inativa.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/1684/1684019.png', verify=False)
        pj_brasil_inativa = Image.open(BytesIO(response.content))
        pj_brasil_inativa.save('pj_brasil_inativa.png')
    pj_brasil_inativa_com_bandeira = Image.open('pj_brasil_inativa.png')

    if not os.path.isfile('endereco.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/83/83909.png', verify=False)
        endereco = Image.open(BytesIO(response.content))
        endereco.save('endereco.png')

    if not os.path.isfile('email.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/3059/3059410.png', verify=False)
        email = Image.open(BytesIO(response.content))
        email.save('email.png')

    if not os.path.isfile('telefone.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/159/159832.png', verify=False)
        telefone = Image.open(BytesIO(response.content))
        telefone.save('telefone.png')

    if not os.path.isfile('bandeira_vermelha.png'):
        response = requests.get('https://cdn-icons-png.flaticon.com/512/395/395841.png', verify=False)
        bandeira_vermelha = Image.open(BytesIO(response.content))
        bandeira_vermelha.save('bandeira_vermelha.png')

    # Abre imagem PNG da bandeira vermelha, usada de forma sobreposta a outras imagens.
    bandeira_vermelha = Image.open('bandeira_vermelha.png')

    # Redefine tamanho da imagem
    bandeira_vermelha = bandeira_vermelha.resize((400, 400))

    # Alterações "inplace". Números se referem à posição da imagem da frente em relação ao canto superior esquerdo (x, y)
    # Bandeira à esquerda: -58, 210
    # Bandeira à direita: 168, 210
    # Para alterar a posição da bandeira, necessário apagar imagens anteriores salvas no diretório local
    pf_homem_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha) 
    
    # Salva imagens sobrepostas como um novo arquivo
    pf_homem_com_bandeira.save("pf_homem_com_bandeira.png")

    pf_mulher_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha)
    pf_mulher_com_bandeira.save("pf_mulher_com_bandeira.png")

    pj_exterior_ativa_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha)
    pj_exterior_ativa_com_bandeira.save("pj_exterior_ativa_com_bandeira.png")

    pj_exterior_inativa_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha)
    pj_exterior_inativa_com_bandeira.save("pj_exterior_inativa_com_bandeira.png")

    pj_brasil_ativa_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha)
    pj_brasil_ativa_com_bandeira.save("pj_brasil_ativa_com_bandeira.png")

    pj_brasil_inativa_com_bandeira.paste(bandeira_vermelha, (-58, 210), bandeira_vermelha)
    pj_brasil_inativa_com_bandeira.save("pj_brasil_inativa_com_bandeira.png")



    # Extrai nomes da colunas do dataframe 'df_no'
    nomes_cols = df_no.columns
    print('Colunas do dataframe "df_no":')
    pprint(list(nomes_cols))
    print()

    # Calcula o grau do grafo
    grau = int(len(list(G.degree())))
    
    # Adiciona imagens aos nós
    for i in lista_de_nos:
    
        # Se o nó ainda não existir no grafo, inclui ele (ele vai ficar sem conexão com os demais)
        if not G.has_node(i):
            G.add_node(i)

            # Atualiza variável 'num_nos' para incluir os nós sem conexão
            # Só incrementa se o grafo tiver mais de um nó (não for nó único)
            if num_nos > 1:
                num_nos = num_nos + 1


        # DEFINE ATRIBUTOS INDIVIDUAIS DE CADA NÓ NO MÓDULO NETWORKX (image, shape, label, title, size etc.)



        ##########################################################################################################
        # CÁLCULO DA DISTÂNCIA ENTRE O ALVO ÚNICO E OS DEMAIS NÓS E ATRIBUIÇÃO EM CADA NÓ COMO ATRIBUTO 'distance'
        ##########################################################################################################

        # https://www.datacamp.com/tutorial/social-network-analysis-python
        # https://coderzcolumn.com/tutorials/data-science/network-analysis-in-python-node-importance-and-paths-networkx#5.4

        # Verifica se só existe um único parâmetro de busca, não foi realizada consulta livre e exite pelo menos dois nós (grau > 1)
        if ';' not in parametro and lista_3 == None and grau > 1:

            # Calcula o número de nós que separa do nó-alvo (tem que ter pelo menos dois nós)
            distance = nx.shortest_path_length(G2, parametro, i)

        else:
            distance = '-'



        #########################################################
        # ADICIONA 'DEGREE SIZE' EM CADA NÓ COMO ATRIBUTO 'value'
        #########################################################

        # Loop sobre cada tupla da lista
        # Cria lista de tuplas contendo o 'id' e o respectivo grau
        # [('PF_***135248**-MIGUEL ANGEL DIAZ VARGAS', 1), ('PJ_60409075000152', 19)...]
        for j in degrees:

            if j[0] == i:

                # Cria atributo 'value' em cada nó com o valor do respectivo grau
                # https://networkx.org/documentation/stable/reference/generated/networkx.classes.function.set_node_attributes.html
                # https://coderzcolumn.com/tutorials/data-science/network-analysis-in-python-node-importance-and-paths-networkx#5.4
                nx.set_node_attributes(G, {f'{i}': f'{j[1]}'}, 'value')



        #################################################################################
        # EXTRAI ITENS FACULTATIVOS DO DATAFRAME 'df_no' (QUE NEM SEMPRE ESTÃO PRESENTES)
        #################################################################################

        # Cria string 'divida' que será incrementada por cada espécie de dívida (sida, prev e fgts), caso existente
        # Essa variável é usasa para apresentar imagem com bandeira vermelha
        dividas = ''
        
        if i[:3] == 'PJ_' or i[:3] == 'PE_' or i[:3] == 'PF_':

            nome_fantasia = ''

            if 'nome_fantasia' in nomes_cols:
                if f"{df_no.loc[df_no['id'] == i, 'nome_fantasia'].iloc[0]}".strip() != '':
                    nome_fantasia = f"\n{df_no.loc[df_no['id'] == i, 'nome_fantasia'].iloc[0]}"

            if 'pgfn-sida' in nomes_cols:
                temp = df_no.loc[df_no['id'] == i, 'pgfn-sida'].iloc[0]
                if isinstance(temp, str):
                    pgfn_sida = df_no.loc[df_no['id'] == i, 'pgfn-sida'].iloc[0]
                    dividas = dividas + ', PGFN-SIDA: ' + pgfn_sida

            if 'pgfn-fgts' in nomes_cols:
                temp = df_no.loc[df_no['id'] == i, 'pgfn-fgts'].iloc[0]
                if isinstance(temp, str):
                    pgfn_fgts = df_no.loc[df_no['id'] == i, 'pgfn-fgts'].iloc[0]
                    dividas = dividas + ', PGFN-FGTS: ' + pgfn_fgts

            if 'pgfn-prev' in nomes_cols:
                temp = df_no.loc[df_no['id'] == i, 'pgfn-prev'].iloc[0]
                if isinstance(temp, str):
                    pgfn_prev = df_no.loc[df_no['id'] == i, 'pgfn-prev'].iloc[0]
                    dividas = dividas + ', PGFN-PREV: ' + pgfn_prev

            # Só se aplica a PJs. Também é responsável pela inclusão de bandeira vermelha ao nó.
            if 'acordo_leniência' in nomes_cols:
                acordo_leniencia = df_no.loc[df_no['id'] == i, 'acordo_leniência'].iloc[0]
                if isinstance(acordo_leniencia, str):
                    dividas = dividas + ', Acordo Leniência: Sim'
                    
                    # Resumo do acordo de leniência impresso no terminal
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(acordo_leniencia, str):
                        print(f'Acordo de Leniência: {temp}')
                        print(acordo_leniencia)
                        print()

            if 'ceis' in nomes_cols:
                ceis = df_no.loc[df_no['id'] == i, 'ceis'].iloc[0]
                if isinstance(ceis, str):
                    dividas = dividas + ', CEIS: Sim'
                    
                    # Cadastro Nacional de Empresas Inidôneas e Suspensas (CEIS) impresso no terminal
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(ceis, str):
                        print(f'CEIS: {temp}')
                        print(ceis)
                        print()

            if 'cnep' in nomes_cols:
                cnep = df_no.loc[df_no['id'] == i, 'cnep'].iloc[0]
                if isinstance(cnep, str):
                    dividas = dividas + ', CNEP: Sim'
                    
                    # Cadastro Nacional de Empresas Punidas (CNEP) impresso no terminal
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(cnep, str):
                        print(f'CNEP: {temp}')
                        print(cnep)
                        print()

            if 'cepim' in nomes_cols:
                cepim = df_no.loc[df_no['id'] == i, 'cepim'].iloc[0]
                if isinstance(cepim, str):
                    dividas = dividas + ', CEPIM: Sim'
                    
                    # Cadastro de Entidades Privadas Sem Fins Lucrativos Impedidas (CEPIM)
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(cepim, str):
                        print(f'CEPIM: {temp}')
                        print(cepim)
                        print()

            if 'pep' in nomes_cols:
                pep = df_no.loc[df_no['id'] == i, 'pep'].iloc[0]
                if isinstance(pep, str):
                    dividas = dividas + ', PEP: Sim'
                    
                    # Cadastro de Pessoas Expostas Politicamente (PEP)
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(pep, str):
                        print(f'PEP: {temp}')
                        print(pep)
                        print()

            if 'ceaf' in nomes_cols:
                ceaf = df_no.loc[df_no['id'] == i, 'ceaf'].iloc[0]
                if isinstance(ceaf, str):
                    dividas = dividas + ', CEAF: Sim'
                    
                    # Cadastro de Expulsões da Administração Federal (CEAF)
                    temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    if isinstance(ceaf, str):
                        print(f'CEAF: {temp}')
                        print(ceaf)
                        print()

            if 'situacao_fiscal' in nomes_cols:
                situacao_fiscal = df_no.loc[df_no['id'] == i, 'situacao_fiscal'].iloc[0]
                if isinstance(situacao_fiscal, str):
                    dividas = dividas + f', Situação Fiscal: {situacao_fiscal}'

                    # Cadastro de Expulsões da Administração Federal (CEAF)
                    #temp = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
                    #if isinstance(situacao_fiscal, str):
                    #    print(f'Situação Fiscal: {temp}')
                    #    print(situacao_fiscal)
                    #    print()



        # INÍCIO DA INSERÇÃO DOS ATRIBUTOS DOS NÓS
        # PESSOA FÍSICA ########################################################      
        if i[:3] == 'PF_':

            # Inclui imagem de homem
            if df_no.loc[df_no['id'] == i, 'sexo'].iloc[0] == 1.0 or df_no.loc[df_no['id'] == i, 'sexo'].iloc[0] == 0.0:  # homem
                if dividas == '':
                    G.nodes[i]['image'] = 'pf_homem.png'
                else:
                    G.nodes[i]['image'] = 'pf_homem_com_bandeira.png'

            # Inclui imagem de mulher
            elif df_no.loc[df_no['id'] == i, 'sexo'].iloc[0] == 2.0:  # mulher
                if dividas == '':
                    G.nodes[i]['image'] = 'pf_mulher.png'
                else:
                    G.nodes[i]['image'] = 'pf_mulher_com_bandeira.png'

            G.nodes[i]['shape'] = 'image'

            lista_no = [x for x in nx.all_neighbors(G, i)]
            lista_final = []

            # Loop sobre os pares de nós de todas as ligações do grafo
            for node1, node2, data in G.edges.data(): # node1 = origem, node2 = destino
                
                # Quando o nó de origem for igual ao nó i (PF_)
                # Para entender essa parte, abrir arquivo edges.csv no Excel e verificar quem faz par com PF_ sendo origem
                # PF_ pode ter como destino outra PF_
                if node1 == i:

                    # Objetivo é criar variável 'situacao' para adicionar aos nós 'PF_' a informação se as empresas a ele ligadas estão ativas ou baixadas
                    # Verifica se o nó de destino (node2) possui situação ativa ou baixada
                    if node2[:3] == 'PJ_' or node2[:3] == 'PE_':
                        if df_no.loc[df_no['id'] == node2, 'situacao_ativa'].iloc[0] == True:
                            situacao = ' - Ativa -'
                        elif df_no.loc[df_no['id'] == node2, 'situacao_ativa'].iloc[0] == False:
                            situacao = ' - Baixada -'
                    else:
                        situacao = ''

                    # Cria linha com informações sobre os nós adjacentes -> ['--> PJ_CHOCOLATES GAROTO LTDA. - Ativa - (Administrador)']
                    # A variável 'b' está apresentando o seguinte erro:
                    #    b = '--> ' + node2[:3] + df_no.loc[df_no['id'] == node2, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
                    #    UnboundLocalError: local variable 'situacao' referenced before assignment
                    # Solução em teste: criar variável global situacao = '' fora da função
                    # Outra solução possível: incluir try/except para tentar descobrir qual nó está gerando erro
                    a = data['label']
                    b = '--> ' + node2[:3] + df_no.loc[df_no['id'] == node2, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
                    lista_final.append(b)

                # Quando o nó de destino for igual ao nó i
                elif node2 == i:

                    # Verifica se o nó de origem (node1) possui situação ativa ou baixada
                    # Não vimos ainda a hipótese de o nó de origem ser PF e o de destino ser PJ ou PE
                    # Por isso, a condição 'else' é acionada
                    if node1[:3] == 'PJ_' or node1[:3] == 'PE_':
                        if df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == True:
                            situacao = ' - Ativa -'
                        elif df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == False:
                            situacao = ' - Baixada -'
                    else:
                        situacao = ''

                    a = data['label']
                    b = '<-- ' + node1[:3] + df_no.loc[df_no['id'] == node1, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
                    lista_final.append(b)
            
            lista_final = sorted(lista_final)
            nome = df_no.loc[df_no["id"] == i, "descricao"].iloc[0] + f' ({i.split("-")[0]})'
            temp = f'{nome}\n\nNÓ - Distância Alvo: {distance}{dividas}\nNós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_final)
            G.nodes[i]['title'] = temp

            G.nodes[i]['label'] = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]

            G.nodes[i]['size'] = 20


        # PESSOA JURÍDICA ###########################################################
        elif i[:3] == 'PJ_':

            # Pessoa jurídica domiciliada no exterior ativa
            if df_no.loc[df_no['id'] == i, 'uf'].iloc[0] == 'EX' and df_no.loc[df_no['id'] == i, 'situacao_ativa'].iloc[0] == True:
                #G.nodes[i]['image'] = 'pj_exterior_ativa.png' #'https://cdn-icons-png.flaticon.com/512/4402/4402353.png'

                if dividas == '':
                    G.nodes[i]['image'] = 'pj_exterior_ativa.png'
                else:
                    G.nodes[i]['image'] = 'pj_exterior_ativa_com_bandeira.png'

            # Pessoa jurídica domiciliada no exterior baixada
            elif df_no.loc[df_no['id'] == i, 'uf'].iloc[0] == 'EX' and df_no.loc[df_no['id'] == i, 'situacao_ativa'].iloc[0] == False:
                #G.nodes[i]['image'] = 'pj_exterior_inativa.png' #'https://cdn-icons-png.flaticon.com/512/4400/4400008.png'
                
                if dividas == '':
                    G.nodes[i]['image'] = 'pj_exterior_inativa.png'
                else:
                    G.nodes[i]['image'] = 'pj_exterior_inativa_com_bandeira.png'

            # Pessoa jurídica domiciliada no Brasil ativa (condição verificada somente se as anteriores não forem satisfeitas: primeiro ver exterior)
            elif df_no.loc[df_no['id'] == i, 'situacao_ativa'].iloc[0] == True:
                G.nodes[i]['image'] = 'pj_brasil_ativa.png' #'https://cdn-icons-png.flaticon.com/512/1684/1684121.png'

                if dividas == '':
                    G.nodes[i]['image'] = 'pj_brasil_ativa.png'
                else:
                    G.nodes[i]['image'] = 'pj_brasil_ativa_com_bandeira.png'
                    #G.nodes[i]['image'] = json.dumps(str(pj_brasil_ativa_com_bandeira_2.tobitmap()))


            # Pessoa jurídica domiciliada no exterior baixada
            elif df_no.loc[df_no['id'] == i, 'situacao_ativa'].iloc[0] == False:
                #G.nodes[i]['image'] = 'pj_brasil_inativa.png' #'https://cdn-icons-png.flaticon.com/512/1684/1684019.png'

                if dividas == '':
                    G.nodes[i]['image'] = 'pj_brasil_inativa.png'
                else:
                    G.nodes[i]['image'] = 'pj_brasil_inativa_com_bandeira.png'

            G.nodes[i]['shape'] = 'image'

            lista_no = [x for x in nx.all_neighbors(G, i)]
            lista_final = []
            for node1, node2, data in G.edges.data(): # node1 = origem, node2 = destino
                if node1 == i:

                    # Quando o nó-alvo é 'TE_', 'EN_' ou 'EM_', 'id' em 'df_no' tem que ser 'node1' (origem)
                    if node2[:3] == 'TE_' or node2[:3] == 'EM_' or node2[:3] == 'EN_':
                        if df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == True:
                            situacao = ' - Ativa -'
                        elif df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == False:
                            situacao = ' - Baixada -'
                        else:
                            situacao = ''
                    
                    # Noes demais casos, 'id' em 'df_no' tem que ser 'node2' (destino)
                    else:
                        if df_no.loc[df_no['id'] == node2, 'situacao_ativa'].iloc[0] == True:
                            situacao = ' - Ativa -'
                        elif df_no.loc[df_no['id'] == node2, 'situacao_ativa'].iloc[0] == False:
                            situacao = ' - Baixada -'
                        else:
                            situacao = ''

                    a = data['label']
                    if a != '':
                        b = '--> ' + node2[:3] + df_no.loc[df_no['id'] == node2, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
                    else:
                        b = '--> ' + node2
                    lista_final.append(b)
                elif node2 == i:
                    if df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == True:
                        situacao = ' - Ativa -'
                    elif df_no.loc[df_no['id'] == node1, 'situacao_ativa'].iloc[0] == False:
                        situacao = ' - Baixada -'
                    else:
                        situacao = ''
                    a = data['label']
                    if a != '':
                        b = '<-- ' + node1[:3] + df_no.loc[df_no['id'] == node1, 'descricao'].iloc[0] + f'{situacao}' + f' ({a})'
                    else:
                        b = '<-- ' + node1
                    lista_final.append(b)
            lista_final = sorted(lista_final)

            # Descrição completa da rede incluída apenas em PESSOA JURÍDICA
            # INCLUÍDO TRY (nem sempre tem a coluna 'logradouro_complemento')
            try:
                logradouro_complemento = '- ' + df_no.loc[df_no["id"] == i, "logradouro_complemento"].iloc[0] + ' -'
            except:
                logradouro_complemento = '-'
            endereco = f'{df_no.loc[df_no["id"] == i, "logradouro"].iloc[0]} {logradouro_complemento} {df_no.loc[df_no["id"] == i, "municipio"].iloc[0]} - {df_no.loc[df_no["id"] == i, "uf"].iloc[0]}'
            temp = f'{df_no.loc[df_no["id"] == i, "descricao"].iloc[0] + f" ({i})"}{nome_fantasia}\n{endereco}\n\nNÓ - Distância Alvo: {distance}{dividas}\n GRAFO - Nós: {num_nos}, Arestas: {num_arestas}, Camadas Solicitadas: {num_camadas}, Conectado: {conectado}\nMensagem: {mensagem}\n Nós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_final)
            G.nodes[i]['title'] = temp
            
            G.nodes[i]['label'] = df_no.loc[df_no['id'] == i, 'descricao'].iloc[0]
            
            if df_no.loc[df_no['id'] == i, 'uf'].iloc[0] == 'EX':
                G.nodes[i]['size'] = 30
            else:
                G.nodes[i]['size'] = 20


        # PESSOA JURÍDICA DOMICILIADA NO EXTERIOR (OU EXTRANGEIRA?) ############################################
        elif i[:3] == 'PE_': # <<<<<<<<<<< PREFIXO DIFERENTE DE PJ_
            
            # Pessoa jurídica extrangeira não possui informação sobre se está ativa ou inativa (coluna 'situacao' em branco)
            # Por isso, a imagem será sempre de pj domiciliada no externo ativa
            #G.nodes[i]['image'] = 'pj_exterior_ativa.png' #'https://cdn-icons-png.flaticon.com/512/4402/4402353.png'

            if dividas == '':
                G.nodes[i]['image'] = 'pj_exterior_ativa.png'
            else:
                G.nodes[i]['image'] = 'pj_exterior_ativa_com_bandeira.png'

            G.nodes[i]['shape'] = 'image'

            lista_no = [x for x in nx.all_neighbors(G, i)]
            lista_final = []
            for node1, node2, data in G.edges.data():
                if node1 == i:
                    a = data['label']
                    if a != '':
                        b = '--> ' + node2[:3] + df_no.loc[df_no['id'] == node2, 'descricao'].iloc[0] + f' ({a})'
                    else:
                        b = '--> ' + node2
                    lista_final.append(b)
                elif node2 == i:
                    a = data['label']
                    if a != '':
                        b = '<-- ' + node1[:3] + df_no.loc[df_no['id'] == node1, 'descricao'].iloc[0] + f' ({a})'
                    else:
                        b = '<-- ' + node1
                    lista_final.append(b)
            lista_final = sorted(lista_final)
            endereco = f'{df_no.loc[df_no["id"] == i, "logradouro"].iloc[0]} - {df_no.loc[df_no["id"] == i, "logradouro_complemento"].iloc[0]} - {df_no.loc[df_no["id"] == i, "municipio"].iloc[0]} - {df_no.loc[df_no["id"] == i, "uf"].iloc[0]}'
            
            # Ao tentar abrir chocolates garoto em 9 camadas, a empresa OMNICON GROUP INC não exibe razão social e CNPJ corretos no title (parece ser problema na consulta ao banco de dados)
            temp = f'{df_no.loc[df_no["id"] == i, "descricao"].iloc[0] + f" ({i})"}{nome_fantasia}\n{endereco}\n\nNÓ - Distância Alvo: {distance}{dividas}\n GRAFO - Nós: {num_nos}, Arestas: {num_arestas}, Camadas Solicitadas: {num_camadas}, Conectado: {conectado}\nMensagem: {mensagem}\n Nós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_final)
            #temp = f'{df_ligacao.loc[df_ligacao["id"] == i, "descricao"].iloc[0] + f" ({i})"}{nome_fantasia}\n{endereco}\n\nNÓ - Distância Alvo: {distance}{dividas}\n GRAFO - Nós: {num_nos}, Arestas: {num_arestas}, Camadas Solicitadas: {num_camadas}, Conectado: {conectado}\nMensagem: {mensagem}\n Nós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_final)
            
            
            G.nodes[i]['title'] = temp
            G.nodes[i]['label'] = i[3:]
            G.nodes[i]['size'] = 30


        # EMAIL ######################################################################
        elif i[:3] == 'EM_':
            G.nodes[i]['image'] = 'email.png' #'https://cdn-icons-png.flaticon.com/512/3059/3059410.png'
            G.nodes[i]['shape'] = 'image'
            lista_no = ['<-- ' + x[:3] + df_no.loc[df_no["id"] == x, "descricao"].iloc[0] for x in nx.all_neighbors(G, i)]
            lista_no = sorted(lista_no)
            temp = f'{i[3:]}\n\nNÓ - Distância Alvo: {distance}\nNós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_no)
            G.nodes[i]['title'] = temp
            G.nodes[i]['label'] = i[3:]#' '
            G.nodes[i]['size'] = 15


        # ENDEREÇO ####################################################################
        elif i[:3] == 'EN_':
            G.nodes[i]['image'] = 'endereco.png' #'https://cdn-icons-png.flaticon.com/512/83/83909.png'
            G.nodes[i]['shape'] = 'image'
            lista_no = ['<-- ' + x[:3] + df_no.loc[df_no["id"] == x, "descricao"].iloc[0] for x in nx.all_neighbors(G, i)]
            lista_no = sorted(lista_no)
            temp = f'{i[3:]}\n\nNÓ - Distância Alvo: {distance}\nNós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_no)
            G.nodes[i]['title'] = temp
            G.nodes[i]['label'] = ' - '.join(i.split('-')[-2:])
            G.nodes[i]['size'] = 15


        # TELEFONE ###################################################################
        elif i[:3] == 'TE_':
            G.nodes[i]['image'] = 'telefone.png' #'https://cdn-icons-png.flaticon.com/512/159/159832.png'
            G.nodes[i]['shape'] = 'image'
            lista_no = ['<-- ' + x[:3] + df_no.loc[df_no["id"] == x, "descricao"].iloc[0] for x in nx.all_neighbors(G, i)]
            lista_no = sorted(lista_no)
            temp = f'{i[3:]}\n\nNÓ - Distância Alvo: {distance}\nNós Adjacentes (Grau {len(lista_no)}):\n' + '\n'.join(lista_no)
            G.nodes[i]['title'] = temp
            G.nodes[i]['label'] = i[3:]
            G.nodes[i]['size'] = 15
    
    
    # OBS.: Para a aplicação de alguns algorítmos é necessário que a rede esteja completa (não seja adicionado nó posteriormente)
    # seja conectada (não tenha nós isolados)

    ######################################################################
    # DEPTH-FIRST SEARCH - DFS_TREE (BUSCA EM PROFUNDIDADE)
    ######################################################################

    t0 = time.time()

    # Apenas para grafos conectados, parâmetro de busca único, não foi realizada consulta livre e com mais de um nó
    if nx.is_connected(G.to_undirected()) and ';' not in parametro and lista_3 == None and grau > 1:
    
        # Realizada inversão do sentido das setas (G.reverse) antes da aplicação do algorítmo
        dfs_tree = nx.dfs_tree(G.reverse(), source=parametro).reverse()
        print('DEPTH-FIRST SEARCH (DFS_TREE):')
        print('Número de nós da árvore:', dfs_tree.number_of_nodes())
        
        # Extrai todos os nós da árvore
        dfs_dic = {}
        for i in dfs_tree.nodes:
            temp = nx.shortest_path_length(dfs_tree.reverse(), source=parametro, target=i)
            dfs_dic[i] = temp

        # Ordena nós da árvore (key) pelo respectivo grau (value)
        dfs_list = sorted(dfs_dic.items(), key=lambda x : x[1], reverse=False)

        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        dfs2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(dfs_list):
            key = G.nodes[i[0]]['label']
            value = i[1]
            dfs2.insert(n, (key, value))

        pprint(dfs2)

        # Inclui atributo aos nós indicando se pertencem ou não à estrutura da árvore. Esse atributo será utilizado pelo front para destacar os nós da árvore
        # A letra z no início do nome do atributo faz o mesmo ser posicionado ao final da lista de atributos do nó,
        # não prejudicando a substituição dos primeiros atributos por meio do módulo re.
        for i in lista_de_nos:
            if i in dfs_tree.nodes:
                G.nodes[i]['z_ancestor'] = 'true'
            else:
                G.nodes[i]['z_ancestor'] = 'false'

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()


    
    ######################################################################
    # PAGERANK
    ######################################################################

    # Designado especialmente para grafos direcionados, mas pode ser usado com grafos não direcionados (G.to_undirected()).
    # Foi necessário instalar o módulo scipy
    # Ainda não vislumbrei utilidade para esse algorítmo, mas deixei a exibição no terminal
    
    # https://networkx.org/documentation/stable/reference/algorithms/generated/networkx.algorithms.link_analysis.pagerank_alg.pagerank.html
    # https://www.youtube.com/watch?v=P8Kt6Abq_rM&t=319s


    print('=' * 56)
    print('PAGERANK (ARESTAS BIDIRECIONAIS):')
    print('=' * 56)

    t0 = time.time()

    if nx.is_connected(G.to_undirected()):

        # Cria cópia de G para eliminar todos os nós sem relevância (TE, EN e EM)
        G_copy = G.copy()

        # Converte gráfico para não direcionado. Quando o grafo é não direcionada, o algoritmo PageRank converte para bidirecional (aresta com duas setas)
        # https://networkx.org/documentation/stable/reference/algorithms/generated/networkx.algorithms.link_analysis.pagerank_alg.pagerank.html
        G_copy = G_copy.to_undirected()
        
        # Identificar os nós que começam com 'TE_', 'EN_' e 'EM_'
        to_remove = [node for node in G_copy.nodes() if node.startswith('TE_') or node.startswith('EN_') or node.startswith('EM_')]

        # Remover esses nós do grafo
        G_copy.remove_nodes_from(to_remove)
        #print(G_copy.number_of_nodes())
        #print()

        # G.reverse inverte o sentido das setas, direcionando o grafo para os nós de origem (INVERTIDO)
        #pr = sorted(nx.pagerank(G.reverse()).items(), key=lambda x : x[1], reverse=True)[:5]

        pr = sorted(nx.pagerank(G_copy).items(), key=lambda x : x[1], reverse=True)[:5]

        #pr = sorted(nx.pagerank(G_copy).items(), key=lambda x : x[1], reverse=True)[:5]

        # Aponta para os nós com mais setas chegando (TO) - MODO NORMAL
        #pr = sorted(nx.pagerank(G).items(), key=lambda x : x[1], reverse=True)[:5]
    
        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        pr2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(pr):
            if i[0][:3] == 'PJ_':
                key = G.nodes[i[0]]['label'] # Permeneceu G e não G_copy
                value = i[1]
                pr2.insert(n, (key, value))
            else:
                pr2.insert(n, i)
        
        pr2 = [[x[0], round(x[1], 3)] for x in pr2][:3]
        pprint(pr2)

        # Gera relação de nós para exibir na página por meio da tecla u
        pr2_print = [str(pr2.index(x) + 1) + ') ' + str(x[0]) + ': ' + str(x[1]) for x in pr2]
        pr2_print = '\n'.join(pr2_print)
        pr2_print = '\n#####  PAGERANK (ARESTAS BIDIRECIONAIS)  #####\n' + pr2_print

    else:
        print('Grafo não é conectado.')

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()

    

    ######################################################################
    # DETECÇÃO DE COMUNIDADE E ATRIBUIÇÃO EM CADA NÓ COMO ATRIBUTO 'group'
    ######################################################################

    # Aplica algorítimo Louvain Community
    # https://networkx.org/documentation/stable/reference/algorithms/generated/networkx.algorithms.community.louvain.louvain_communities.html#networkx.algorithms.community.louvain.louvain_communities
    
    # Estava dando erro com nó único
    if grau > 1:
    
        # resolution e threshold definem o momento em que o crescimento do grupo gera o nascimento de outro grupo

        #group = nx.algorithms.community.louvain_communities(G2, weight=None, resolution=.7, threshold=0.000001)
        
        #group = nx.algorithms.community.louvain_communities(G, weight=None, resolution=0.01, threshold=0.000000001)

        # Melhor ajuste até o momento. Usando grafo não direcionado (G2)
        group = nx.algorithms.community.louvain_communities(G2, weight=None, resolution=0.1, threshold=0.000000001)
        
        # Algoritmo que não chegou a ser usado
        #from networkx.algorithms.community.centrality import girvan_newman
        #group = girvan_newman(G)
        #pprint(group)

        z_group = []
        for n, i in enumerate(group):
            for k in i:
                z_group.append({k: n + 1})
                
                # Insere índice do grupo no atributo 'z_group'
                G.nodes[k]['z_group'] = n + 1
                
                # Insere total de grupos no atributo 'z_group_total'
                G.nodes[k]['z_group_total'] = len(group)
                

    ##################################################################################################
    # INCLI 'z_nos_novos' EM CADA NÓ PARA DESTACAR OS NÓS ADICIONADOS EM RELAÇÃO À REQUISIÇÃO ANTERIOR
    ##################################################################################################

    if lista_de_nos != None and lista_nos_anterior != None:
        lista_diff = list(set(lista_de_nos) - set(lista_nos_anterior))
        
        if len(lista_diff) > 0:

            for i in lista_de_nos:
            
                if i in lista_diff:
                
                    # Insere no atributo 'z_novo' os números 0 e 1 para indicar se o nó foi ou não adicionado na requisição atual 
                    G.nodes[i]['z_novo'] = 1 # Nó adicionado na requisição atual
                else:
                    G.nodes[i]['z_novo'] = 0





    ######################################################################
    # DOMINANCE 
    ######################################################################

    # NÃO IMPLEMENTADO

    #immediate_dominators = nx.immediate_dominators(G, start=parametro)
    #print('immediate_dominators:')
    #pprint(immediate_dominators)
    #print()

    #dominance_frontiers = nx.dominance_frontiers(G, start=parametro)
    #print('dominance_frontiers:')
    #pprint(dominance_frontiers)
    #print()

    #dominating_set = nx.dominating_set(G, start_with=parametro)
    #print('dominating_set:')
    #pprint(dominating_set)

    #for i in lista_de_nos:
    #    if i in list(dominating_set):

            # ATENÇÃO!
            # A letra z no início do nome do atributo faz o mesmo ser posicionado ao final da lista de atributos do nó,
            # não prejudicando a localização dos primeiros atributos do nó por meio do módulo re.
    #        G.nodes[i]['z_dominating_set'] = 'true'
    #    else:
    #        G.nodes[i]['z_dominating_set'] = 'false'

    #print(len(dominating_set))
    #print()
    


    ###################################################
    # EXCENTRICITY - NÚMERO DE CAMADAS A PARTIR DO ALVO
    ###################################################

    # Verifica se só existe um único parâmetro de busca
    if ';' not in parametro and lista_3 == None and grau > 1:
        excentricidade = nx.eccentricity(G2, parametro)
        print('Excentricidade do grafo:', excentricidade, '(número máximo de camadas a partir do nó-alvo)')
        print()
    else:
        print('Excentricidade do grafo: Não calculada por não ser parâmetro de busca único.')
        print()
        


    ######################################
    # DEGREE CENTRALITY
    ######################################

    # https://coderzcolumn.com/tutorials/data-science/network-analysis-in-python-node-importance-and-paths-networkx#5.4

    print('=' * 56)
    print('DEGREE CENTRALITY:')
    print('=' * 56)
    
    if nx.is_connected(G.to_undirected()):

        t0 = time.time()        

        # DEGREE CENTRALITY NÃO ESCLUI TE, EN E EM, POIS SUA MEDIDA ORIGINAL (INCLUINDO TE, EN E EM) É USADA NO TAMANHO DOS NÓS PARA AUMENTAR AS MASSAS

        # Gera lista de tuplas
        dc = sorted(nx.degree_centrality(G2).items(), key=lambda x : x[1], reverse=True)[:5]
        # Tutorial: https://www.youtube.com/watch?v=gOAG507pi14

        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        dc2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(dc):
            if i[0][:3] == 'PJ_':
                key = G.nodes[i[0]]['label']
                value = i[1]
                dc2.insert(n, (key, value))
            else:
                dc2.insert(n, i)

        dc2 = [[x[0], round(x[1], 3)] for x in dc2][:3]
        pprint(dc2)

        # Gera relação de nós para exibir na página por meio da tecla u
        dc2_print = [str(dc2.index(x) + 1) + ') ' + str(x[0]) + ': ' + str(x[1]) for x in dc2]
        dc2_print = '\n'.join(dc2_print)
        dc2_print = '\n#####  DEGREE CENTRALITY  #####\n' + dc2_print

    
    else:
        print('Grafo não é conectado.')

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()



    ######################################
    # BETWEENNESS CENTRALITY (PADRÃO)
    ######################################

    # https://coderzcolumn.com/tutorials/data-science/network-analysis-in-python-node-importance-and-paths-networkx#5.4

    print('=' * 56)
    print('BETWEENNESS CENTRALITY:')
    print('=' * 56)

    t0 = time.time()
    
    if nx.is_connected(G.to_undirected()):

        # Cria cópia de G para eliminar todos os nós sem relevância (TE, EN e EM)
        G2_copy = G2.copy()
        
        # Identificar os nós que começam com 'TE_', 'EN_' e 'EM_'
        to_remove = [node for node in G2_copy.nodes() if node.startswith('TE_') or node.startswith('EN_') or node.startswith('EM_')]

        # Remover esses nós do grafo
        G2_copy.remove_nodes_from(to_remove)

        # Demora um pouco a processar
        # Gera lista de tuplas
        #bc = sorted(nx.betweenness_centrality(G2, normalized=False).items(), key=lambda x : x[1], reverse=True)[:5]
        # Tutorial: https://www.youtube.com/watch?v=gOAG507pi14

        bc = sorted(nx.betweenness_centrality(G2_copy, normalized=False).items(), key=lambda x : x[1], reverse=True)[:5]
        
        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        bc2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(bc):
            if i[0][:3] == 'PJ_':
                key = G.nodes[i[0]]['label'] # Permaneceu G e não G2_copy
                value = i[1]
                bc2.insert(n, (key, value))
            else:
                bc2.insert(n, i)
        
        bc2 = [[x[0], round(x[1], 3)] for x in bc2][:3]
        pprint(bc2)

        # Gera relação de nós para exibir na página por meio da tecla u
        bc2_print = [str(bc2.index(x) + 1) + ') ' + str(x[0]) + ': ' + str(x[1]) for x in bc2]
        bc2_print = '\n'.join(bc2_print)
        bc2_print = '\n#####  BETWEENNESS CENTRALITY  #####\n' + bc2_print


    else:
        print('Grafo não é conectado.')

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()



    ######################################
    # EIGENVECTOR CENTRALITY
    ######################################

    # https://www.youtube.com/watch?v=teRuQnQ3v7o

    print('=' * 56)
    print('EIGENVECTOR CENTRALITY:')
    print('=' * 56)

    t0 = time.time()
    
    # O grafo tem que ser conectado e ter mais de 1 nó
    if nx.is_connected(G.to_undirected()) and grau > 1:

        # Cria cópia de G para eliminar todos os nós sem relevância (TE, EN e EM)
        G2_copy = G2.copy()
        
        # Identificar os nós que começam com 'TE_', 'EN_' e 'EM_'
        to_remove = [node for node in G2_copy.nodes() if node.startswith('TE_') or node.startswith('EN_') or node.startswith('EM_')]

        # Remover esses nós do grafo
        G2_copy.remove_nodes_from(to_remove)


        # O algoritmo retorna dicionários de nós com o eigenvector centrality como valor
        # O algoritmo não permitiu usar directed graphs (networkx.exception.PowerIterationFailedConvergence) como sugerido aqui:
        # https://networkx.org/documentation/stable/reference/algorithms/generated/networkx.algorithms.centrality.eigenvector_centrality.html#networkx.algorithms.centrality.eigenvector_centrality
        # Gera lista de tuplas
        # Alterados os parâmetros: padrão max_iter=100, tol=1e-06
        #ec = sorted(nx.eigenvector_centrality(G2, max_iter=100, tol=1e-03).items(), key=lambda x : x[1], reverse=True)[:5]

        ec = sorted(nx.eigenvector_centrality(G2_copy, max_iter=100, tol=1e-03).items(), key=lambda x : x[1], reverse=True)[:5]
        
        # Tutorial: https://www.youtube.com/watch?v=gOAG507pi14
        
        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        ec2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(ec):
            if i[0][:3] == 'PJ_':
                key = G.nodes[i[0]]['label']
                value = i[1]
                ec2.insert(n, (key, value))
            else:
                ec2.insert(n, i)

        ec2 = [[x[0], round(x[1], 3)] for x in ec2][:3]
        pprint(ec2)

        # Gera relação de nós para exibir na página por meio da tecla u
        ec2_print = [str(ec2.index(x) + 1) + ') ' + str(x[0]) + ': ' + str(x[1]) for x in ec2]
        ec2_print = '\n'.join(ec2_print)
        ec2_print = '\n#####  EIGENVECTOR CENTRALITY  #####\n' + ec2_print


    else:
        print('Grafo não é conectado.')

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()



    ######################################
    # CLOSENESS CENTRALITY
    ######################################

    # https://www.youtube.com/watch?v=teRuQnQ3v7o

    print('=' * 56)
    print('CLOSENESS CENTRALITY:')
    print('=' * 56)

    t0 = time.time()
    
    if nx.is_connected(G.to_undirected()):

        # Cria cópia de G para eliminar todos os nós sem relevância (TE, EN e EM)
        G2_copy = G2.copy()
        
        # Identificar os nós que começam com 'TE_', 'EN_' e 'EM_'
        to_remove = [node for node in G2_copy.nodes() if node.startswith('TE_') or node.startswith('EN_') or node.startswith('EM_')]

        # Remover esses nós do grafo
        G2_copy.remove_nodes_from(to_remove)


        # O algoritmo retorna dicionários de nós com o closeness centrality como valor

        # Algoritmo permitiu usar directed graph G, que foi revertido (G.reverse()) para priorizar a origem das setas (outward - para fora)
        # Vide https://networkx.org/documentation/stable/reference/algorithms/generated/networkx.algorithms.centrality.closeness_centrality.html#networkx.algorithms.centrality.closeness_centrality
       
        # Gera lista de tuplas
        #cc = sorted(nx.closeness_centrality(G.reverse()).items(), key=lambda x : x[1], reverse=True)[:5]
        #cc = sorted(nx.closeness_centrality(G2).items(), key=lambda x : x[1], reverse=True)[:5]
        # Tutorial: https://www.youtube.com/watch?v=gOAG507pi14

        cc = sorted(nx.closeness_centrality(G2_copy).items(), key=lambda x : x[1], reverse=True)[:5]
        
        # Substitui CNPJ pela razão social da empresa para facilitar leitura
        cc2 = [] # Cria lista apenas para exibição
        for n, i in enumerate(cc):
            if i[0][:3] == 'PJ_':
                key = G.nodes[i[0]]['label']
                value = i[1]
                cc2.insert(n, (key, value))
            else:
                cc2.insert(n, i)

        cc2 = [[x[0], round(x[1], 3)] for x in cc2][:3]
        pprint(cc2)
        
        # Gera relação de nós para exibir na página por meio da tecla u
        cc2_print = [str(cc2.index(x) + 1) + ') ' + str(x[0]) + ': ' + str(x[1]) for x in cc2]
        cc2_print = '\n'.join(cc2_print)
        cc2_print = '\n#####  CLOSENESS CENTRALITY  #####\n' + cc2_print

    else:
        print('Grafo não é conectado.')
        #print()

    t1 = time.time() - t0
    print('Tempo:', round(t1, 2))
    print()



    # REUNIÃO DAS INFORMAÇÕES PARA EXIBIÇÃO NA JANEJA ALERT, NA PÁGINA HTML, QUANTO SE PRESSIONA A TECLA U

    # Foi utilizado a condição 'if' mais rigorosa (Betweenness Centrality)
    if nx.is_connected(G.to_undirected()) and grau > 1:
    
        # INCLUI RESULTADOS DOS ALGORITMOS NO ATRIBUTO 'algoritmo' DOS NÓS PJ (para consulta via Javascript na página HTML)
        # INCLUI NÚMERO DE CAMADAS NO ATRIBUTO 'camadas_solicitadas'
        temp = f"Betweenness Centrality: {str(bc2[0][0])}\nCloseness Centrality:      {str(cc2[0][0])}\nEigenvector Centrality:   {str(ec2[0][0])}\nDegree Centrality:          {str(dc2[0][0])}\nPagerank (Bidirecional): {str(pr2[0][0])}\n{bc2_print}{cc2_print}{ec2_print}{pr2_print}"
        for i in lista_de_nos:
            if i[:3] == 'PJ_':
                G.nodes[i]['algoritmos'] = f'{temp}'
                G.nodes[i]['camadas_solicitadas'] = num_camadas

    else:
        temp = "Não foi possível calcular as medidas de centralidade. Nó único ou grafo não conectado."
        for i in lista_de_nos:
            if i[:3] == 'PJ_':
                G.nodes[i]['algoritmos'] = f'{temp}'
                G.nodes[i]['camadas_solicitadas'] = num_camadas




    # SELECIONA ALGORÍTMO
    # if nx.is_connected(G.to_undirected()): # Necessária inclusão: múltiplos resultados da consulta com a TECLA s apresentava erro (bc not defined)

    #     if algoritmo == 'betweenness':
    #         alg = bc
    #         print('Algoritmo selecionado: Betweenness Centrality')
    #         print()

    #     elif algoritmo == 'closeness':
    #         alg = cc
    #         print('Algoritmo selecionado: Closeness Centrality')
    #         print()

    #     elif algoritmo == 'eigenvector':
    #         alg = ec
    #         print('Algoritmo selecionado: Eigenvector Centrality')
    #         print()
        
    #     else:
    #         alg = bc
    #         print('Algoritmo selecionado: Betweenness Centrality')

    # PAREI AQUI

    ######################################################################################
    # DESCOBRE TODOS OS CAMINHOS MAIS CURTOS ENTRE OS NÓS SELECIONADOS (ALL SHORTEST PATH)
    ######################################################################################

    # Quando os CNPJ são copiados do texto, 'destacar_ligacoes' é sempre 'True'.
    # Quando os itens são copiados do grafo, 'destacar_ligacoes' só é 'True' se forem selecionados apenas dois nós.
    if destacar_ligacoes == True:

        # Gera grafo com ligações sem setas
        G2 = G.to_undirected()

        # Cria lista de parâmetros de pesquisa
        lista = parametro.replace('\n', '').split(';')

        # Cria conjunto (set) vazio
        set_n = set()

        # 
        for x in lista:

            # Exclui prefixo da variável 'x'
            # Para evitar erro quando são colocados juntos 02879926000124 e PJ_02879926000124
            if x[:3] == 'PJ_':
                x = x[3:]

            # Loop sobre nós
            for n in G2.nodes:
                if x in n: # Serve tanto como '==' quanto como 'in'
                    set_n.add(n)
        
        # Converte set em list
        lista_n = list(set_n)

        # Verifica 
        if len(lista_n) >= 2: # <<<<<<<<<<<<<<<<<<<<<<<<<< VÁLIDO SOMENTE PARA DOIS NÓS SELECIONADOS (MUDOU. PASSOU A VALER PARA DOIS OU MAIS NÓS)

            # Gera lista com todas as combinações possíveis de 2 elementos da lista_n
            comb = list(combinations(lista_n, 2))

            count = 0
            for node0, node1 in comb:
                if nx.has_path(G2, node0, node1):
                    
                    # Retorna lista com os nós do caminho mais curto
                    #path2 = nx.shortest_path(G2, node0, node1)

                    # Retorna lista de listas com todos os caminhos mais curtos entre os dois nós
                    path2 = list(nx.all_shortest_paths(G2, node0, node1)) # Demora muito e gera lista de listas (precisa adequar o código)
                    #pprint(path2)

                    # https://seaborn.pydata.org/tutorial/color_palettes.html
                    # https://www.geeksforgeeks.org/seaborn-color-palette/
                    # cores originais = ['#023eff', '#ff7c00', '#1ac938', '#e8000b', '#8b2be2', '#9f4800', '#f14cc1', '#a3a3a3', '#ffc400', '#00d7ff']
                    colors = ['#1ac938', '#8b2be2', '#9f4800', '#f14cc1', '#a3a3a3', '#ffc400', '#00d7ff', '#023eff', '#ff7c00', '#ff0000']

                    # Loop sobre listas de caminhos contidas na lista 'path2'
                    for k in path2:
                    
                        # Loop sobre nós de cada caminho
                        for n, i in enumerate(k):

                            # Variável 'count' usada para permitir que as cores sejam utilizadas de forma cíclica, independentemente do número de ligações
                            if count < 10:
                                x = copy(count)
                            elif count >= 10:
                                x = count % 10

                            # Altera as propriedades 'color' e 'value' do nó
                            try:
                                G.edges[(k[n], k[n + 1])]['color'] = colors[x] #'#FF0000'
                                G.edges[(k[n], k[n + 1])]['value'] = 12
                            except:
                                G.edges[(k[n + 1], k[n])]['color'] = colors[x] #'#FF0000'
                                G.edges[(k[n + 1], k[n])]['value'] = 12
                            
                            # Sai do loop
                            if n == len(k) - 2:
                                break
                        
                        print(f'PATH {count + 1}:')
                        print(k)
                        print()

                        count += 1

                        # Avisa ao usuário a existência de conexão entre os nós (soa apenas uma vez)
                        if count == 1:
                            winsound.Beep(2000, 1000)




    ###################################################################################
    #                                  MÓDULO PYVIS
    ###################################################################################

    net = Network(
                  height= f'{screen_height}px',
                  width='100%', 
                  directed=True, 
                  notebook=False, 
                  bgcolor="#ffffff", 
                  font_color= "#777777", 
                  layout=None,
                  #heading="Fireworks",
                  neighborhood_highlight=False, # Oculta labels dos nós além da 1ª camada
                  #select_menu=True,
                  #filter_menu=True,
                  )

    # Cria grafo Pyvis (Javascript) a partir de grafo criado com Networkx (Python)
    net.from_nx(G)

    # Página HTML/Javascript que toca áudio armazenado no computador: https://collaboratescience.com/so_34243/audio.html
    
    # Configura optções do módulo Vis JS (Javascript) por meio do módulo Pyvis (Python)
    # Usar apenas aspas duplas. Virgula excedente dá erro
    # "enabled": true --> exibe controles na página do grafo
    net.set_options("""var options = {
    "autoResize": true,
    "height": "100%",
    "width": "100%",
    "configure": {
        "enabled": false,
        "showButton": true,
        "filter": true
    },
    "nodes": {
        "color": {
            "border": "transparent",
            "background": "transparent",
            "highlight": {
                "border": "transparent",
                "background": "rgba(0, 255, 0, 0.3)"
            },
            "hover": {
                "border": "transparent",
                "background": "transparent"
            }
        },
        "shapeProperties": {
            "useBorderWithImage": true,
            "interpolation": true
        },
        "scaling": {
            "min": 20,
            "max": 100,
            "label": {
                "enabled": true,
                "min": 15,
                "max": 40,
                "maxVisible": 30,
                "drawThreshold": 8
            }
        },
        "font": {
            "color": "#777777",
            "strokeWidth": 4,
            "strokeColor": "#ffffff",
            "bold": {
                "color": "#ff0000"
            }
        }
    },
    "edges": {
        "hoverWidth":4,
        "smooth": {
            "enabled":false
        },
        "scaling": {
            "label": {
                "enabled": true,
                "min": 30,
                "max": 40,
                "maxVisible": 30,
                "drawThreshold": 8
            }
        },
        "font": {
            "align": "top",
            "color": "#777777",
            "size": 15
        },
        "selectionWidth":4,
        "color": {
            "color": "#2B7CE9",
            "highlight": "#FF0000",
            "inherit":false,
            "hover":"#FF0000"
        },
        "arrowStrikethrough": false
    },
    "interaction": {
        "hover": true,
        "multiselect": true,
        "navigationButtons": true,
		"hideEdgesOnDrag": true,
		"hideEdgesOnZoom": true,
        "tooltipDelay": 1500
    },
    "physics": {
        "enabled": true,
        "barnesHut": {
            "gravitationalConstant": -10000,
            "springLength": 200,
            "avoidOverlap": 0
            }
    },
    "layout": {
        "improvedLayout": false
    }
    }""")

    # Cria variável com endereço do diretório local do arquivo sinarc.py
    diretorio = os.getcwd()
    print('diretório local do arquivo sinarc.py:', diretorio)

    #net.set_template(path_to_template='D:\\PROJETO EDUCAÇÃO MPC\\siade\\dist\\templates\\') # Full os path string value of the template directory
    net.set_template_dir(template_directory=diretorio + '/templates/', template_file='template.html') # Template_directory path: template directory :template_file path: name of the template file that is going to be used to generate the html doc

    # Exibe painéis de ajustes dos parâmetros do grafo
    #net.show_buttons(filter_=['physics', 'layout'])
    
    net.show(name='grafo_inicial.html')



    ###################################################################################
    #                               VIS JS (JAVASCRIPT)
    ###################################################################################

    # https://visjs.org/
    # https://www.hlt.inesc-id.pt/~david/wiki/pt/extensions/vis/docs/network.html


    #########################################################################################################
    # ALTERA CÓDIGO EM JAVASCRIPT NO ARQUIVO HTML - INCLUI FUNDOS COLORIDOS SEGUNDO O COMMUNUNITIES DETECTION
    #########################################################################################################

    #fin = open("grafo_inicial.html", "r")
    #temp = fin.read()
    #fin.close()

    # Paleta de cores alto contrates (5 primeiras cores)
    # https://www.schemecolor.com/high-contrast.php
    #hcc = ["rgba(137, 49, 239, 0.5)",
    #       "rgba(242, 202, 25, 0.5)",
    #       "rgba(255, 0, 189, 0.5)",
    #       "rgba(0, 87, 233, 0.5)",
    #       "rgba(135, 233, 17, 0.5)"
    #      ]

    # Cria dicionário com nós (key) e seus respectivos atributos (values)
    #lista_indices = nx.get_node_attributes(G, 'z_group')

    #for i in lista_de_nos:

        # Calcula índice cíclico para a lista de cores
    #    x = lista_indices[i] % len(hcc)
    #    cor = hcc[x]
        
    #    antes = f'"color": "#97c2fc", "font": {{"color": "#777777"}}, "id": "{i}"'
    #    print(antes)

    #    depois = f'"color": "{cor}", "font": {{"color": "#777777"}}, "id": "{i}"'
    #    print(depois)
        # Substitui padrão '#ffffff" por padrão rgba(255, 255, 255, 1)
    #    temp = temp.replace(antes, depois)

    #fout = open("grafo_final.html", "w")
    #fout.write(temp)
    #fout.close()
    #exit()



    ########################################################
    # DESTACA ALVOS DA REDE INCLUINDO BORDA VERMELHA AOS NÓS
    ########################################################

    # A ESTRATÉGIA UTILIZADA FOI ABRIR O ARQUIVO HTML CRIADO E ALTERAR SEU CÓDIGO COM PRECISÃO
    # 1) Abre arquivo HTML (grafo_inicial.html) com 'open'
    # 2) Localiza trecho com 're.search'
    # 3) Substitui trecho com 'replace' e 're.sub'
    # 4) Salva arquivo HTML (grafo_final.html) com 'open'

    #fin = open(f"{diretorio}grafo_inicial.html", "r")
    fin = open("grafo_inicial.html", "r")
    temp = fin.read()
    fin.close()

    # Não destaca nós no caso de CONSULTA LIVRE na base de dados utilizando caracteres curinga
    #if ' * ' in parametro or '?' in parametro:
    if lista_3 != None:# or inserir_borda_vermelha == False:
        pass
        #for par in lista_3: # par = parâmetro de busca ('PJ_28053619000183')
        #    a_labels = [x[1] for x in list(G.nodes.data('label'))]
        #    for n, i in enumerate(list(G.nodes.data('label'))):
        #        if par.upper() in a_labels[n].upper(): 
        #            try:
        #                antes = re.search(f""".{{51}}id.:..{i[0]}.""", temp).group(0)
        #            except:
        #                antes = re.search(r"""(.color.{45}id.:..[A-Z]{2}_[^,]+),\s.image""", temp).group(1)
        #            #break # Sai do 'loop for' porque achou o nó que deve ser destacado
        #            
        #            # Substitui trecho no arquivo HTML
        #            depois = f'"color": {{"border": "red", "background": "transparent", "hover": {{"background": "transparent", "border": "red"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "red"}}}}, "borderWidth": 4, "shapeProperties": {{"useBorderWithImage":true}}, "borderWidthSelected": 4, "id": "{i[0]}"'
        #            temp = temp.replace(antes, depois)

    else:
        lista = [x.strip() for x in parametro.split(';')] # Lista de parâmetros
        #a_labels = [x[1] for x in list(G.nodes.data('label'))] # Lista de labels dos nós
        for par in lista: # par = parâmetro de busca ('PJ_28053619000183')

            for n, i in enumerate(net.node_ids): # Lista dos ids dos nós (ordem diferente de a_labels)
                if par.upper() in i.upper():# or a_labels[n].upper() in par.upper(): 
                    try:
                        i2 = i.replace(r'*', r'\*') # Usado para regex não usar * como um comando
                        antes = re.search(rf'.{{56}}"{i2}"', temp).group(0) # FUNCIONANDO!!!!!
                        
                        #antes = re.search(fr'(.color(.*?).id.:."{i2}"),..image.:', temp).group(1) # EM TESTE!!!!!
                        
                        #print('ANTES:', antes)
                    except Exception as e:
                        print(e)
                        
                    break # Sai do 'loop for' porque achou o nó que deve ser destacado
                    
            # Substitui trecho no arquivo HTML
            depois = f'"color": {{"border": "red", "background": "transparent", "hover": {{"background": "transparent", "border": "red"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "red"}}}}, "borderWidth": 4, "shapeProperties": {{"useBorderWithImage":true}}, "borderWidthSelected": 4, "id": "{i}"'
            #print('DEPOIS:', depois)
            temp = temp.replace(antes, depois)

    #fout = open(f"{diretorio}grafo_final.html", "w")
    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()

   

    ###################################################################
    # DESTACA COM TRACEJAMENTO A BORDA DOS NÓS COM O MAIOR CENTRALIDADE
    ###################################################################

    # https://jsfiddle.net/api/post/library/pure/

    # Verifica se G é um grafo conectado. Se não for, não faz sentido falar em centralidade
    # Acrescentada a condição 'grau > 1', pois resultado com nó único estava gerando erro.
    # Essas duas condições são as mesmas exigidas pele algorítmo Betweenness Centrality
    if nx.is_connected(G.to_undirected()) and grau > 1:

        # Extrai o primeiro nó do ranking gerado pelo algorítmo Betweenness Centrality
        #nos_centrais = alg[:1]

        # Pode ser usado para qualquer número de algorítmos!  Todos serão destacados no grafo
        temp = [bc[0], ec[0], cc[0], pr[0]] # Degree Centrality não precisa, pois é usado no tamanho dos nós

        # Elimina nós de centralidade duplicados para evitar erro na localização do trecho do código HTML referente a nó que já foi alterado
        nos_centrais = []
        temp2 = []
        for i in temp:
            if i[0] not in temp2:
                temp2.append(i[0])
                nos_centrais.append(i)

        # Extrai código do arquivo HTML
        fin = open("grafo_final.html", "r")
        temp = fin.read()
        fin.close()

        for id_no_central in nos_centrais:

            id_no_central = id_no_central[0]
            
            # Se o primeiro nó do ranking não estiver entre os alvos, faz borda verde tracejada
            if id_no_central not in lista:

                # Inclui '\' antes do * para evitar erro na utilização do módulo regex
                id_no_central = id_no_central.replace(r'*', r'\*')

                antes = re.search(rf'.{{56}}"{id_no_central}"', temp).group(0)
                #print('ANTES:', antes) # NÃO APAGAR, POIS É USADO PARA IDENTIFICAR ERROS NA SUBSTITUIÇÃO DE CÓDIGOS NO ARQUIVO HTML

                # Desfaz inclusão de '\'
                id_no_central = id_no_central.replace(r'\*', r'*')

                # Substitui trecho no arquivo HTML (azul das setas: #2B7CE9, cinza: #777777)
                depois = f'"color": {{"border": "#aaaaaa", "background": "transparent", "hover": {{"background": "transparent", "border": "#aaaaaa"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "#aaaaaa"}}}}, "borderWidth": 4, "shapeProperties": {{"useBorderWithImage":true, borderDashes: [15, 5]}}, "borderWidthSelected": 4, "id": "{id_no_central}"'
                #print('DEPOIS:', depois)
                #print()

                temp = temp.replace(antes, depois)

            # o primeiro nó do ranking estiver entre os alvos, faz borda vermelha tracejada
            else:
                for i in lista:

                    if id_no_central == i:

                        #id_no_central = id_no_central.replace(r'*', r'\*')
                        antes = f'"color": {{"border": "red", "background": "transparent", "hover": {{"background": "transparent", "border": "red"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "red"}}}}, "borderWidth": 4, "shapeProperties": {{"useBorderWithImage":true}}, "borderWidthSelected": 4, "id": "{i}"'
                        #print('ANTES:', antes)

                        # Substitui trecho no arquivo HTML (inclui borderDashes: [15, 5])
                        depois = f'"color": {{"border": "red", "background": "transparent", "hover": {{"background": "transparent", "border": "red"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "red"}}}}, "borderWidth": 4, "shapeProperties": {{"useBorderWithImage":true, borderDashes: [15, 5]}}, "borderWidthSelected": 4, "id": "{i}"'
                        #print('DEPOIS:', depois)
                        #print()

                        temp = temp.replace(antes, depois)

        # Atualiza código do arquivo HTML
        fout = open("grafo_final.html", "w")
        fout.write(temp)
        fout.close()


    
    #########################################
    # INCLUI FUNDO VERDE NOS NÓS SELECIONADOS
    #########################################

    #fin = open(f"{diretorio}grafo_final.html", "r")
    fin = open("grafo_final.html", "r")
    temp = fin.read()
    fin.close()

    temp = temp.replace('"color": "#97c2fc", ', '')

    # lista = [x.strip() for x in parametro.split(';')]
    # for par in lista: # par = parâmetro de busca ('PJ_28053619000183')
 
    #     for i in net.node_ids: # Loop sobre lista de todos os nós
    #         if par not in i:
    #             try:
    #                 antes = re.search(fr""".{{51}}id.:..{i}.""", temp).group(0)
    #             except:
    #                 antes = re.search(r"""(.color.{42}...id.:..[A-Z]{2}_[^,]+),\s.image""", temp).group(1)
    #             depois = f'"color": {{"border": "transparent", "background": "transparent", "hover": {{"background": "transparent", "border": "transparent"}}, "highlight": {{"background": "rgba(0, 255, 0, 0.3)", "border": "transparent"}}}}, "borderWidth": 0, "shapeProperties": {{"useBorderWithImage":true}}, "borderWidthSelected": 0, "id": "{i}"'
    #             temp = temp.replace(antes, depois)
 
    #fout = open(f"{diretorio}grafo_final.html", "w")
    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()



    #####################################################################################
    # INCLUI URL DE IMAGENS PARA PERMITIR A PORTABILIDADE DOS ARQUIVOS HTML (BROKENIMAGE)
    #####################################################################################

    # ATENÇÃO!
    # A portabilidade dos arquivos HTML implica perda da indicação gráfica da existência de dívidas com a União (bandeira vermelha sobreposta)
    # As informações sobre as dívidas permanecem disponíveis no popup de cada nó.
    # A bandeira vermelha sobreposta só aparece se o usuário estiver executando o arquivo executável do Sistema Sinarc.

    fin = open("grafo_final.html", "r")
    temp = fin.read()
    fin.close()

    # Inclui atributo 'brokenImage' a todos os nós do grafo para permitir que qualquer pessoa possa abrir os arquivos HTML do grafo.
    # Utiliza-se a mesma imagem (sem bandeira) para figuras com e sem bandeira (PF_, PJ_ e PE_)
    temp = temp.replace('"image": "pf_homem.png"', '"image": "pf_homem.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/3220/3220358.png"')
    temp = temp.replace('"image": "pf_homem_com_bandeira.png"', '"image": "pf_homem_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/3220/3220358.png"')

    temp = temp.replace('"image": "pf_mulher.png"', '"image": "pf_mulher.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/3220/3220315.png"')
    temp = temp.replace('"image": "pf_mulher_com_bandeira.png"', '"image": "pf_mulher_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/3220/3220315.png"')
    
    temp = temp.replace('"image": "pj_brasil_ativa.png"', '"image": "pj_brasil_ativa.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/1684/1684121.png"')
    temp = temp.replace('"image": "pj_brasil_ativa_com_bandeira.png"', '"image": "pj_brasil_ativa_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/1684/1684121.png"')

    temp = temp.replace('"image": "pj_brasil_inativa.png"', '"image": "pj_brasil_inativa.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/1684/1684019.png"')
    temp = temp.replace('"image": "pj_brasil_inativa_com_bandeira.png"', '"image": "pj_brasil_inativa_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/1684/1684019.png"')

    temp = temp.replace('"image": "pj_exterior_ativa.png"', '"image": "pj_exterior_ativa.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/4402/4402353.png"')
    temp = temp.replace('"image": "pj_exterior_ativa_com_bandeira.png"', '"image": "pj_exterior_ativa_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/4402/4402353.png"')

    temp = temp.replace('"image": "pj_exterior_inativa.png"', '"image": "pj_exterior_inativa.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/4400/4400008.png"')
    temp = temp.replace('"image": "pj_exterior_inativa_com_bandeira.png"', '"image": "pj_exterior_inativa_com_bandeira.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/4400/4400008.png"')

    temp = temp.replace('"image": "endereco.png"', '"image": "endereco.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/83/83909.png"')
    temp = temp.replace('"image": "email.png"', '"image": "email.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/3059/3059410.png"')
    temp = temp.replace('"image": "telefone.png"', '"image": "telefone.png", "brokenImage": "https://cdn-icons-png.flaticon.com/512/159/159832.png"')

    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()



    #########################################################################################################
    # ALTERA CÓDIGO EM JAVASCRIPT NO ARQUIVO HTML - INCLUI FUNDOS COLORIDOS SEGUNDO O COMMUNUNITIES DETECTION
    #########################################################################################################

    #fin = open("grafo_final.html", "r")
    #temp = fin.read()
    #fin.close()

    # Paleta de cores alto contrates (5 primeiras cores)
    # https://www.schemecolor.com/high-contrast.php
    #hcc = ["rgba(137, 49, 239, 0.5)",
    #       "rgba(242, 202, 25, 0.5)",
    #       "rgba(255, 0, 189, 0.5)",
    #       "rgba(0, 87, 233, 0.5)",
    #       "rgba(135, 233, 17, 0.5)"
    #      ]

    #temp = temp.replace("#8931EF", "rgba(137, 49, 239, 0.5)")
    #temp = temp.replace("#F2CA19", "rgba(242, 202, 25, 0.5)")
    #temp = temp.replace("#FF00BD",  "rgba(255, 0, 189, 0.5)")
    #temp = temp.replace("#0057E9", "rgba(0, 87, 233, 0.5)")
    #temp = temp.replace("#87E911", "rgba(135, 233, 17, 0.5)")
    
    
    # Cria dicionário com nós (key) e seus respectivos atributos (values)
    #lista_indices = nx.get_node_attributes(G, 'z_group')

    #for i in lista_de_nos:

        # Calcula índice cíclico para a lista de cores
    #    x = lista_indices[i] % len(hcc)
    #    cor = hcc[x]
        
    #    antes = f'"color": "#97c2fc", "font": {{"color": "#777777"}}, "id": "{i}"'
    #    print(antes)

    #    depois = f'"color": "{cor}", "font": {{"color": "#777777"}}, "id": "{i}"'
    #    print(depois)
        # Substitui padrão '#ffffff" por padrão rgba(255, 255, 255, 1)
    #    temp = temp.replace(antes, depois)

    #fout = open("grafo_final.html", "w")
    #fout.write(temp)
    #fout.close()



    ##################################
    # SUBSTITUI CHARSET DA PÁGINA HTML
    ##################################

    fin = open("grafo_final.html", "r")
    temp = fin.read()
    fin.close()

    temp = temp.replace('<meta charset="utf-8">', '<meta charset="ISO-8859-1">')

    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()



    ############################
    # SUBSTITUI SRC DE SCRIPT JS
    ############################

    fin = open("grafo_final.html", "r")
    temp = fin.read()
    fin.close()

    temp = temp.replace('<script src="lib/bindings/utils.js"></script>', '<script src="./pyvis/templates/lib/bindings/utils.js"></script>')

    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()



    #############################################
    # ALTERA CÓDIGO EM JAVASCRIPT NO ARQUIVO HTML
    #############################################

    # Extrai código do arquivo HTML
    fin = open("grafo_final.html", "r")
    temp = fin.read()
    fin.close()

    # Cria variável contendo todo o código Javascript a ser inserido no arquivo HTML
    code = """

        // ####################################
        // CRIA FUNÇÃO DE MENSAGENS TEMPORÁRIAS
        // ####################################

        // As mensagens são exibidas na tela pelo tempo definido no parâmetro da função
        // 178, 255, 178, 1

        var timerId;
        function tempAlert(msg, duration) {

            let divRemover = document.getElementById("el-id");

            // verifica se a div existe
            if (divRemover) {

                // Cancela o temporizador anterior
                clearTimeout(timerId);
                
                // Remova a div usando o método remove()
                divRemover.remove();
            }


            var el = document.createElement("div");
            el.id = "el-id";
            el.setAttribute("style","position:absolute;top:30px;right:21px;border:5px solid rgba(245, 245, 245, 1);border-radius: 5px;background-color:rgba(245, 245, 245, 1);height: auto;");
            el.innerHTML = msg;
            timerId = setTimeout(function(){
                el.parentNode.removeChild(el);
            }, duration);
            document.body.appendChild(el);
        }



        // #################################################
        // CRIA FUNÇÃO QUE AVISA QUANDO CAPSLOCK FOI ATIVADA
        // #################################################

        // EXIBE AVISO DE CAPSLOCK ATIVADA
        function capsLockAtivada(msg, duration) {
            var el = document.createElement("div");
            el.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
            el.innerHTML = msg;
            setTimeout(function(){
            el.parentNode.removeChild(el);
            },duration);
            document.body.appendChild(el);
        }

        // EXIBE AVISO DE CAPSLOCK DESATIVADA
        function capsLockDesativada(msg, duration) {
            var el = document.createElement("div");
            el.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(178, 255, 178, 1);border-radius: 5px;background-color:rgba(178, 255, 178, 1);");
            el.innerHTML = msg;
            setTimeout(function(){
            el.parentNode.removeChild(el);
            },duration);
            document.body.appendChild(el);
        }



        var el_2 = ''
        var toggle_caps = 0;
        document.addEventListener('keydown', (event) => {

            if(event.code === "CapsLock"){

                let isCapsLockOn = event.getModifierState("CapsLock");

                if(isCapsLockOn) {

                    el_2 = document.createElement("div");
                    el_2.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                    el_2.innerHTML = "Tecla CapsLock Ativada";
                    document.body.appendChild(el_2);
                    //console.log("Caps Lock turned on");

                    toggle_caps = 1;

                } else if (!isCapsLockOn && toggle_caps == 1) {

                    el_2.style.display = "none";
                    capsLockDesativada("Tecla CapsLock Desativada", 3000);
                    //console.log("Caps Lock turned off");

                }
            }
        });


        // EXIBE AVISO NO CARREGAMENTO DA PÁGINA DE QUE A TECLA CAPSLOCK ESTÁ ATIVADA
        //document.addEventListener("DOMContentLoaded", function(event) {
        //    if (event.getModifierState("CapsLock")) {
        //        //el_2 = document.createElement("div");
        //        //el_2.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
        //        //el_2.innerHTML = "Tecla CapsLock Ativada";
        //        //document.body.appendChild(el_2);
        //    }
        //});



        //setInterval(function() {
        //    var capsLockOn = navigator.getKeyboardState ? navigator.getKeyboardState()[0].shiftKey : false; // verifica se Caps Lock está ativado
            
        //    if (capsLockOn) {
        //        var div = document.createElement('div'); // cria um novo elemento de div
        //        div.textContent = 'A tecla Caps Lock está ativada!'; // define o conteúdo da div
        //        div.style.color = 'red'; // define a cor da fonte da div
        //        div.style.fontSize = '20px'; // define o tamanho da fonte da div
        //        div.style.fontWeight = 'bold'; // define o peso da fonte da div
        //        div.style.marginTop = '10px'; // define a margem superior da div
                
        //        var container = document.getElementById('visualization'); // obtém o elemento do contêiner da biblioteca vis.js
        //        container.appendChild(div); // adiciona a div ao contêiner
        //        console.log('oi')
        //    }
        //}, 2000); // verifica o estado da tecla Caps Lock a cada segundo




        // ##########################
        // ACUMULA NÓS - TECLAS + E -
        // ##########################

        // OS ITENS DO ARRAY SALVO NO 'localStorage' SÃO ABERTOS PELA TECLA 'y'

        // localStorage: https://codetheweb.blog/javascript-localstorage/
        
        // CRIA ARRAY VAZIO SE ELE AINDA NÃO EXISTIR E SALVA EM 'localStorage'
        if (localStorage.getItem("nos_acum") == null) {
            var empty_array = new Array();
            localStorage.setItem('nos_acum', JSON.stringify(empty_array));
        }
        
        document.addEventListener('keydown', (event) => {

                // INCLUI NÓS SELECIONADOS NO ARRAY 'nos_acum' E SALVA EESSE ARRAY NO 'localStorage'
                if (event.key == '+') {
                    
                    let s = network.getSelectedNodes();

                    let nos_acum = []

                    if (s.length > 0) {
                    
                        for (let i=0; i<s.length; i++) {
                            nos_acum.push(s[i]);
                        }

                        // RESTAURA O ARRAY
                        let temp = JSON.parse(localStorage.getItem("nos_acum"));

                        // ADICIONA ITENS AO ARRAY
                        temp = temp.concat(nos_acum)

                        // ELIMINA DUPLICIDADES DO ARRAY
                        temp = [...new Set(temp)];

                        // GRAVA NOVO ARRAY NO 'localStorage' (SUBSTITUI O ANTERIOR)
                        localStorage.setItem('nos_acum', JSON.stringify(temp));

                        // EXIBE NOVO ARRAY
                        tempAlert(`Tecla (${event.key}) - Adiciona item à lista de nós-alvos (${temp.length}): ${temp.join(' ||| ')}`, 3000);
                    
                    
                    // EXIBE ARRAY SE NENHUM NÓ ESTIVER SELECIONADO
                    } else {

                        // RESTAURA O ARRAY
                        let temp = JSON.parse(localStorage.getItem("nos_acum"));

                        tempAlert(`Tecla (${event.key}) - Lista de nós-alvos (${temp.length}): ${temp.join(' ||| ')}`, 3000);
                    }
                    
                // EXCLUI ÚLTIMO ITEM DO ARRAY 'nos_acum' SALVO NO 'localStorage'
                } else if (event.key == '-') {

                    // RESTAURA O ARRAY
                    let temp = JSON.parse(localStorage.getItem("nos_acum"));

                    // EXCLUI ÚLTIMO ITEM DO ARRAY
                    temp.pop();

                    // GRAVA NOVO ARRAY NO 'localStorage' (SUBSTITUI O ANTERIOR)
                    localStorage.setItem('nos_acum', JSON.stringify(temp));

                    tempAlert(`Tecla (${event.key}) - Remove último item da lista de nós-alvos (${temp.length}): ${temp.join(' ||| ')}`, 3000);
                }            
        });



        // ###########################################
        // ATIVA E DESATIVA O MODO DE CAPTURA - TECLA |
        // ###########################################

        var toggle_ponto_e_virgula = true

        document.addEventListener('keydown', (event) => {

            if (event.key == '|' && toggle_ponto_e_virgula == true) {

                async function capturaDesativa(param) {
                    await navigator.clipboard.writeText(param)
                }

                let no_unico = network.getSelectedNodes();
                let idno = ''
                if (no_unico.length == 1) {
                    idno = no_unico[0]
                }

                let temp = "AAAdesativaAAA" + idno;
                
                setTimeout(() => {capturaDesativa(temp)}, 0)

                toggle_ponto_e_virgula = false

                el_1 = document.createElement("div");
                el_1.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                el_1.innerHTML = `Tecla ${event.key} - Modo de captura desativado`;
                document.body.appendChild(el_1);
            
            } else if (event.key == '|' && toggle_ponto_e_virgula == false) {

                async function capturaAtiva(param) {
                    await navigator.clipboard.writeText(param)
                }

                let temp = "AAAativaAAA";
                
                setTimeout(() => {capturaAtiva(temp)}, 0)

                el = document.createElement("div");
                el.setAttribute("style","position:absolute;top:30px;left:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                el.innerHTML = `Tecla ${event.key} - Modo de captura ativado`;

                toggle_ponto_e_virgula = true

                el_1.style.display = "none";

                // Exibe informações na tela
                capsLockDesativada(`Tecla ${event.key} - Modo de captura ativado`, 3000);
            }
        });



        // ########################################################
        // ABRE PÁGINA COM PESQUISA AVANÇADA - TECLA 1 A 9
        // ########################################################

        document.addEventListener('keydown', (event) => {

            // TECLA 1 - VALIDAR DADOS DO CNPJ NO SITE DA RECEITA FEDERAL
            if (event.key == '1') {
                let sel = network.getSelectedNodes();
                if (sel.length == 1 && sel[0].slice(0, 3) == 'PJ_') {
                    let cnpj = sel[0].slice(3);
                    let link_consulta_cnpj = "https://solucoes.receita.fazenda.gov.br/servicos/cnpjreva/Cnpjreva_Solicitacao.asp?cnpj=" + cnpj;
                    window.open(link_consulta_cnpj, target="_blank", windowFeature='left=50,top=100,width=800,height=600');
                    
                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Validar CNPJ no site da Receita Federal`, 3000);
                }
            }



            // TECLA 2 - VALIDAR CEIS, CEPIM, CNEP e Acordo de Leniência
            if (event.key == '2') {

                let sel = network.getSelectedNodes();
                if (sel.length == 1 && sel[0].slice(0, 3) == 'PJ_') {
                    let cnpj = sel[0].slice(3);
                    let link_consulta_cnpj = `https://portaldatransparencia.gov.br/sancoes/consulta?paginacaoSimples=true&tamanhoPagina=&offset=&direcaoOrdenacao=asc&cpfCnpj=${cnpj}&colunasSelecionadas=linkDetalhamento%2Ccadastro%2CcpfCnpj%2CnomeSancionado%2CufSancionado%2Corgao%2CcategoriaSancao%2CdataPublicacao%2CvalorMulta%2Cquantidade`;
                    window.open(link_consulta_cnpj, target="_blank", windowFeature='left=50,top=100,width=800,height=600');
                    
                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Validar CNPJ no site da Receita Federal`, 3000);
                }
            }



            // TECLA 3 - VALIDAR DÍVIDAS COM A UNIÃO
            // Ainda não implementado



            // TECLA 4 - ABRE ARQUIVOS EXCEL COM NÓS E ARESTAS
            if (event.key == '4') {
                
                async function copyOperation(param) {
                    await navigator.clipboard.writeText(param)
                };

                setTimeout(() => {copyOperation('***ABRE-EXCEL***')}, 0)

                tempAlert(`Tecla ${event.key} - Abre arquivos Excel de nós e de arestas`, 3000);
            }



            // TECLA 5 - ABRE SITE REDE CNPJ
            if (event.key == "5") {

                window.open("https://www.redecnpj.com.br/rede/", "_blank", "left=50,top=100,width=600,height=600");

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Abre site Rede CNPJ`, 3000);

            }
            


            // TECLA 6 - EXIBE PÁGINA DA RECEITA FEDERAL E DA PGFN CONTENDO BASES DE DADOS
            if (event.key == '6') {

                let link_pep = "https://portaldatransparencia.gov.br/download-de-dados/pep";
                window.open(url=link_pep, target="_blank", windowFeature='left=200,top=140,width=600,height=500');
                
                let link_cnpj = "https://dados.gov.br/dados/conjuntos-dados/cadastro-nacional-da-pessoa-juridica---cnpj";
                window.open(url=link_cnpj, target="_blank", windowFeature='left=150,top=110,width=600,height=500');

                let link_cgu = "https://portaldatransparencia.gov.br/sancoes";
                window.open(url=link_cgu, target="_blank", windowFeature='left=100,top=80,width=600,height=500');

                let link_pgfn = "https://www.gov.br/pgfn/pt-br/assuntos/divida-ativa-da-uniao/transparencia-fiscal-1/copy_of_dados-abertos"
                window.open(url=link_pgfn, target="_blank", windowFeature='left=50,top=50,width=600,height=500');

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Bases de dados da Receita Federal e da PGFN`, 3000);
            }



            // TECLA 7 - EXIBE ARQUIVO JSON DA EMPRESA
            if (event.key == '7') {
                let sel = network.getSelectedNodes();
                if (sel.length == 1) {
                    if (sel[0].slice(0, 3) == 'PJ_') {
                        let parametro = sel[0].slice(3);
                        let link = "http://jsonviewer.stack.hu/#http://publica.cnpj.ws/cnpj/" + parametro;
                        window.open(link, target="_blank", windowFeature='left=50,top=100,width=600,height=600');
                    }

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Dados da empresa no formato JSON`, 3000);

                }
            }



            // TECLA 8 - PESQUISA POR FORNECEDORES PJ - UNIÃO
            if (event.key == '8') {

                let sel = network.getSelectedNodes();

                if (sel.length == 1) {

                    // PESSOA JURÍDICA
                    if (sel[0].slice(0, 3) == 'PJ_') {
                        let parametro = nodes.get(sel[0]);
                        parametro = parametro.id.slice(3)

                        let link_fornecedores_uniao = "https://compras.dados.gov.br/fornecedores/doc/fornecedor_pj/" + parametro;
                        window.open(url=link_fornecedores_uniao, target="_blank", windowFeature='left=50,top=50,width=600,height=500');
                    }

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Pesquisa nó selecionado entre fornecedores da União`, 3000);
                }
            }



            // TECLA 9 - DISPONÍVEL



            //  TECLA 0 - NÃO ATRIBUIR ATALHO PARA A TECLA 0, POIS IMPEDE AJUSTAR O ZOOM DO NAVEGADOR AO VALOR PADRÃO (100%) POR MEIO DE CTRL + 0

        });



        // ################################
        // SELECIONA TODOS OS NÓS - TECLA A
        // ################################

        document.addEventListener('keydown', (event) => {
            if (event.key == "A") {
                network.selectNodes(nodes.getIds());
                
                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Seleciona todos os nós (${network.getSelectedNodes().length})`, 3000);
            }

        })



        // ###############################################################################################
        // SELECIONA NÓS ADJACENTES AOS NÓS SELECIONADOS DE FORMA ACUMULATIVA, CAMADA POR CAMADA - TECLA a
        // ###############################################################################################

        var sel_anterior = [];
        document.addEventListener('keydown', (event) => {

            if (event.key == "a" && network.getSelectedNodes().length > 0) {

                // Desseleciona todos os nós quando a seleção alcança todos os nós
                if (network.getSelectedNodes().length > 0 && network.getSelectedNodes().length == nodes.getIds().length) {
                    tempAlert(`Tecla ${event.key} - Desseleciona todos os nós (${nodes.getIds().length})`, 3000);
                }

                // Exibe mensagem informativa
                //if (network.getSelectedNodes().length > 0 && network.getSelectedNodes().length == nodes.getIds().length) {
                //    tempAlert(`Tecla ${event.key} - Desseleciona todos os nós (${nodes.getIds().length})`, 3000);
                //}


                // DESATIVA A FÍSICA NO INÍCIO DA SELEÇÃO
                //network.physics.physicsEnabled = false;
                //network.physics.stopSimulation();

                // VERIFICA SE TODOS OS NÓS FORAM SELECIONADOS
                if ([...new Set(sel_anterior)].length == nodes.getIds().length) {
                    network.unselectAll();
                    sel_anterior = [];
                }

                let sel = network.getSelectedNodes();

                // IMPEDE QUE DOIS OU MAIS NÓS SEJAM SELECIONADOS NO INÍCIO DO SCRIPT
                if (sel.length > 1 && sel_anterior.length == 0) {
                    for (let i=0; i<sel.length; i++) {
                        nodes.update({id: sel[i], font: {color: '#777777'}})
                    }
                    network.unselectAll();
                    network.physics.physicsEnabled = true;
                    network.physics.startSimulation();
                    return;
                }

                for (let i=0; i<sel.length; i++) {
                    if (!(sel[i] in sel_anterior)) {
                        let temp_a = network.getConnectedNodes(sel[i]);
                        sel_anterior = sel_anterior.concat(temp_a);
                    }
                }
                
                if (sel.length == 1) {
                    sel_anterior = sel_anterior.concat(sel);
                    network.selectNodes(sel_anterior);
                    nodes.update({id: sel[0], font: {color: '#777777'}});
                } else {
                    network.selectNodes(sel_anterior);
                }

                // REATIVA A FÍSICA APÓS A ÚLTIMA CAMADA SER SELECIONADA E TODA A REDEE SER DESSELECIONADA
                //if (network.getSelectedNodes().length == 0) {
                //    network.physics.physicsEnabled = true;
                //    network.physics.startSimulation();
                //}


                if (network.getSelectedNodes().length > 0) {
                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes (${network.getSelectedNodes().length}/${nodes.getIds().length})`, 3000);
                }
            
            } else if (event.key == "a" && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um nó para selecionar seus nós adjacentes`, 3000);
            }
        });
        


        // ######################################################
        // AJUSTE AUTOMÁTICO DA ALTURA DA ÁREA DO GRAFO - TECLA b
        // ######################################################
        
        document.addEventListener('keydown', (event) => {
            if (event.key == 'b') {
                    // EXTRAI DIMENSÕES DA ÁREA DO GRAFO
                    //let box = document.querySelector('#mynetwork');
                    //let width = box.offsetWidth;
                    //let height = box.offsetHeight;
                    
                    // EXTRAI DIMENSÕES DA JANELA DO NAVEGADOR
                    //var w = window.innerWidth;
                    var h = window.innerHeight;
                    
                    // AJUSTA ALTURA DA ÁREA DO GRAFO À ALTURA DA JANELA
                    document.querySelector('#mynetwork').style.height = h - 20;

                    // ENQUADRA GRAFO. FUNCIONA APENAS APÓS A ALTURA TER SIDO AJUSTADA ()
                    //network.fit();

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Ajusta altura da janela de visualização`, 3000);

            }
        });


        // ##############################################
        // CENTRALIZA NÓS SELECIONADOS - TECLA c (CENTER)
        // ##############################################

        var toggle_c = true;

        document.addEventListener('keydown', (event) => {
            
            // SÓ FUNCIONA NO LAYOUT BARNES-HUT E COM APENAS UM NÓ SELECIONADO
            if (event.key == 'c' && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length == 1 && toggle_c == true) {

                // DESABILITA ELEMENTOS PARA EVITAR TRAVAMENTO DA IMAGEM (MELHORA DESEMPENHO)
                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();

                // PARTE PRINCIPAL
                //network.fit({nodes: con_nodes, maxZoomLevel: 1, animation: false});

                let ss = network.getSelectedNodes();

                // PARTE PRINCIPAL
                for (let i=0; i<ss.length; i++) {
                    p = network.getPosition(ss[i]);
                    network.moveTo({
                        position: {x: p['x'], y: p['y']},
                        scale: 1,
                        offset: {x:0, y:0}
                    });
                }
            
                toggle_c = false;

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Centraliza nó selecionado`, 3000);

            // EXIBE NÓS CONECTADOS AOS NÓS SELECIONADOS
            } else if (event.key == 'c' && network.physics.options.solver == 'barnesHut') {

                // DESABILITA ELEMENTOS PARA EVITAR TRAVAMENTO DA IMAGEM (MELHORA DESEMPENHO)
                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();

                let sel_nodes = network.getSelectedNodes();

                // EXTRAI NOS CONECTADOS AO NÓ SELECIONADO
                let con_nodes = []
                let temp = []
                if (sel_nodes.length > 0) {
                    for (var i=0; i<sel_nodes.length; i++) {
                        temp = network.getConnectedNodes(sel_nodes[i]);
                        temp2 = temp.concat(sel_nodes);
                        con_nodes = con_nodes.concat(temp2)
                    }
                }

//                if (network.getSelectedNodes().length == 1 && toggle_c == false) {

//                    // PARTE PRINCIPAL
//                    network.fit({nodes: con_nodes, maxZoomLevel: 1, animation: false});
//                    toggle_c = true
                
//                } else if (network.getSelectedNodes().length > 1) {

                    // PARTE PRINCIPAL
//                    network.fit({nodes: con_nodes, maxZoomLevel: 1, animation: false});
//                    toggle_c = true
//                }

                // PARTE PRINCIPAL
                network.fit({nodes: con_nodes, maxZoomLevel: 1, animation: false});
                toggle_c = true

                // Exibe informações na tela
                if (network.getSelectedNodes().length == 0) {
                    tempAlert(`Tecla ${event.key} - Enquadra todos os nós`, 3000);
                } else {
                    tempAlert(`Tecla ${event.key} - Enquadra nós adjacentes aos nós selecionados`, 3000);
                }
            }
        });



        // #####################################################################
        // SELECIONA NÓS ÚNICOS (EN, EM E TE) PARA SIMPLIFICAR O GRAFO - TECLA ç
        // #####################################################################

        var toggle_ç = 0;
        document.addEventListener("keypress", function(event) {
            if (event.key === "ç" && network.getSelectedNodes().length == 0 && toggle_ç == 0) {
                let nod = nodes.getIds();
                let nodToSelect = [];
                for (var i = 0; i < nod.length; i++) {
                    if ((nod[i].slice(0, 3) == "TE_") || (nod[i].slice(0, 3) == "EN_") || (nod[i].slice(0, 3) == "EM_")) {
                        if (network.body.nodes[nod[i]].edges.length == 1) {
                            nodToSelect.push(nod[i]);
                            //console.log(nod[i]);
                        }
                    }
                }
                network.selectNodes(nodToSelect);

                let percentual = (nodToSelect.length / nod.length) * 100;
                percentual = percentual.toFixed(1);

                tempAlert(`Tecla ${event.key} - Seleciona nós EN, EM e TE com aresta única (${nodToSelect.length}) ${percentual}%`, 3000);

                toggle_ç = 1;
            
            } else if (event.key === "ç" && toggle_ç == 1) {
                let nod = nodes.getIds();
                let nodToSelect = [];
                for (var i = 0; i < nod.length; i++) {
                    if ((nod[i].slice(0, 3) == "TE_") || (nod[i].slice(0, 3) == "EN_") || (nod[i].slice(0, 3) == "EM_")) {
                        nodToSelect.push(nod[i]);
                    }
                }
                network.selectNodes(nodToSelect);

                let percentual = (nodToSelect.length / nod.length) * 100;
                percentual = percentual.toFixed(1);

                tempAlert(`Tecla ${event.key} - Seleciona todos os nós EN, EM e TE (${nodToSelect.length}) ${percentual}%`, 3000);

                toggle_ç = 2;



            //} else if (event.key === "ç" && toggle_ç == 2) {

            //    let nod = nodes.getIds();
            //    let nodToSelect = [];
            //    for (var i = 0; i < nod.length; i++) {
            //        if (network.body.nodes[nod[i]].edges.length == 1) {
            //            nodToSelect.push(nod[i]);
            //        }
            //    }
            //    network.selectNodes(nodToSelect);

            //    let percentual = (nodToSelect.length / nod.length) * 100;
            //    percentual = percentual.toFixed(1);

            //    tempAlert(`Tecla ${event.key} - Seleciona todos os nós com apenas uma aresta (${nodToSelect.length}) ${percentual}%`, 3000);

            //    toggle_ç = 3;
            //}




            } else if (event.key === "ç" && toggle_ç == 2) {
                network.unselectAll();

                tempAlert(`Tecla ${event.key} - Desseleciona todos os nós (${nodes.getIds().length})`, 3000);

                toggle_ç = 0;
            }
        });



        // ######################################################################
        // ALTERNA ENTRE A SELEÇÃO DE NÓS COM BASE NO NÚMERO DE ARESTAS - TECLA Ç
        // ######################################################################

        var count_Ç = 0;

        document.addEventListener("keypress", function(event) {
            
            if (event.key === "Ç") {

                network.unselectAll();
                
                let nod = nodes.getIds();
                

                let num_arestas = [];
                for (var i = 0; i < nod.length; i++) {
                    let temp = network.body.nodes[nod[i]].edges.length;
                    temp = parseInt(temp);
                    num_arestas.push(temp);
                }

                num_arestas = [...new Set(num_arestas)];

                num_arestas.sort((a, b) => a - b);
                
                //console.log(num_arestas);

                let nodToSelect = [];
                for (var i = 0; i < nod.length; i++) {
                    if (network.body.nodes[nod[i]].edges.length == num_arestas[count_Ç]) {
                        nodToSelect.push(nod[i]);
                    }
                }
                
                network.selectNodes(nodToSelect);
            
                let percentual = (nodToSelect.length / nod.length) * 100;
                percentual = percentual.toFixed(1);


                if (num_arestas.length == count_Ç) {
                    tempAlert(`Tecla ${event.key} - Desseleciona todos os nós`, 3000);
                    count_Ç = 0;
                
                } else {
                    tempAlert(`Tecla ${event.key} - Seleciona nós com ${num_arestas[count_Ç]} arestas (${nodToSelect.length}) ${percentual}%`, 3000);
                    count_Ç += 1;
                }
            }
        });



        // #####################################################
        // ABRE PÁGINA COM DETALHES DA PESSOA JURÍDICA - TECLA d
        // #####################################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'd' && network.getSelectedNodes().length == 1) {
                var sel = network.getSelectedNodes();
                for (var i=0; i<sel.length; i++) {
                    if (sel.length > 0) {
                        for (var i=0; i<sel.length; i++) {

                            // PARA CONSULTA DE CNPJ
                            if (sel[i].slice(0, 3) == "PJ_") {
                                var cnpj = sel[i].slice(3);
                                var link = "http://cnpj.info/" + cnpj;
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');

                            // PARA CONSULTA DE ENDEREÇO
                            } else if (sel[i].slice(0, 3) == "EN_") {
                                var endereco = sel[i].slice(3).replace();
                                endereco = endereco.replaceAll(' ', '+');
                                var link = "https://maps.google.com/maps?q=" + endereco;
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');

                            // PARA CONSULTA DE E-MAIL
                            } else if (sel[i].slice(0, 3) == "EM_") {
                                var email = sel[i].slice(3).replace();
                                email = email.replaceAll(' ', '+').replace('@', '%40');
                                var link = "https://maps.google.com/search?q=" + '"' + email + '"';
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');
                            
                            // PARA CONSULTA DE PESSOA FÍSICA
                            } else if (sel[i].slice(0, 3) == "PF_") {
                                var nome_socio = sel[i].slice(15);
                                nome_socio = nome_socio.replaceAll(' ', '-');
                                var link = "http://www.consultasocio.com/q/sa/" + nome_socio;
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');
                            }

                            // Exibe informações na tela
                            tempAlert(`Tecla ${event.key} - Exibe detalhes do nó selecionado (exceto TE)`, 3000);

                        }
                    }
                }

            } else if (event.key == 'd' && network.getSelectedNodes().length != 1) {
                tempAlert(`Tecla ${event.key} - Selecione um nó para exibir detalhes (exceto TE)`, 3000);
            }
        });



        // ############################################
        // ABRE PESQUISA NOS DIÁRIOS OFICIAIS - TECLA D
        // ############################################

        document.addEventListener('keydown', (event) => {

            // ABRE PESQUISA NOS DIÁRIOS OFICIAIS (BUSCA EXATA)
            if (event.key == 'D' && network.getSelectedNodes().length == 1) {

                let sel = network.getSelectedNodes();

                if (sel.length == 1) {

                    // PESSOA FÍSICA
                    if (sel[0].slice(0, 3) == 'PF_') {
                        let parametro = sel[0].slice(15);

                        parametro = parametro.replaceAll('%20', '+');
                        let link_jusbrasil = "https://www.jusbrasil.com.br/busca?q=" + '"' + parametro + '"';
                        window.open(url=link_jusbrasil, target="_blank", windowFeature='left=300,top=200,width=600,height=500');

                        parametro = parametro.replaceAll('+', '%20');
                        let link_portal_transp = "https://portaldatransparencia.gov.br/busca?termo=" + '"' + parametro + '"';
                        window.open(url=link_portal_transp, target="_blank", windowFeature='left=250,top=170,width=600,height=500');

                        parametro = parametro.replaceAll(' ', '%20');
                        let link_querido_diario = "https://queridodiario.ok.org.br/pesquisa?term=" + "%22" + parametro + "%22";
                        window.open(url=link_querido_diario, target="_blank", windowFeature='left=200,top=140,width=600,height=500');
                        
                        parametro = parametro.replaceAll(' ', '%20');
                        let link_dom_es = "https://ioes.dio.es.gov.br/buscanova/#/dom/p=1&q=" + "%22" + parametro + "%22";
                        window.open(url=link_dom_es, target="_blank", windowFeature='left=150,top=110,width=600,height=500');
                        
                        parametro = parametro.replaceAll(' ', '%20');
                        let link_dio_es = "https://ioes.dio.es.gov.br/buscanova/#/p=1&q=" + "%22" + parametro + "%22";
                        window.open(url=link_dio_es, target="_blank", windowFeature='left=100,top=80,width=600,height=500');

                        parametro = parametro.replaceAll('%20', '+');
                        let link_dou = "https://www.in.gov.br/consulta/-/buscar/dou?q=" + "%22" + parametro + "%22&s=todos&exactDate=all&sortType=0";
                        window.open(url=link_dou, target="_blank", windowFeature='left=50,top=50,width=600,height=500');

                    
                    // PESSOA JURÍDICA
                    } else if (sel[0].slice(0, 3) == 'PJ_') {
                        let parametro = nodes.get(sel[0]);
                        parametro = parametro.label
                        
                        parametro = parametro.replaceAll('%20', '+');
                        let link_jusbrasil = "https://www.jusbrasil.com.br/busca?q=" + '"' + parametro + '"';
                        window.open(url=link_jusbrasil, target="_blank", windowFeature='left=300,top=200,width=600,height=500');
                        
                        parametro = parametro.replaceAll(' ', '%20');
                        let link_portal_transp = "https://portaldatransparencia.gov.br/busca?termo=" + '"' + parametro + '"';
                        window.open(url=link_portal_transp, target="_blank", windowFeature='left=250,top=170,width=600,height=500');

                        parametro = parametro.replaceAll(' ', '%20');
                        let link_querido_diario = "https://queridodiario.ok.org.br/pesquisa?term=" + "%22" + parametro + "%22";
                        window.open(url=link_querido_diario, target="_blank", windowFeature='left=200,top=140,width=600,height=500');

                        parametro = parametro.replaceAll(' ', '%20');
                        let link_dom_es = "https://ioes.dio.es.gov.br/buscanova/#/dom/p=1&q=" + "%22" + parametro + "%22";
                        window.open(url=link_dom_es, target="_blank", windowFeature='left=150,top=110,width=600,height=500');
                        
                        parametro = parametro.replaceAll(' ', '%20');
                        let link_dio_es = "https://ioes.dio.es.gov.br/buscanova/#/p=1&q=" + "%22" + parametro + "%22";
                        window.open(url=link_dio_es, target="_blank", windowFeature='left=100,top=80,width=600,height=500');

                        parametro = parametro.replaceAll('%20', '+');
                        let link_dou = "https://www.in.gov.br/consulta/-/buscar/dou?q=" + "%22" + parametro + "%22&s=todos&exactDate=all&sortType=0";
                        window.open(url=link_dou, target="_blank", windowFeature='left=50,top=50,width=600,height=500');

                    }

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Pesquisa nó selecionado no DIO-ES, DOM-ES, DOU, Querido Diário, Portal da Transparência e Jusbrasil`, 3000);
                }

            } else if (event.key == 'D' && network.getSelectedNodes().length != 1) {
                tempAlert(`Tecla ${event.key} - Selecione um nó para pesquisar no DIO-ES, DOM-ES, DOU, Querido Diário, Portal da Transparência e Jusbrasil`, 3000);
            }
        });



        // ###################################
        // DELETE NÓS SELECIONADOS - TECLA DEL
        // ###################################

        var del_array = []
        var total_n = nodes.getIds();

        document.addEventListener('keydown', (event) => {
            if (event.key == "Delete") {

                let sel = network.getSelectedNodes();
                network.deleteSelected(sel);

                del_array = del_array.concat(sel);

                let percentual = (del_array.length / total_n.length) * 100;
                percentual = percentual.toFixed(1);

                // INCLUIR AVISO PERMANENTE QUE NÓS FORAM DELETADOS
                let el = document.createElement("div");
                el.setAttribute("style","position:absolute;top:70px;right:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                el.innerHTML = `Nós deletados (${del_array.length}/${total_n.length}) ${percentual}%`;
                document.body.appendChild(el);

                // INCLUÍDO PARA EVITAR ERRO APÓS DELETAR NÓS E RETIRAR O MOUSE DE CIMA DE NÓS ADJACENTES AOS QUE FORAM DELETADOS
                network.unselectAll();

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Deleta nós selecionados (${sel.length})`, 3000);

            }
        });



        // ###############################################
        // OCULTA E EXIBE LABELS DE TODOS OS NÓS - TECLA e
        // ###############################################

        var toggle_no = false
        var toggle_no_2 = false
        var toggle_no_3 = false
        document.addEventListener('keydown', (event) => {
            if (event.key == 'e') {

                // 0 0 0
                if (toggle_no_3 == false && toggle_no_2 == false && toggle_no == false) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 100
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    //network.redraw();
                    toggle_no_3 = false
                    toggle_no_2 = false
                    toggle_no = true

                    
                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Oculta rótulos dos nós`, 3000);

                // 0 0 1
                } else if (toggle_no_3 == false && toggle_no_2 == false && toggle_no == true) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 30,
                                max: 30,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 8
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    //network.redraw();
                    toggle_no_3 = false
                    toggle_no_2 = true
                    toggle_no = false

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Uniformiza tamanho dos nós`, 3000);
                    

                // 0 1 0
                } else if (toggle_no_3 == false && toggle_no_2 == true && toggle_no == false) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 30,
                                max: 30,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 100
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    //network.redraw();
                    toggle_no_3 = false
                    toggle_no_2 = true
                    toggle_no = true

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Oculta rótulos dos nós`, 3000);

                // 0 1 1
                } else if (toggle_no_3 == false && toggle_no_2 == true && toggle_no == true) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 50,
                                max: 50,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 100
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    toggle_no_3 = true
                    toggle_no_2 = false
                    toggle_no = false

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Aumenta tamanho dos nós`, 3000);

                // 1 0 0
                } else if (toggle_no_3 == true && toggle_no_2 == false && toggle_no == false) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 70,
                                max: 70,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 100
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    toggle_no_3 = true
                    toggle_no_2 = false
                    toggle_no = true

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Aumenta tamanho dos nós`, 3000);

                // 1 0 1
                } else if (toggle_no_3 == true && toggle_no_2 == false && toggle_no == true) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 90,
                                max: 90,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 15,
                                    drawThreshold: 100
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    toggle_no_3 = true
                    toggle_no_2 = true
                    toggle_no = false

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Aumenta tamanho dos nós`, 3000);

                // 1 1 0
                } else if (toggle_no == false && toggle_no_2 == true && toggle_no_3 == true) {
                    network.setOptions({
                        nodes: {
                            scaling: {
                                min: 20,
                                max: 100,
                                label: {
                                    enabled: true,
                                    min: 15,
                                    max: 40,
                                    drawThreshold: 8
                                }
                            },
                            physics: true
                        }
                    });
                    //network.stabilize();
                    //network.body.view.scale = 0.8
                    toggle_no_3 = false
                    toggle_no_2 = false
                    toggle_no = false

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Retorna ao tamanho inicial dos nós`, 3000);
                }
            }
        });



        // ##############################
        // CENTRALIZA O GRAFO - TECLA ESC
        // ##############################

        document.addEventListener('keydown', (event) => {

            if (event.key == "Escape") {
                network.fit();

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Enquadra todos os nós`, 3000);
            }
        });



        // ###############################################
        // CONGELA E DESCONGELA NÓS SELECIONADOS - TECLA f
        // ###############################################

        // https://www.toptal.com/developers/keycode

        var fixos = [];
        document.addEventListener('keydown', (event) => {

            // Alterna congelamento dos nós selecionados
            if (event.key == "f" && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length > 0) {

                sel_f = network.getSelectedNodes();
                
                if (sel_f.length > 0) {

                    for (let i=0; i<sel_f.length; i++) {

                        // Se o nó selecionado não estiver em fixos, congela e inclui em fixos
                        if (network.body.nodes[sel_f[i]].options.font.background != 'yellow') {
                            nodes.update({id: sel_f[i], physics: false});
                            nodes.update({id: sel_f[i], font: {strokeColor: "transparent"}});
                            nodes.update({id: sel_f[i], font: {background: "yellow"}});
                            fixos.push(sel_f[i]);

                            // Exibe informações na tela
                            tempAlert(`Tecla ${event.key} - Fixa nós selecionados na área do grafo`, 3000);
                        }

                        // Se estiver, exclui de fixos e descongela
                        else if (network.body.nodes[sel_f[i]].options.font.background == 'yellow') {
                            nodes.update({id: sel_f[i], physics: true});
                            nodes.update({id: sel_f[i], font: {strokeColor: "#ffffff"}});
                            nodes.update({id: sel_f[i], font: {background: "transparent"}});
                            fixos.slice(sel_f[i], 1);

                            // Exibe informações na tela
                            tempAlert(`Tecla ${event.key} - Desafixa nós selecionados na área do grafo`, 3000);
                        }
                    }
                }

            } else if (event.key == "f" && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um ou mais nós para fixá-los na área do grafo`, 3000);
            }
        });



        // #####################################################################
        // AJUSTA ALTURA DA JANELA DE VISUALIZAÇÃO AO PRESSIONAR F11 - TECLA F11
        // #####################################################################

        // Só funciona na primeira vez que se aperta F11. No retorno, deve-se apertar tecla b

        document.addEventListener('keydown', (event) => {

            if (event.key == 'F11') {

                function f11() {
                    // EXTRAI DIMENSÕES DA JANELA DO NAVEGADOR
                    //let w = window.innerWidth;
                    let h = window.innerHeight;
                    
                    // AJUSTA ALTURA DA ÁREA DO GRAFO À ALTURA DA JANELA
                    document.querySelector('#mynetwork').style.height = h - 20;
                }
                setTimeout(f11, 200);

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Ativa/desativa modo tela cheia`, 3000);

            } 
        });



        // ##########################################
        // REALIZA PESQUISA GERAL NO GOOGLE - TECLA g
        // ##########################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'g' && network.getSelectedNodes().length == 1) {
                let sel = network.getSelectedNodes();
                for (let i=0; i<sel.length; i++) {
                    if (sel.length > 0) {
                        for (let i=0; i<sel.length; i++) {

                            if (sel[i].slice(0, 3) == 'PF_') {
                                let parametro = sel[i].slice(15);
                                parametro = parametro.replaceAll(' ', '+');
                                //console.log('parametro1:', parametro);
                                let link = "https://www.google.com/search?q=" + parametro;
                                //console.log(link);
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');

                            } else if (sel[i].slice(0, 3) == 'PJ_') {
                                let parametro = nodes.get(sel[i]);
                                parametro = parametro.label
                                parametro = parametro.replaceAll(' ', '+');
                                //console.log('parametro2:', parametro);
                                let link = "https://www.google.com/search?q=" + parametro;
                                //console.log(link);
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');
                            
                            } else if (sel[i].slice(0, 3) == 'PE_') {
                                let parametro = nodes.get(sel[i]);
                                parametro = parametro.id.slice(3)
                                parametro = parametro.replaceAll(' ', '+');
                                //console.log('parametro2:', parametro);
                                let link = "https://www.google.com/search?q=" + parametro;
                                //console.log(link);
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');

                            // PARA CONSULTA DE ENDEREÇO
                            } else if (sel[i].slice(0, 3) == "EN_") {
                                let endereco = sel[i].slice(3).replace();
                                endereco = endereco.replaceAll(' ', '+');
                                var link = "https://maps.google.com/search?q=" + endereco;
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');

                            // PARA CONSULTA DE E-MAIL
                            } else if (sel[i].slice(0, 3) == "EM_") {
                                let email = sel[i].slice(3).replace();
                                email = email.replaceAll(' ', '+').replace('@', '%40');
                                var link = "https://maps.google.com/search?q=" + email;
                                window.open(link, 'popup','left=50,top=100,width=600,height=600');
                            }

                            // Exibe informações na tela
                            tempAlert(`Tecla ${event.key} - Pesquisa nó selecionado no Google (exceto TE)`, 3000);

                        }  
                    }
                }

            } else if (event.key == 'g' && network.getSelectedNodes().length != 1) {
                tempAlert(`Tecla ${event.key} - Selecione um nó para pesquisar no Google (exceto TE)`, 3000);
            }
        });



        // #############################################
        // REALIZA PESQUISA AVANÇADA NO GOOGLE - TECLA G
        // #############################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'G' && network.getSelectedNodes().length == 1) {
                let sel = network.getSelectedNodes();
                for (let i=0; i<sel.length; i++) {
                    if (sel.length > 0) {
                        for (let i=0; i<sel.length; i++) {

                            // PESSOA FÍSICA
                            if (sel[i].slice(0, 3) == 'PF_') {
                                let parametro = sel[i].slice(15);
                                parametro = parametro.replaceAll(' ', '+');
                                
                                let link_es_gov = "https://www.google.com/search?q=" + "site%3Aes.gov.br+" + '"' + parametro + '"';
                                window.open(url=link_es_gov, target="_blank", windowFeature='left=250,top=200,width=600,height=500');
                                
                                let link_gov = "https://www.google.com/search?q=" + "site%3Agov.br+" + '"' + parametro + '"';
                                window.open(url=link_gov, target="_blank", windowFeature='left=200,top=170,width=600,height=500');

                                let link_jus = "https://www.google.com/search?q=" + "site%3Ajus.br+" + '"' + parametro + '"';
                                window.open(url=link_jus, target="_blank", windowFeature='left=150,top=110,width=600,height=500');

                                let link_mp = "https://www.google.com/search?q=" + "site%3Amp.br+" + '"' + parametro + '"';
                                window.open(url=link_mp, target="_blank", windowFeature='left=100,top=80,width=600,height=500');

                                let link_tc = "https://www.google.com/search?q=" + "site%3Atc.br+" + '"' + parametro + '"';
                                window.open(url=link_tc, target="_blank", windowFeature='left=50,top=50,width=600,height=500');

                            
                            // PESSOA JURÍDICA
                            } else if (sel[i].slice(0, 3) == 'PJ_') {
                                let parametro = nodes.get(sel[i]);
                                parametro = parametro.label
                                parametro = parametro.replaceAll(' ', '+');
                                
                                let link_es_gov = "https://www.google.com/search?q=" + "site%3Aes.gov.br+" + parametro;
                                window.open(url=link_es_gov, target="_blank", windowFeature='left=250,top=170,width=600,height=500');

                                let link_gov = "https://www.google.com/search?q=" + "site%3Agov.br+" + parametro;
                                window.open(url=link_gov, target="_blank", windowFeature='left=200,top=140,width=600,height=500');
                                
                                let link_jus = "https://www.google.com/search?q=" + "site%3Ajus.br+" + parametro;
                                window.open(url=link_jus, target="_blank", windowFeature='left=150,top=110,width=600,height=500');

                                let link_mp = "https://www.google.com/search?q=" + "site%3Amp.br+" + parametro;
                                window.open(url=link_mp, target="_blank", windowFeature='left=100,top=80,width=600,height=500');

                                let link_tc = "https://www.google.com/search?q=" + "site%3Atc.br+" + parametro;
                                window.open(url=link_tc, target="_blank", windowFeature='left=50,top=50,width=600,height=500');

                            }

                            // Exibe informações na tela
                            tempAlert(`Tecla ${event.key} - Pesquisa nó selecionado (apenas PF e PJ) no Google nos domínios TC.br, MP.br, JUS.br, GOV.br e ES.GOV.br`, 3000);
                        }
                    }
                }

            } else if (event.key == 'G' && network.getSelectedNodes().length != 1) {
                tempAlert(`Tecla ${event.key} - Selecione um nó (apenas PF e PJ) para pesquisar no Google nos domínios TC.br, MP.br, JUS.br, GOV.br e ES.GOV.br`, 3000);
            }
        });



        // #####################################################
        // ARBRE ARQUIVO HTML CONTENDO MANUAL DO SINAR - TECLA h
        // #####################################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'h') {

                // O nome 'popup_help' faz com que essa janela seja exclusiva, não sendo usada por outros popups do programa.
                window.open('help.html', 'popup_help','left=50,top=100,width=600,height=600');

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Abre manual do SINARC`, 3000);
            }
        });



        // ########################################
        // ALTERNA ENTRE NÓS SELECIONADOS - TECLA i
        // ########################################

        var counter = 0 // Deve iniciar com valor zero para funcionar
        document.addEventListener('keydown', (event) => {
            if (event.key == "i" && network.getSelectedNodes().length > 0) {
                
                let s = network.getSelectedNodes();
                //console.log('counter:', counter);
                //console.log('s:', s[counter]);
                //network.fit({nodes: network.getConnectedNodes(s[counter]), maxZoomLevel: 0.7, minZoomLevel: 0.7});


                //if (s.length > 0 ) {
                //    network.setOptions({physics: {enabled: false}});
                //    let p = '';
                //    let num = 0;
                //    num = 700 + nodes.length / 10;
                //    console.log(num + 'ms');
                //    for (let i=0; i<s.length; i++) {
                //        esconde();
                //        p = network.getPosition(s[i]);
                //        network.moveTo({
                //            position: {x: p['x'], y: p['y']},
                //            scale: 1,
                //            offset: {x:0, y:0},
                //            animation: {
                //                duration: 500,
                //                easingFunction: 'linear'
                //            }
                //       });
                //        setTimeout(exibe, num);
                //        network.releaseNode();
                //    }
                //    network.setOptions({physics: {enabled: true}});

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Centraliza e alterna entre nós selecionados (${counter + 1}/${s.length})`, 3000);

                network.fit({nodes: [s[counter]], maxZoomLevel: 1, minZoomLevel: 1});
                if (counter == s.length - 1) {
                    counter = 0;
                } else if (counter < s.length) {
                    counter = counter + 1;
                }

            } else if (event.key == "i" && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione pelo menos um nó para centralizá-los na tela de forma alternada`, 3000);
            }
        });



        // ################################
        // ALTERNA SELEÇÃO DE NÓS - TECLA I
        // ################################

        var se_1 = [];
        var se_2 = [];
        var toggle_I = true

        document.addEventListener('keydown', (event) => {

            
            if (event.key == "I" && network.getSelectedNodes().length > 0 && toggle_I == true) {

                // EXTRAI NÓS SELECIONADOS
                se_1 = network.getSelectedNodes();

                // EXTRAI NÓS NÃO SELECIONADOS
                se_2 = nodes.getIds().filter(x => !se_1.includes(x));

                // SELECIONA NÓS QUE NÃO ESTAVAM SELECIONADOS
                network.selectNodes(se_2);

                // COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
                for (let i=0; i<se_1.length; i++) {
                    nodes.update({id: se_1[i], font: {color: '#777777'}});
                }

                network.redraw();

                toggle_I = false

                tempAlert(`Tecla ${event.key} - Inverte seleção de nós (${se_2.length}/${nodes.getIds().length})`, 3000);

                
            } else if (event.key == "I" && network.getSelectedNodes().length > 0  && toggle_I == false) {
                
                network.unselectAll();

                // SELECIONA NÓS QUE NÃO ESTAVAM SELECIONADOS
                network.selectNodes(se_1);
                network.redraw();

                toggle_I = true

                tempAlert(`Tecla ${event.key} - Inverte seleção de nós (${se_1.length}/${nodes.getIds().length})`, 3000);

            } else if (event.key == "I" && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione pelo menos um nó para inverter a seleção`, 3000);
            }

            // RETORNA COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
            //for (let i=0; i<no_sel_pv.length; i++) {
            //    nodes.update({id: no_sel_pv[i], font: {color: '#777777'}});
            //}

        });



        // ###################################################
        // SELECIONA NÓS COMUNS AOS NÓS SELECIONADOS - TECLA j
        // ###################################################

        document.addEventListener('keydown', (event) => {
            if (event.key == "j" && network.getSelectedNodes().length >= 2) {
                let sel = network.getSelectedNodes();
                if (sel.length >= 2) {
                    let caixa = [];
                    for (let i=0; i<sel.length; i++) {
                        let temp = network.getConnectedNodes(sel[i]);
                        for (let k=0; k<temp.length; k++) {
                            caixa.push(temp[k]);
                        }    
                    }
                    let counts = {};
                    for (const num of caixa) {
                    counts[num] = counts[num] ? counts[num] + 1 : 1;
                    }
                    let comum = [];
                    for (const v in counts) {
                        if (counts[v] == sel.length) {
                            if (!(v in comum)) {
                                comum.push(v);
                            }
                        }
                    }
                    //console.log(comum);
                    network.unselectAll();
                    network.selectNodes(comum);

                    //network.fit({nodes: comum});

                    for (let i=0; i<sel.length; i++) {
                        nodes.update({id: sel[i], font: {color: '#777777'}});
                    }

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes comuns (${comum.length})`, 3000);

                }

            } else if (event.key == "j" && network.getSelectedNodes().length < 2) {
                tempAlert(`Tecla ${event.key} - Selecione pelo menos dois nós`, 3000);
            }
        });



        // ###############################################################n
        // SELECIONA NÓS QUE NÃO SÃO COMUNS AOS NÓS SELECIONADOS - TECLA J
        // ###############################################################

        document.addEventListener('keydown', (event) => {
            if (event.key == "J" && network.getSelectedNodes().length >= 2) {

                let sel = network.getSelectedNodes();
                //console.log('sel:', sel)

                if (sel.length >= 2) {

                    let caixa = [];
                    for (let i=0; i<sel.length; i++) {
                        let temp = network.getConnectedNodes(sel[i]);
                        for (let k=0; k<temp.length; k++) {
                            caixa.push(temp[k]);
                        }    
                    }
                    let counts = {};
                    for (const num of caixa) {
                        counts[num] = counts[num] ? counts[num] + 1 : 1;
                    }
                    //console.log(counts);
                    
                    let comum = [];
                    for (const v in counts) {
                        if (counts[v] == sel.length) {
                            if (!(v in comum)) {
                                comum.push(v);
                            }
                        }
                    }
                    //console.log(comum);

                    let diff = caixa
                        .filter(x => !comum.includes(x))
                        .concat(comum.filter(x => !caixa.includes(x)));


                    let diff2 = diff.filter(x => !sel.includes(x));
                    console.log('diff2:', diff2);

                    //network.fit({nodes: diff2});


                    //diff = [...new Set(diff)];
                    //console.log('diff:', diff);
                    
                    network.unselectAll();
                    network.selectNodes(diff2);

                    //for (let j=0; j<sel.length; j++) {
                    //    network.body.nodes[sel[j]].selected = false;
                    //    network.redraw();
                    //}

                    for (let i=0; i<sel.length; i++) {
                        nodes.update({id: sel[i], font: {color: '#777777'}});
                    }

                    // Exibe informações na tela
                    //let n = nodes.getIds().length - comum.length - sel.length
                    
                    //tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes que não são comuns (${diff2.length})`, 3000);
                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes que não são comuns (${network.getSelectedNodes().length})`, 3000);

                }

            } else if (event.key == "J" && network.getSelectedNodes().length < 2) {
                tempAlert(`Tecla ${event.key} - Selecione pelo menos dois nós`, 3000);
            }
        });
        


        // ###############################
        // ALTERNA ENTRE LAYOUTS - TECLA k
        // ###############################

        var toggle_hieraquic = false
        var toggle_hieraquic_2 = false
        var toggle_hieraquic_3 = false
        document.addEventListener('keydown', (event) => {
            if (event.key == 'k') {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}})
                }

                //network.view.body.view.scale = 0.8;
                //network.fit();
                //network.redraw();
                
                // GRAFO NA HORIZONTAL (LEFT TO RIGHT)
                
                // 0 0 0
                if (toggle_hieraquic_3 == false && toggle_hieraquic_2 == false && toggle_hieraquic == false) {
                    network.setOptions({
                        layout: {
                            hierarchical: {
                            enabled: true,
                            levelSeparation: 500,
                            nodeSpacing: 100,
                            direction: "LR",
                            sortMethod: "directed",
                            shakeTowards: "leaves"
                            }
                        },
                        physics: {
                            hierarchicalRepulsion: {
                            centralGravity: 0
                            },
                            minVelocity: 0.75,
                            solver: "hierarchicalRepulsion",
                        }
                    });
                    
                    //network.setOptions({physics: {enabled: false}});
                    //network.stabilize();
                    
                    //network.redraw();
                    network.setOptions({nodes: {physics: false}});
                    network.setOptions({nodes: {physics: true}});
                    //network.fit();
                    toggle_hieraquic_3 = false
                    toggle_hieraquic_2 = false
                    toggle_hieraquic = true
                    //console.log('direction: LR, sortMethod: directed, shakeTowaords: leaves');

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Layout hierárquico a partir das folhas (destino das setas)`, 3000);

                // 0 0 1
                } else if (toggle_hieraquic_3 == false && toggle_hieraquic_2 == false && toggle_hieraquic == true) {
                    network.setOptions({
                        layout: {
                            hierarchical: {
                            enabled: true,
                            levelSeparation: 500,
                            nodeSpacing: 100,
                            direction: "LR",
                            sortMethod: "directed",
                            shakeTowards: "roots"
                            }
                        },
                        physics: {
                            hierarchicalRepulsion: {
                            centralGravity: 0
                            },
                            minVelocity: 0.75,
                            solver: "hierarchicalRepulsion",
                        }
                    });
                    
                    //network.setOptions({physics: {enabled: false}});
                    //network.stabilize();
                    
                    //network.redraw();
                    network.setOptions({nodes: {physics: false}});
                    network.setOptions({nodes: {physics: true}});
                    //network.fit();
                    toggle_hieraquic_3 = false
                    toggle_hieraquic_2 = true
                    toggle_hieraquic = false
                    //console.log('direction: LR, sortMethod: directed, shakeTowaords: roots');

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Layout hierárquico a partir das raízes (origem das setas)`, 3000);

                // 0 1 0
                //} else if (toggle_hieraquic_3 == false && toggle_hieraquic_2 == true && toggle_hieraquic == false) {
                //    network.setOptions({
                //        layout: {
                //            hierarchical: {
                //            enabled: true,
                //            levelSeparation: 500,
                //            nodeSpacing: 100,
                //            direction: "LR",
                //            sortMethod: "hubsize",
                //            shakeTowards: "leaves"
                //            }
                //        },
                //        physics: {
                //            hierarchicalRepulsion: {
                //            centralGravity: 0
                //            },
                //            minVelocity: 0.75,
                //            solver: "hierarchicalRepulsion",
                //        }
                //    });
                    
                    //network.stabilize();
                    //network.setOptions({physics: {enabled: false}});
                    
                //    network.setOptions({nodes: {physics: false}});
                //    network.setOptions({nodes: {physics: true}});
                    //network.fit();
                //    toggle_hieraquic_3 = false
                //    toggle_hieraquic_2 = true
                //    toggle_hieraquic = true
                //    console.log('direction: LR, sortMethod: hubsize, shakeTowaords: leaves');

                // 0 1 0
                } else if (toggle_hieraquic_3 == false && toggle_hieraquic_2 == true && toggle_hieraquic == false) {
                    network.setOptions({
                        layout: {
                            hierarchical: {
                            enabled: false,
                            direction: "LR",
                            sortMethod: "directed",
                            shakeTowards: "leaves"
                            }
                        }
                    });
                    
                    network.setOptions({
                        edges: {
                            smooth: false
                        }
                    });
                    
                    network.setOptions({
                        physics: {
                            solver: 'barnesHut',
                            barnesHut: {
                            gravitationalConstant: -10000,
                            springLength: 200
                            },
                            minVelocity: 0.75
                        }
                    });
                    //network.setOptions({physics: {enabled: true}});
                    //network.stabilize();
                    //network.fit();

                    // 1 second delay
                    //setTimeout(function(){
                    //    network.fit();
                    //}, 1000);

                    //network.view.body.view.scale = 0.8;
                    network.redraw();
    
                    toggle_hieraquic_3 = false
                    toggle_hieraquic_2 = false
                    toggle_hieraquic = false
                    //console.log('solver: BarnesHut');

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Layout gravitacional`, 3000);
                }
            }
        });



        // #####################################
        // EXIBE NOVOS LAYOUTS CRIADOS - TECLA K
        // #####################################

        // Criado com o auxílio do Chat GPT: criar código em Python que posicione n pontos de forma equidistante sobre uma circunferência no plano cartesiano
        // Validação realizada no site para plotagem de pontos no plano cartesiano: https://www.desmos.com/calculator?lang=pt-BR
        // O Critério adotado para ordenar os nós sobre a circunferência foi o número de conexões.
        
        var toggle_K = 0;

        document.addEventListener('keydown', (event) => {

            if (event.key == 'K' && toggle_K == 0) {

                network.physics.physicsEnabled = true;
                network.physics.startSimulation()
                
                let nos2 = nodes.getIds();
                
                // CALCULA PONTOS NA CIRCUNFERÊNCIA
                let n2 = nos2.length;

                // O multiplicador do raio (20) controla a separação entre os nós sobre a circunferência
                function pontosCircunferencia(n, R=n2 * 20) {
                    
                    let d = 2 * R * Math.sin(Math.PI / n2);
                    let pontos = [];
                    
                    for (let i = 0; i < n2; i++) {
                        let angulo = 2 * Math.PI * i / n2 - Math.PI / 2;
                        let x = R * Math.cos(angulo);
                        let y = R * Math.sin(angulo);
                        let ponto = [x, y];
                        pontos.push(ponto);
                    }
                    
                    let pontosDistancia = [];
                    
                    for (let i = 0; i < n2; i++) {
                        let pontoAtual = pontos[i];
                        let pontoAnterior = pontos[(i - 1 + n2) % n2];
                        let pontoProximo = pontos[(i + 1) % n2];
                        let distanciaAnterior = Math.sqrt(Math.pow((pontoAtual[0] - pontoAnterior[0]), 2) + Math.pow((pontoAtual[1] - pontoAnterior[1]), 2));
                        let distanciaProxima = Math.sqrt(Math.pow((pontoAtual[0] - pontoProximo[0]), 2) + Math.pow((pontoAtual[1] - pontoProximo[1]), 2));
                        if (Math.abs(distanciaAnterior - d) < 1e-9 && Math.abs(distanciaProxima - d) < 1e-9) {
                            pontosDistancia.push(pontoAtual);
                        }
                    }
                    
                    return pontosDistancia;
                }
                
                let pontos = pontosCircunferencia(n2);


                // Create a new array from the list of nodes
                const nodeArray = nodes.get({
                    fields: ['id']
                });
                
                // Sort the array of nodes based on their degree
                nodeArray.sort((a, b) => {
                const degreeA = edges.get({
                    filter: edge => edge.from === a.id || edge.to === a.id
                }).length;
                const degreeB = edges.get({
                    filter: edge => edge.from === b.id || edge.to === b.id
                }).length;
                return degreeB - degreeA;
                });

                
                // Extrai ids dos nós
                nos3 = [];
                for (let i=0; i<nodeArray.length; i++) {
                    nos3.push(nodeArray[i].id);
                }
                
                
                // ARREDONDA TUPLAS PARA DUAS CASAS DECIMAIS
                function arredondaTuplas(array) {
                let novoArray = [];
                for (let i = 0; i < array.length; i++) {
                    let tupla = [];
                    for (let j = 0; j < array[i].length; j++) {
                    tupla.push(parseFloat(array[i][j].toFixed(2)));
                    }
                    novoArray.push(tupla);
                }
                return novoArray;
                }
                
                lista = arredondaTuplas(pontos);
                
                
                // ALTERA A POSIÇÃO DOS NÓS
                for (let i=0; i<nos3.length; i++) { 
                network.body.nodes[nos3[i]].x = lista[i][0];
                network.body.nodes[nos3[i]].y = lista[i][1];
                }
                
                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();
                
                network.redraw();

                network.fit();

                toggle_K = 1;

                tempAlert(`Tecla ${event.key} - Layout circular`, 3000);
            
            } else if (event.key == 'K' && toggle_K == 1) {

                network.physics.physicsEnabled = true;
                network.physics.startSimulation()

                toggle_K = 0;

                tempAlert(`Tecla ${event.key} - Layout gravitacional`, 3000);
                
            }
        });



        // ###################################
        // OCULTA LABELS DAS ARESTAS - TECLA l
        // ###################################

        var toggle_label = false
        var toggle_label_2 = false
        document.addEventListener('keydown', (event) => {
            if (event.key == 'l') {

                if (toggle_label == false && toggle_label_2 == false) {
                    network.setOptions({'edges': {'font': {'size': 0}}});
                    //console.log('Oculta texto das arestas')
                    //toggle_label = !toggle_label
                    toggle_label = true

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Oculta rótulos das arestas`, 3000);


                } else if (toggle_label == true && toggle_label_2 == false) {
                    network.setOptions({"edges": {"hidden": true}});
                    //console.log('Oculta arestas')
                    //toggle_label_2 = !toggle_label_2
                    toggle_label_2 = true

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Oculta arestas`, 3000);


                } else if (toggle_label == true && toggle_label_2 == true) {
                    //console.log('Reexibe arestas e textos')
                    network.setOptions({'edges': {'font': {'size': 15}}});
                    network.setOptions({"edges": {"hidden": false}});
                    toggle_label = false
                    toggle_label_2 = false

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Reexibe arestas e rótulos`, 3000);

                } 
            }
        });



        // ########################################
        // EXIBE DISTRIBUIÇÃO DAS ARESTAS - TECLA L
        // ########################################

        document.addEventListener('keydown', (event) => {
            
            if (event.key == 'L') {

                let ta = edges.getIds();
                let aa = [];
                for (let i=0; i<ta.length; i++) {
                    let temp = network.body.edges[ta[i]].title;
                    aa.push(temp);
                }
                
                
                const counts = {};
                for (const num of aa) {
                counts[num] = counts[num] ? counts[num] + 1 : 1;
                }
                
                
                let tex = `DISTRIBUIÇÃO DAS ARESTAS POR TIPO (${ta.length}):\\n`;
                Object.keys(counts)
                    .sort()
                    .forEach(function(v, i) {
                        tex = tex + v + ': ' + counts[v] + '\\n'
                    });
                
                alert(tex);

                tempAlert(`Tecla ${event.key} - Exibe distribuição das aresta por tipo (${ta.length})`, 3000);
            }
        });



        // ######################################################
        // AUMENTA MASSA DE TODOS OS NÓS EM 0.5 UNIDADE - TECLA m
        // ######################################################

        var massa = 1

        document.addEventListener('keydown', (event) => {

            if (event.key == 'm' && network.physics.options.solver == 'barnesHut') {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}})
                } 

                // AUMENTA MASSA EM 0.5               
                massa = massa + 0.5
                
                // AJUSTA A MASSA DE TODOS OS NÓS, EXCETO OS QUE JÁ TIVERAM AS MASSAS AJUSTADAS PELAS TECLAS n E N
                network.setOptions({
                    nodes: {
                        mass: massa
                    }
                });

                //network.selectNodes(nodes.getIds());

                network.redraw();

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Aumenta massa dos nós em 0.5 unidade (geral). Fator: ${massa}`, 3000);
            }

        });



        // ######################################################
        // DIMINUI MASSA DE TODOS OS NÓS EM 0.5 UNIDADE - TECLA M
        // ######################################################

        // Variável 'massa' foi declarada na tecla anterior

        document.addEventListener('keydown', (event) => {

            if (event.key == 'M' && network.physics.options.solver == 'barnesHut') {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}})
                }

                massa = massa - 0.5

                if (massa <= 1) {
                    massa = 1
                }

                // AJUSTA A MASSA DE TODOS OS NÓS, EXCETO OS QUE JÁ TIVERAM AS MASSAS AJUSTADAS PELAS TECLAS n E N
                network.setOptions({
                    nodes: {
                        mass: massa
                    }
                });

                network.redraw();

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Reduz massa dos nós em 0.5 unidade (geral). Fator: ${massa}`, 3000);
            }
        });



        // #######################################################
        // AUMENTA MASSA DOS NÓS COM TAMANHO ACIMA DE 40 - TECLA n
        // #######################################################

        // A partir do momento em que se altera as massas dos nós com as teclas n e N, as teclas m e M passam a não mais atuar sobre esse nós.
        // Enquanto as teclas n e N não forem acionadas, as teclas m e M atuam sobre todos os nós.

        var par_n = 1

        document.addEventListener('keydown', (event) => {

            if (event.key == 'n' && network.physics.options.solver == 'barnesHut') {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}})
                }
                
                let nod = nodes.getIds();

                let sel_e = []

                for (let i=0; i<nod.length; i++) {

                    let size = network.body.nodes[nod[i]].options.size

                    if (size >= t_ref) {
                        sel_e.push(nod[i])
                    }

                    if (size >= t_ref) {

                        // AUMENTA EM 5
                        nodes.update({
                            id: nod[i],
                            mass: par_n + 5
                        });
                    }
                }

                network.selectNodes(sel_e);

                // AUMENTA EM 5
                par_n = par_n + 5

                network.redraw();

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Aumenta massa dos nós com tamanho >= ${t_ref} em 5 unidades. Fator: ${par_n} (${network.getSelectedNodes().length})`, 3000);

            }
        })



        // ##########################################################
        // REDUZ A MASSA DOS NÓS COM TAMANHO ACIMA DO t_ref - TECLA N
        // ##########################################################

        // A variável global 'par_n' foi criada na tecla anterior (Tecla n)

        document.addEventListener('keydown', (event) => {

            if (event.key == 'N' && network.physics.options.solver == 'barnesHut') {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}})
                }

                // REDUZ EM 5
                par_n = par_n - 5

                if (par_n <= 1) {
                    par_n = 1;
                }

                let nod = nodes.getIds();

                let sel_e = [];

                for (let i=0; i<nod.length; i++) {

                    let size = network.body.nodes[nod[i]].options.size

                    if (size >= t_ref) {
                        sel_e.push(nod[i])
                    }

                    if (size >= t_ref) {
                        nodes.update({id: nod[i], mass: par_n});
                    }
                }

                network.selectNodes(sel_e);

                network.redraw();
                
                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Reduz massa dos nós com tamanho >= ${t_ref} em 5 unidades. Fator: ${par_n} (${network.getSelectedNodes().length})`, 3000);
            }
        })



        // ###############################################################
        // COPIA NÓS E NÚMERO DE CAMADAS PARA A ÁREA DE TRABALHO - TECLA o
        // ###############################################################

        document.addEventListener('keydown', (event) => {

            if (event.key == 'o' && network.getSelectedNodes().length > 0) {

                let get_sel = network.getSelectedNodes();


                let num_cam = prompt("Digite o número de camadas adjacentes (ENTER para 1 camada):");

                // INTERROMPE FUNÇÃO SE O BOTÃO "CANCELA" FOR PRESSIONADO
                if (num_cam == null) {
                    return;
                }


                //let del_te_en_em = ''
                //del_te_en_em = confirm("Deseja excluir Telefones, Endereços e E-mails?");



                let conf_destaca = ''
                if (get_sel.length > 1) {
                    conf_destaca = confirm("Deseja destacar caminhos mais curtos entre os nós?");
                }

                async function copyOperation(num_cam) {
                    await navigator.clipboard.writeText(num_cam)
                };

                let temp = [];
                for (let i = 0; i < get_sel.length; i++) {
                    let j = get_sel[i] + "***" + num_cam + "***" + conf_destaca;
                    temp.push(j);
                }

                setTimeout(() => {copyOperation(temp)}, 500)

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Carrega nós selecionados em nova aba (${temp.length})`, 3000);
            
            } else if (event.key == 'o' && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um ou mais nós para abri-los em nova aba`, 3000);
            }
        });



        // ###############################################################
        // CONSULTA PELA EXISTÊNCIA DE FILIAIS DO NÓ SELECIONADO - TECLA O
        // ###############################################################

        document.addEventListener('keydown', (event) => {

            if (event.key == 'O' && network.getSelectedNodes().length == 1) {
                
                let sel = network.getSelectedNodes();
                
                //if (sel.length > 1) {
                //    alert('Selecione apenas uma pessoa jurídica.')
                //    return;
                //}

                let texto = '';
                
                // APENAS PESSOAS JURÍDICAS
                if (sel.length == 1 && ((sel[0].slice(0, 3) == 'PJ_') || (sel[0].slice(0, 3) == 'PE_'))) {

                    texto = '#_#' + sel[0].slice(3, 11) + '@' + '9999' + '#_#';

                    //console.log(texto);

                    async function copyOperation(num_cam) {
                        await navigator.clipboard.writeText(num_cam);
                    }
                    setTimeout(() => {copyOperation(texto)}, 500);

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Carrega todas as filiais do nó selecionado em nova aba`, 3000);
                
                //} else {
                    //alert('Para consulta de filiais, selecione uma - e apenas uma - pessoa juridica.');
                    // Exibe informações na tela
                //    tempAlert(`Tecla ${event.key} - ALERTA: Para consultar filiais, selecione uma - e apenas uma - pessoa jurídica`, 3000);

                }

            } else if (event.key == 'O' && network.getSelectedNodes().length != 1) {
                    tempAlert(`Tecla ${event.key} - Selecione uma pessoa jurídica para abrir suas filiais em nova aba`, 3000);
            }

            //};
        });



        // ###################################
        // ATIVA/DESATIVA PHYSICS - TECLA p
        // ###################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'p') {

                if (network.physics.physicsEnabled == true) {
                    network.physics.physicsEnabled = false;
                    network.physics.stopSimulation();

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Para a interação gravitacional (congela todos os nós)`, 3000);

                } else if (network.physics.physicsEnabled == false) {
                    network.physics.physicsEnabled = true;
                    network.physics.startSimulation();

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Retoma a interação gravitacional (descongela todos os nós)`, 3000);

                }

            }
        });



        // #########################################
        // CONSULTA LIVRE DOS NÓS NO GRAFO - TECLA q
        // #########################################

        // Realiza consulta nos textos dos labels e dos ids dos nós e das arestas. Após a consulta, usar tecla I para alternar entre os nós selecionados.

        var target_id = [];
        //var nodLista = [];
        var con = [];
        
        document.addEventListener('keydown', (event) => {

            if (event.key == 'q') {

                let a = network.getSelectedNodes();
                let temp = '';
                if (a.length == 1) {
                    temp = `id selecionado: ${network.body.nodes[a[0]].id}\\n`
                }

                // ZERA VARIÁVEL GLOBAL. EXIGIDO PELO SCRITP SEGUINTE (TECLA I)
                counter = 0

                let text = prompt(temp + "Digite o texo a ser localizado nos rótulos e nos ids dos NÓS:");
                //console.log('Termo pesquisado:', text);

                // SE FOR FORNECIDO ALGUM PARÂMETRO
                if (text != '') {
                    let a = nodes.getIds();
                    //let target_id = [];
                    //let con = [];
                    for (var i=0; i<a.length; i++) {
                        if (nodes.get(a[i])['label'].toLowerCase().includes(text.toLowerCase()) || a[i].toLowerCase().includes(text.toLowerCase())) {
                            let temp = network.getConnectedNodes(a[i]);
                            con = con.concat(temp);
                            con = con.concat([a[i]])
                            target_id.push(a[i]);
                        }
                    }

                    // Elimina duplicidades
                    con = [...new Set(con)];
                    target_id = [...new Set(target_id)];

                    network.selectNodes(target_id);

                    //console.log('con:', con);
                    //network.fit({nodes: con, maxZoomLevel: 0.7, minZoomLevel: 0.7});

                    if (con.length > 0) {
                        network.fit({nodes: con});
                    }
                    //network.fit({nodes: target_id});

                    //console.log('Itens localizados:', target_id);
                    //alert(target_id.length.toString() + ' itens localizados');

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Pesquisa por "${text}" nos nós (${target_id.length})`, 3000);


                // SE NÃO FOR FORNECIDO NENHUM PARÂMETRO
                } else if (text == '') {

                    let a = edges.getIds();
                    let edgLista = [];
                    let param = prompt('Digite o texo a ser localizado nos rótulos das ARESTAS:');

                    if (param == '') {
                        return;
                    }
                    
                    for (let i=0; i<a.length; i++) {
                        if (network.body.edges[a[i]].options.label != undefined) {
                            let temp = network.body.edges[a[i]].options.label;
                            temp = temp.toLowerCase();
                            
                            function removeAccents(str) {
                                const accents = "ÀÁÂÃÄÅàáâãäåÒÓÔÕÕÖØòóôõöøÈÉÊËèéêëðÇçÐÌÍÎÏìíîïÙÚÛÜùúûüÑñŠšŸÿýŽž";
                                const accentsOut = "AAAAAAaaaaaaOOOOOOooooooEEEEeeeeeCcDIIIIiiiiUUUUuuuuNnSsYyyZz";
                                str = str.split('');
                                str.forEach((letter, index) => {
                                    const i = accents.indexOf(letter);
                                    if (i !== -1) {
                                    str[index] = accentsOut[i];
                                    }
                                });
                                return str.join('');
                            }
                            temp = removeAccents(temp)
                                
                            param = param.toLowerCase();
                            if (temp.includes(param)) {
                                edgLista.push(network.body.edges[a[i]].id);
                            }
                        }
                    }
                    
                    let nodLista = [];
                    for (let i=0; i<edgLista.length;i++) {
                        nodLista.push(network.body.edges[edgLista[i]].fromId);
                        target_id.push(network.body.edges[edgLista[i]].fromId);
                    }
                    
                    for (let i=0; i<nodLista.length; i++) {
                        let temp = network.getConnectedNodes(nodLista[i]);
                        con = con.concat(temp);
                    }

                    con = con.concat(target_id);

                    // Elimina duplicidades
                    con = [...new Set(con)];
                    target_id = [...new Set(target_id)];

                    //network.selectNodes(nodLista);

                    //target_id = target_id.concat(nodLista);

                    network.selectNodes(target_id);

                    if (con.length > 0) {
                        network.fit({nodes: con});
                    }

                    //tempAlert(`Tecla ${event.key} - Pesquisa por "${param}" nas arestas (${nodLista.length})`, 3000);
                    tempAlert(`Tecla ${event.key} - Pesquisa por "${param}" nas arestas (${target_id.length})`, 3000);

                }

                
            }
        });



        // ############################
        // SELECIONA NÓS-ALVOS - TECLA Q
        // ############################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'Q') {
                let todos = nodes.getIds();
                let nos_sel = []
                for (let i=0; i<todos.length; i++) {
                    let node = network.body.nodes[todos[i]];
                    if (node.shape.options.color.border == 'red') {
                        nos_sel.push(todos[i]);
                    }
                }
                network.selectNodes(nos_sel);

                if (nos_sel.length > 0) {
                    tempAlert(`Tecla ${event.key} - Seleciona nós-alvos (${nos_sel.length})`, 3000);
                } else if (nos_sel.length == 0) {
                    tempAlert(`Tecla ${event.key} - Não há nós-alvos no grafo`, 3000);
                }
            }
        });



        // ########################################################
        // DELETA TODOS OS NÓS QUE NÃO ESTÃO SELECIONADOS - TECLA r
        // ########################################################

        document.addEventListener('keydown', (event) => {
    
            if (event.key == "r" && network.getSelectedNodes().length > 0) {

                if (network.getSelectedNodes().length > 0) {

                    let n_1 = network.getSelectedNodes();

                    let tudo_1 = nodes.getIds();

                    let diff_2 = tudo_1.filter(x => !n_1.includes(x));

                    network.selectNodes(diff_2)

                    network.deleteSelected();

                    temp = nodes.getIds();

                    for (let i=0; i<temp.length; i++) {
                        nodes.update({id: temp[i], font: {color: '#777777'}})
                    }

                    network.unselectAll();

                    // VARIÁVEL GLOBAL DEFINIDA NA FUNÇÃO DA TECLA DEL
                    del_array = del_array.concat(diff_2);

                    // total_n: VARIÁVEL GLOBAL DENIFINDA NA FUNÇÃO DA TECLA DEL
                    let percentual = (del_array.length / total_n.length) * 100;
                    percentual = percentual.toFixed(1);

                    // INCLUIR AVISO PERMANENTE QUE NÓS FORAM DELETADOS
                    let el = document.createElement("div");
                    el.setAttribute("style","position:absolute;top:70px;right:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                    el.innerHTML = `Nós deletados (${del_array.length}/${total_n.length}) ${percentual}%`;
                    document.body.appendChild(el);

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Deleta nós não selecionados (${diff_2.length})`, 3000);
                }

            } else if (event.key == "r" && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione os nós que devem permanecer no grafo`, 3000);
            }
        });



        // ###############################################################################
        // EXIBE APENAS NÓS E ARESTAS DOS CAMINHOS MAIS CURTOS ENTRE OS NÓS-ALVOS - TECLA R
        // ###############################################################################

        var e_color = [];
        document.addEventListener('keydown', (event) => {
    
            if (event.key == "R") {

                // EXTRAI ARESTAS COLORIDAS
                let e = edges.getIds();
                //let e_color = [];
                for (let i=0; i<e.length; i++) {
                    let temp = network.body.edges[e[i]].options.width
                    if (temp == 8) {
                        e_color.push(e[i]);
                    }
                }

                if (e_color.length > 0) {
                
                    // EXTRAI NÓS CONECTADOS ÀS ARESTAS COLORIDAS
                    let n_color = [];
                    for (let i=0; i<e_color.length; i++) {
                        let temp = network.getConnectedNodes(e_color[i]);
                        for (let k=0; k<temp.length; k++) {
                            n_color.push(temp[k]);
                        }
                    }
                    
                    
                    // ELIMINA DUPLICIDADES
                    n_color = [...new Set(n_color)];
                    

                    // CALCULA DIFERENÇA E DELETA
                    let n_diff = nodes.getIds().filter(x => !n_color.includes(x));
                    network.setSelection({nodes: n_diff})
                    network.deleteSelected();
                    
                    let e_diff = edges.getIds().filter(x => !e_color.includes(x));	
                    network.setSelection({edges: e_diff})
                    network.deleteSelected();


                    // VARIÁVEL GLOBAL DEFINIDA NA FUNÇÃO DA TECLA DEL
                    del_array = del_array.concat(n_diff);

                    // total_n: VARIÁVEL GLOBAL DENIFINDA NA FUNÇÃO DA TECLA DEL
                    let percentual = (del_array.length / total_n.length) * 100;
                    percentual = percentual.toFixed(1);

                    // INCLUIR AVISO PERMANENTE QUE NÓS FORAM DELETADOS
                    let el = document.createElement("div");
                    el.setAttribute("style","position:absolute;top:70px;right:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                    el.innerHTML = `Nós deletados (${del_array.length}/${total_n.length}) ${percentual}%`;
                    document.body.appendChild(el);

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Exibe apenas nós interligados por arestas coloridas (${n_color.length})`, 3000);
                
                } else if (e_color.length == 0) {
                    tempAlert(`Tecla ${event.key} - Não há nós interligados por arestas coloridas`, 3000);
                }
            }
        })

        


        // ###################################
        // RECEBE CONSULTA LIVRE - TECLA s
        // ###################################

        document.addEventListener('keydown', (event) => {

            let texto = '';
            if (event.key == 's') {

                let a = network.getSelectedNodes();
                let temp = '';
                if (a.length == 1) {
                    temp = `id selecionado: ${network.body.nodes[a[0]].id}\\n`
                }
                //console.log(temp)

                let texto = prompt(temp + "Texto para consulta livre (curingas de palavras: * ):");

                //console.log(texto);
                if (texto != null && texto != '') {
                    async function copyOperation(texto) {
                        await navigator.clipboard.writeText(texto)
                    }

                    // Exibe informações na tela
                    tempAlert(`Tecla ${event.key} - Realiza consulta livre por "${texto}"`, 3000);

                    texto = '#_#' + texto + '#_#';
                    setTimeout(() => {copyOperation(texto)}, 1000);
                    //console.log('Texto copiado:', texto);

                }
            }
        });



        //function ajusta_altura() {

        //    // EXTRAI DIMENSÕES DA JANELA DO NAVEGADOR
        //    var w = window.innerWidth;
        //    let h = window.innerHeight;

        //    // AJUSTA ALTURA DA ÁREA DO GRAFO À ALTURA DA JANELA
        //    document.querySelector('#mynetwork').style.height = h - 20;
        //};

        //setTimeout(() => {ajusta_altura()}, 0);



        // ########################################
        // EXIBE IMAGENS DOS NÓS POR TIPO - TECLA t
        // ########################################

        // 15 itens (0 a 14)
        var img_names = ['pf_homem.png',
                        'pf_homem_com_bandeira.png',
                        'pf_mulher.png',
                        'pf_mulher_com_bandeira.png',
                        'pj_brasil_ativa.png',
                        'pj_brasil_ativa_com_bandeira.png', 
                        'pj_brasil_inativa.png',
                        'pj_brasil_inativa_com_bandeira.png',
                        'pj_exterior_ativa.png',
                        'pj_exterior_ativa_com_bandeira.png', 
                        'pj_exterior_inativa.png', 
                        'pj_exterior_inativa_com_bandeira.png', 
                        'endereco.png', 
                        'telefone.png',
                        'email.png'
                        ]

        var cont_t = 0
        var temp_t = nodes.getIds();

        document.addEventListener('keydown', (event) => {
            
            if (event.key == 't' && network.physics.options.solver == 'barnesHut') {
                
                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();
                
                // EXECUTAR APENAS UMA VEZ
                if (cont_t == 0) {
                    network.setOptions({nodes: {opacity: 0.1}, edges: {hidden: true}});
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Ativa modo transparência por tipos de nós - oculta rede (${temp_t.length})`, 3000);

                    
                } else if (cont_t == 1) {
                    let temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[0]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe homens (${temp.length}/${nodes.getIds().length})`, 3000);
                    

                    
                } else if (cont_t == 2) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[1]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[0]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe homens com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 3) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[2]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[1]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe mulheres (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 4) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[3]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[2]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe mulheres com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 5) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[4]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[3]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs ativas domiciliadas no Brasil (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 6) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[5]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[4]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs ativas domiciliadas no Brasil com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 7) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[6]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[5]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs inativas domiciliadas no Brasil (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 8) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[7]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[6]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs inativas domiciliadas no Brasil com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);
                

                } else if (cont_t == 9) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[8]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[7]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs ativas domiciliadas no exterior (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 10) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[9]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[8]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs ativas domiciliadas no exterior com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 11) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[10]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[9]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs inativas domiciliadas no exterior (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 12) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[11]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[10]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe PJs inativas domiciliadas no exterior com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 13) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[12]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[11]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe endereços (${temp.length}/${nodes.getIds().length})`, 3000);

                    
                } else if (cont_t == 14) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[13]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[12]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe telefones (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t == 15) {
                    temp = [];
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[14]) {
                            nodes.update({id: temp_t[i], opacity: 1});
                            temp.push(temp_t[i]);
                        } else if (network.body.nodes[temp_t[i]].options.image == img_names[13]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = cont_t + 1;
                    tempAlert(`Tecla ${event.key} - Exibe e-mails (${temp.length}/${nodes.getIds().length})`, 3000);

                    
                } else if (cont_t == 16) {
                    //window.location.reload();
                    for (var i = 0; i < temp_t.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t[i]].options.image == img_names[14]) {
                            nodes.update({id: temp_t[i], opacity: 0.1});
                        }
                    }
                    cont_t = 1;
                    tempAlert(`Tecla ${event.key} - Oculta rede (${temp_t.length})`, 3000);
                }

                //network.physics.physicsEnabled = true;
                //network.physics.startSimulation();
            }
        });



        // ##################################
        // ATIVA MODO TRANSPARÊNCIA - TECLA T
        // ##################################

        //var toggle_T = true;
        var con_t = [];
        var edg_t = [];
        var con_sel = [];
        var edg_sel = []

        document.addEventListener('keydown', (event) => {

            //  SÓ FUNCIONA NO MODO BARNES-HUT
            if (event.key == 'T' && network.physics.options.solver == 'barnesHut') {

                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();

                // TORNA TODOS OS NÓS E ARESTAS TRANSPARENTES
                network.setOptions({nodes: {opacity: 0.1}, edges: {hidden: true}});

                // AO FICAR EM CIMA DO NÓ
                network.on('hoverNode', function f(params) {

                    // EXTRAI NÓS CONECTADOS AO NÓ HOVER (con_t)
                    let node = params.node;
                    //console.log('hover node:', node);
                    con_t = network.getConnectedNodes(node);
                    con_t = con_t.concat(node);
                    //console.log('con_t:', con_t);

                    // EXTRAI ARESTAS CONECTADAS AO NÓ HOVER (edg_t)
                    edg_t = network.getConnectedEdges(node);
                    //console.log('edg_t:', edg_t);

                    // FAZ APARECER NÓS CONECTADOS AO NÓ HOVER
                    for (let i=0; i<con_t.length; i++) {
                        nodes.update({id: con_t[i], opacity: 1});
                    }

                    // FAZ APARECER ARESTAS CONECTADAS AO NÓ HOVER
                    for (let i=0; i<edg_t.length; i++) {
                        edges.update({id: edg_t[i], hidden: false});
                    }
                });

                // A SELEÇÃO DOS NÓS OCORRE POR MEIO DE OUTRO SCRIPT (CLICK)
                

                // ######################################################### ACIMA ESTÁ OK!

                // AO SAIR DE CIMA DO NÓ
                network.on('blurNode', function f(params) {

                    // OBTÉM NÓ BLUR (APENAS UM)
                    let blur_node = params.node;

                    // OBTÉM ARRAY DE NÓS CONECTADOS AOS NÓS SELECIONADOS (PODE SER MAIS DE UM)
                    let sel = network.getSelectedNodes();

                    // SE O NÓ BLUR ESTIVER ENTRE OS NÓS SELECIONADOS, ADICIONA OS NÓS E ARESTAS CONECTADOS AOS ARRAYS con_sel E edg_sel
                    for (let i=0; i<sel.length; i++) {                
                        if (blur_node == sel[i]) {
                            
                            //for (let j=0; j<sel.length; j++) {
                            let temp = network.getConnectedNodes(sel[i]);
                            con_sel = con_sel.concat(temp);
                            con_sel = con_sel.concat(sel[i]);
                            //}
                            //console.log('con_sel:', con_sel);
                
                            // OBTÉM ARESTAS CONECTADAS AO NÓ SELECIONADO
                            //for (let k=0; k<con_sel.length; k++) {
                            temp = network.getSelectedEdges(con_sel[i]);
                            edg_sel = edg_sel.concat(temp);
                            //}
                            //console.log('edg_sel:', edg_sel);
                        }
                    }

                    // SE NÃO TIVER NÓ SELECIONADO, FAZ OS NÓS E ARESTAS (con_t e edg_t) VOLTAREM A FICAR TRANSPARENTES
                    if (con_sel.length == 0) {
                        for (let i=0; i<con_t.length; i++) {
                            nodes.update({id: con_t[i], opacity: 0.1});
                        }
                        for (let i=0; i<edg_t.length; i++) {
                            edges.update({id: edg_t[i], hidden: true});
                        }

                    // SE TIVER NÓS E ARESTAS SELECIONADOS, SUBTRAI NÓS E ARESTAS SELECIONADOS (con_sel E edg_sel) DOS NÓS E ARESTAS CONTIDOS EM con_t E OS FAZ TRANSPARENTES
                    } else if (con_sel.length > 0) {
                        let diff_node = con_t.filter(x => !con_sel.includes(x));
                        for (let i=0; i<diff_node.length; i++) {
                            nodes.update({id: diff_node[i], opacity: 0.1});
                        }

                        // FALTA FAZER COM QUE AS ARESTAS DESAPAREÇAM APÓS BLUR
                        let diff_edg = edg_t.filter(x => !edg_sel.includes(x));
                        for (let i=0; i<diff_edg.length; i++) {
                            edges.update({id: diff_edg[i], hidden: true});
                        }
                    }
                });
                //toggle_T = false;

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Ativa modo transparência por conexões - oculta rede`, 3000);
                
            } else if (event.key == 'T' && network.physics.options.solver == 'barnesHut') {
                window.location.reload();               
                //ajusta_altura();
                //toggle_T = true;

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Desativa modo transparência`, 3000);
            }
        });



        // ####################################
        // EXIBE CENTRALIDADE DOS NÓS - TECLA u
        // ####################################

        // Encontrar uma forma de passar a informação apenas uma vez e não em todos os nós PJ
        // Descobrir forma de exibir informações em outro elemento e não mais em alert

        document.addEventListener('keydown', (event) => {
            if (event.key == "u") {

                let n = nodes.getIds();
                for (let i=0; i<n.length; i++) {
                    if (n[i].slice(0, 3) == 'PJ_') {
                        
                        // As informações exibidas estão no atributo 'algoritmos' de todos os nós PJ
                        let temp = network.body.nodes[n[i]].options.algoritmos;
                        alert(temp)
                        break;
                    }
                }

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Exibe informações sobre as medidas de centralidade`, 3000);
            }
        });



        // ################################################
        // EXIBE DISTRIBUIÇÃO DOS NÓS POR TAMANHO - TECLA U
        // ################################################

        //var a_tudo = nodes.getIds();
        var a_tudo = []

        document.addEventListener('keydown', (event) => {

            a_tudo = nodes.getIds();
            
            if (event.key == "U") {
                
                let a_100 = [];
                let a_95 = [];
                let a_90 = [];
                let a_85 = [];
                let a_80 = [];
                let a_75 = [];
                let a_70 = [];
                let a_65 = [];
                let a_60 = [];
                let a_55 = [];
                let a_50 = [];
                let a_45 = [];
                let a_40 = [];
                let a_35 = [];
                let a_30 = [];
                let a_25 = [];
                let a_20 = [];
                
                for (let i=0; i<a_tudo.length; i++) {

                    if (network.body.nodes[a_tudo[i]].options.size > 99.9999999999 && network.body.nodes[a_tudo[i]].options.size <= 100) {
                        a_100.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 99.9999999999 && network.body.nodes[a_tudo[i]].options.size >= 95) {
                        a_95.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 95 && network.body.nodes[a_tudo[i]].options.size >= 90) {
                        a_90.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 90 && network.body.nodes[a_tudo[i]].options.size >= 85) {
                        a_85.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 85 && network.body.nodes[a_tudo[i]].options.size >= 80) {
                        a_80.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 80 && network.body.nodes[a_tudo[i]].options.size >= 75) {
                        a_75.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 75 && network.body.nodes[a_tudo[i]].options.size >= 70) {
                        a_70.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 70 && network.body.nodes[a_tudo[i]].options.size >= 65) {
                        a_65.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 65 && network.body.nodes[a_tudo[i]].options.size >= 60) {
                        a_60.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 60 && network.body.nodes[a_tudo[i]].options.size >= 55) {
                        a_55.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 55 && network.body.nodes[a_tudo[i]].options.size >= 50) {
                        a_50.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 50 && network.body.nodes[a_tudo[i]].options.size >= 45) {
                        a_45.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 45 && network.body.nodes[a_tudo[i]].options.size >= 40) {
                        a_40.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 40 && network.body.nodes[a_tudo[i]].options.size >= 35) {
                        a_35.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 35 && network.body.nodes[a_tudo[i]].options.size >= 30) {
                        a_30.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 30 && network.body.nodes[a_tudo[i]].options.size >= 25) {
                        a_25.push(a_tudo[i]);
                    }
                    else if (network.body.nodes[a_tudo[i]].options.size < 25 && network.body.nodes[a_tudo[i]].options.size >= 20) {
                        a_20.push(a_tudo[i]);
                    }

                }
                //console.log(a_100);
                //console.log(a_95);
                //console.log(a_90);
                //console.log(a_85);
                //console.log(a_80);
                //console.log(a_75);
                //console.log(a_70);
                //console.log(a_65);
                //console.log(a_60);
                //console.log(a_55);
                //console.log(a_50);
                //console.log(a_45);
                //console.log(a_40);
                //console.log(a_35);
                //console.log(a_30);
                //console.log(a_25);
                //console.log(a_20);
                
                let texto_espaco = `DISTRIBUIÇÃO DOS NÓS POR TAMANHO:\nTamanho      100:   ${a_100.length}\nTamanho 95 - 99:   ${a_95.length}\nTamanho 90 - 94:   ${a_90.length}\nTamanho 85 - 90:   ${a_85.length}\nTamanho 80 - 84:   ${a_80.length}\nTamanho 75 - 79:   ${a_75.length}\nTamanho 70 - 74:   ${a_70.length}\nTamanho 65 - 69:   ${a_65.length}\nTamanho 60 - 64:   ${a_60.length}\nTamanho 55 - 59:   ${a_55.length}\nTamanho 50 - 54:   ${a_50.length}\nTamanho 45 - 49:   ${a_45.length}\nTamanho 40 - 44:   ${a_40.length}\nTamanho 35 - 39:   ${a_35.length}\nTamanho 30 - 34:   ${a_30.length}\nTamanho 25 - 29:   ${a_25.length}\nTamanho 20 - 24:   ${a_20.length}\n-----------------------------\nTOTAL:                    ${a_tudo.length}`;
                
                texto_espaco = texto_espaco.replaceAll(" 0\\n", "  \\n");

                tempAlert(`Tecla ${event.key} - Exibe distribuição dos nós por tamanho`, 3000);
                
                alert(texto_espaco);
            }
        });



        // ###########################
        // EXIBE COMUNIDADES - TECLA v
        // ###########################

        // Quando os atributos dos nós são alterados por meio de "nodes.update", "network.setOptions" não funciona para retornar ao estado anterior
        // Ao mudar paletar de cores, alterar também a variável 'num_cores' (resolvido: criado mecanismo de variação automática)

        var toggle_v = true;
        var n_comunidades = 0;

        document.addEventListener('keydown', (event) => {
            if (event.key == "v" && toggle_v == true) {

                //network.physics.physicsEnabled = false;
                //network.physics.stopSimulation();

                let a = nodes.getIds();

                if (a.length > 500) {
                    let conf = confirm('A identificacao de comunidades em mais de 500 nós pode demorar vários minutos. Deseja continuar?')
                    if (!conf) {
                        return;
                    }
                }

                // PALETA DE CORES ALTO CONTRASTE
                // https://www.schemecolor.com/high-contrast.php
                //var hcc = ["#8931EF", "#F2CA19", "#FF00BD", "#0057E9", "#87E911", "#E11845"]
                //var hcc = ["#8931EF", "#F2CA19", "#FF00BD", "#0057E9", "#87E911"]

                // tab10 (tableau 10)
                //var hcc = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"]

                var hcc = ["rgba(137, 49, 239, 0.5)",
                           "rgba(242, 202, 25, 0.5)",
                           "rgba(255, 0, 189, 0.5)",
                           "rgba(0, 87, 233, 0.5)",
                           "rgba(135, 233, 17, 0.5)",
                           "rgba(225, 24, 69, 0.5)",

                           "rgba(0, 255, 0, 0.5)",
                           "rgba(255, 255, 0, 0.5)",
                           "rgba(0, 255, 255, 0.5)",
                           "rgba(255, 0, 255, 0.5)"
                          ]

                //var hcc = ["rgba(0, 255, 0, 1)",
                //           "rgba(255, 255, 0, 1)",
                //           "rgba(0, 255, 255, 1)",
                //           "rgba(255, 0, 255, 1)"
                //          ]

                //var hcc = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"]

                for (let i=0; i<a.length; i++) {
                    
                    let n = network.body.nodes[a[i]].options.z_group

                    if (n > n_comunidades) {
                        n_comunidades = n;
                    }
                    
                    // CALCULA ÍNDICE CÍCLICO DO ARRAY
                    //let num_cores = hcc.length
                    index_cir = Math.max(0, n % hcc.length) || 0;

                    // MUDA COR DO FUNDO DO NÓ
                    nodes.update({id: a[i], color: {background: hcc[index_cir]}});

                    //console.log(i)
                }
                toggle_v = false;

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Adiciona cores aos nós para destacar comunidades (${n_comunidades})`, 3000);

            } else if (event.key == "v" && toggle_v == false) {
                
                let a = nodes.getIds();
                for (let i=0; i<a.length; i++) {
                    nodes.update({id: a[i], color: {background: 'transparent'}});
                }
                toggle_v = true;

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Remove cores dos nós`, 3000);
            }

            //network.physics.physicsEnabled = true;
            //network.physics.startSimulation();

        });



        // #################################
        // EXIBE NÓS DE CADA GRUPO - TECLA V
        // #################################

        var num_grupos = 1;
        var num_grupos_total = network.body.nodes[nodes.getIds()[0]].options.z_group_total;
        var arr2 = [];

        document.addEventListener('keydown', (event) => {
            
            if (event.key == 'V') {

                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();

                network.setOptions({nodes: {opacity: 0.0}, edges: {hidden: true}});
                
                let arr1 = nodes.getIds();
                
                if (arr2.length > 0) {
                    for (let i=0; i<arr2.length; i++) {
                        nodes.update({id: arr2[i], opacity: 0.0});
                    }
                }

                arr2 = [];
                
                for (let i=0; i<arr1.length; i++) {
                    let temp = network.body.nodes[arr1[i]].options.z_group;
                    if (temp == num_grupos) {
                        arr2.push(network.body.nodes[arr1[i]].options.id)
                        nodes.update({id: arr1[i], opacity: 1});
                    }
                }

                
                if (num_grupos == num_grupos_total) {
                    tempAlert(`Tecla ${event.key} - 1 Seleciona grupo ${num_grupos}/${num_grupos_total} (${arr2.length})`, 3000);
                    num_grupos = 1;
                
                } else if (num_grupos < num_grupos_total) {
                    tempAlert(`Tecla ${event.key} - 2 Seleciona grupo ${num_grupos}/${num_grupos_total} (${arr2.length})`, 3000);
                    num_grupos = num_grupos + 1
                }
                
                network.physics.physicsEnabled = true;
                network.physics.startSimulation();
            }
        });



        // #############################################################
        // ADICIONA COR AO FUNDO DOS NÓS ANCESTRAIS DO NÓ-ALVO - TECLA w
        // #############################################################

        // CRIA ARRAY COM NÓS ANCESTRAIS
        var ancestors = [];
        var nodes_all = nodes.getIds();
        for (let i=0; i<nodes_all.length; i++) {
            if (network.body.nodes[nodes_all[i]].options.z_ancestor == 'true') {
                ancestors.push(nodes_all[i]);
            }
        }

        // INSERE COR NO FUNDO DOS NÓS ANCESTRAIS
        var toggle_w = true;
        document.addEventListener('keydown', (event) => {

            if (event.key == 'w') {

                if (toggle_w == true) {
                    for (let i=0; i<ancestors.length; i++) {
                        nodes.update({id: ancestors[i], color: {background: 'rgb(0, 0, 255, 0.2)', hover: {background: 'rgb(0, 0, 255, 0.2)'}}});
                        //nodes.update({id: ancestors[i], color: {background: 'rgb(0, 255, 0, 1)', hover: {background: 'rgb(0, 255, 0, 1)'}}});
                    }

                    if (ancestors.length > 0) {
                        tempAlert(`Tecla ${event.key} - Adiciona cores para destacar cadeia de nós originária do nó-alvo (${ancestors.length})`, 3000);
                        toggle_w = false;

                    } else if (ancestors.length == 0) {
                        tempAlert(`Tecla ${event.key} - Não há nós-alvos para destacar cadeia originária`, 3000);
                    }
                    
                } else if (toggle_w == false) {
                    for (let i=0; i<ancestors.length; i++) {
                        nodes.update({id: ancestors[i], color: {background: 'transparent', hover: {background: 'transparent'}}});
                    }
                    toggle_w = true;

                    // Exibe informações na tela (não está exibindo)
                    tempAlert(`Tecla ${event.key} - Remove cores dos nós (${ancestors.length})`, 3000);
                }
            }
        });



        // ##################################
        // SELECIONA NÓS ANCESTRAIS - TECLA W
        // ##################################

        document.addEventListener('keydown', (event) => {

            if (event.key == 'W') {
                arr1 = nodes.getIds();
                arr2 = ancestors;

                if (arr2.length == 0) {
                    tempAlert(`Tecla ${event.key} - Não há nó-alvo único para selecionar cadeia originária`, 3000);
                    return;
                }

                let difference = arr1.filter(x => !arr2.includes(x));
                network.selectNodes(difference);
                network.deleteSelected();


                // VARIÁVEL GLOBAL DEFINIDA NA FUNÇÃO DA TECLA DEL
                del_array = del_array.concat(difference);

                // total_n: VARIÁVEL GLOBAL DENIFINDA NA FUNÇÃO DA TECLA DEL
                let percentual = (del_array.length / total_n.length) * 100;
                percentual = percentual.toFixed(1);

                // INCLUIR AVISO PERMANENTE QUE NÓS FORAM DELETADOS
                let el = document.createElement("div");
                el.setAttribute("style","position:absolute;top:70px;right:21px;border:5px solid rgba(255, 255, 123, 1);border-radius: 5px;background-color:rgba(255, 255, 123, 1);");
                el.innerHTML = `Nós deletados (${del_array.length}/${total_n.length}) ${percentual}%`;
                document.body.appendChild(el);

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Exclui nós que não pertecem à cadeia de nós originária do nó-alvo (${ancestors.length})`, 3000);
            }
        });



        // ####################################################
        // ADICIONA COR AO FUNDO DOS NÓS SELECIONADOS - TECLA x
        // ####################################################

        // Verde, amarelo, azul, lilás
        var cores_fundo = ["rgba(0, 255, 0, 1)", "rgba(255, 255, 0, 1)", "rgba(0, 255, 255, 1)", "rgba(255, 0, 255, 1)"];
        var c_color = 0;
        var nos_coloridos = [];

        document.addEventListener('keydown', (event) => {
            if (event.key == 'x' && network.getSelectedNodes().length > 0) {

                let todos = network.getSelectedNodes();
                for (let i=0; i<todos.length; i++) {
                    nodes.update({id: todos[i], color: {background: cores_fundo[c_color], hover: {background: cores_fundo[c_color]}}});
                    nodes.update({id: todos[i], font: {color: '#777777'}});
                    nos_coloridos.push(todos[i]);
                    network.unselectAll();
                }
                
                if (network.getSelectedNodes().length == 0) {
                    c_color = c_color + 1;
                }
                
                if (c_color > cores_fundo.length - 1) {
                    c_color = 0;
                }

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Adiciona cores sequenciais aos nós selecionados (verde, amarelo, azul e lilás)`, 3000);
            
            } else if (event.key == 'x' && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione nós em sequência para adicionar cores (verde, amarelo, azul e lilás)`, 3000);
            }
        });



        // ####################################################
        // REMOVE COR AO FUNDO DOS NÓS SELECIONADOS - TECLA X
        // ####################################################

        document.addEventListener('keydown', (event) => {
            if (event.key == 'X') {
                for (let i=0; i<nos_coloridos.length; i++) {
                    nodes.update({id: nos_coloridos[i], color: {background: 'transparent', hover: {background: 'transparent'}}});
                    nodes.update({id: nos_coloridos[i], font: {color: '#777777'}});
                }
                network.unselectAll();
                nos_coloridos = [];
                c_color = 0;

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Tecla ${event.key} - Remove cores dos nós`, 3000);
            }
        });



        // #################################
        // CONSULTA NÓS ACUMULADOS - TECLA y
        // #################################

        document.addEventListener('keydown', (event) => {
            
            if (event.key == 'y') {
                
                // COPIA PARA O CLIPBOARD
                async function copyOperation(param) {
                    await navigator.clipboard.writeText(param)    
                }

                // RESTAURA O ARRAY ARMAZENADO EM 'localStorage' E GERADO PELAS TECLAS '+' E '-'
                let temp = JSON.parse(localStorage.getItem("nos_acum"));
                console.log(temp);


                let conf_destaca = ''
                if (temp.length > 1) {
                    conf_destaca = confirm("Deseja destacar caminhos mais curtos entre os nós?");
                }


                let temp2 = [];
                for (let i = 0; i < temp.length; i++) {
                    let j = temp[i] + "***" + 1 + "***" + conf_destaca;
                    temp2.push(j);
                }


                setTimeout(() => {copyOperation(temp2)}, 500)

                tempAlert(`Tecla ${event.key} - Abre lista de nós-alvos em nova aba (${temp.length})`, 3000);
            }
        });



        // ########################################
        // EXIBE TODOS OS NÓS PF, PJ E PE - TECLA z
        // ########################################

        var controle_z = true;
        var toggle_z = true;
        var last_central_node = '';
        var last_node = '';
        var old_title = '';
        var new_title = '';
        var nod = [];
        var num_nos = 0;
    
        document.addEventListener('keydown', (event) => {
            if (event.key == "z" && toggle_z == true) {

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Ativa modo automático de visualização`, 3000);

                toggle_z = false;

                network.physics.physicsEnabled = false;
                network.physics.stopSimulation();

                var pos = network.getPositions();
                var a = nodes.getIds();

                maiores = [];

                // LOOP PARA OBTER APENAS SEQUÊNCIA NUMÉRICA
                for (let i=0; i<a.length; i++) {

                    // LOOP SOBRE IDS DOS NÓS
                    for (let k=0; k<a.length; k++) {
                        let value_no = network.body.nodes[a[k]].options.value;
                        if (value_no == i) {

                            // APENAS NÓS COM 2 OU MAIS NÓS ADJACENTES
                            if (network.getConnectedNodes(a[k]).length > 1) {
                                maiores.push(a[k]);
                            }
                        }
                    }
                }

                // NÓS ORDENADOS SEGUNDO MAIOR NÚMERO DE ADJACÊNCIAS
                maiores.reverse();

                // CRIA LISTA FINAL COM NÓS COM MAIOR GRAU, SEGUIDOS DE SEUS ADJACENTES
                for (let i=0; i<maiores.length; i++) {
                    nod.push([maiores[i], maiores[i]]);
                    let temp = network.getConnectedNodes(maiores[i]);
                    for (let k=0; k<temp.length; k++) {
                        let temp2 = []
                        temp2 = temp2.concat(temp[k]);
                        temp2 = temp2.concat(maiores[i]);
                        nod.push(temp2);
                    }
                }

                //console.log('nod:', nod)

                network.setOptions({'edges': {'font': {'size': 0}}});
                network.setOptions({"edges": {"hidden": false}});

                // INÍCIO DA FUNÇÃO DE LOOP
                var i = 0;              
                function myLoop() {        
                setTimeout(function() {

                    //console.log('nod:', nod)

                    // CONDIÇÃO QUE VERIFICA O FINAL DO ARRAY 'nod' E ENCERRA A FUNÇÃO
                    if (i == nod.length) {
                        nodes.update({id: last_node, font: {color: '#777777'}});
                        //nodes.update({id: last_central_node, color: {background: 'transparent'}});
                        nodes.update({id: last_central_node, color: {background: 'transparent', hover: {background: 'transparent'}}});
                        nodes.update({id: last_node, label: old_title});
                        //num_nos = 0;
                        network.unselectAll();
                        network.fit({
                            position: {x:0, y:0},
                            scale: 1.0,
                            animation: true
                        });
                        nod = [];
                        network.physics.physicsEnabled = true;
                        network.physics.startSimulation();
                        toggle_z = true;
                        return;
                    }
                    
                    //console.log('nod:', nod);

                    // SAI DO PROGRAMA (SEM ESTA CONDIÇÃO, ESTAVA DANDO ERRO NA PRÓXIMA LINHA)
                    if (nod.length == 0) {
                        return
                    }

                    x = pos[nod[i][0]]['x'];
                    y = pos[nod[i][0]]['y'];

                    last_central_node = nod[i][1];
                    last_node = nod[i][0]

                    nodes.update({id: nod[i][1], color: {background: 'rgb(0, 0, 255, 0.2)', hover: {background: 'rgb(0, 0, 255, 0.2)'}}});
                    
                    if (i >= 1) {
                        if (nod[i][1] != nod[i - 1][1]) {
                            nodes.update({id: nod[i - 1][1], font: {color: '#777777'}});
                            nodes.update({id: nod[i - 1][1], color: {background: 'transparent', hover: {background: 'transparent'}}});
                            setTimeout(function() {
                                network.moveTo({
                                    position: {x:x, y:y},
                                    scale: 0.5,
                                    animation: true
                                });
                            }, 1000);
                        }

                        // ESTOU VENDO ESTE ITEM
                        network.body.nodes[nod[i - 1][0]].options.label = old_title;
                        nodes.update({id: nod[i - 1][0], font: {color: '#777777'}});
                    }

                    // CRIANDO NOVO RÓTULO
                    let edg1 = network.getConnectedEdges(nod[i][0]);
                    let edg2 = network.getConnectedEdges(nod[i][1]);
                    let common_edg = edg1.filter(value => edg2.includes(value));
                    let edg_title = network.body.edges[common_edg[0]].title;
                    old_title = network.body.nodes[nod[i][0]].options.label;

                    var num_nos_cen = network.body.nodes[nod[i][1]].options.value
                    if (num_nos <= parseInt(num_nos_cen)) {
                        if (num_nos == 0) {
                            num_nos = 0;
                        } 
                    }

                    // PARA OS CASOS DE TE_, EN_, EM_ E PARA O PRIMEIRO ITEM DA LISTA
                    if (edg_title == undefined || i == 0) {
                        edg_title = '';
                    }
                    new_title = old_title + '\\n' + edg_title + ' (' + num_nos + ' de ' + num_nos_cen + ')';

                    num_nos = num_nos + 1;
                    if (num_nos > parseInt(num_nos_cen)) {
                        num_nos = 0;
                    }

                    network.body.nodes[nod[i][0]].options.label = new_title;
                    
                    network.moveTo({
                        position: {x:x, y:y},
                        scale: 1.0,
                        animation: true
                    });
                    
                    network.selectNodes([nod[i][0]]);
                    nodes.update({id: nod[i][0], font: {color: 'red'}});
                    if (controle_z == false) {
                        controle_z = true;
                        nodes.update({id: last_node, font: {color: '#777777'}});
                        nodes.update({id: last_central_node, color: {background: 'transparent', hover: {background: 'transparent'}}});
                        network.unselectAll();
                        network.redraw();
                        return;
                    }

                    if (i < nod.length) {        
                        myLoop();            
                        if (i == nod.length - 1) {
                            network.setOptions({'edges': {'font': {'size': 15}}});
                            network.setOptions({"edges": {"hidden": false}});
                        }
                    }

                    i++;               

                }, 3500)
                    
                }
                
                myLoop();

                //nodes.update({id: nod[i][1], color: {background: 'transparent', hover: {background: 'transparent'}}});
                

            
            } else if (event.key == "z" && toggle_z == false) {

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Desativa modo automático de visualização`, 3000);

                controle_z = false;
                toggle_z = true;
                network.setOptions({'edges': {'font': {'size': 15}}});
                network.setOptions({"edges": {"hidden": false}});
                nodes.update({id: last_node, font: {color: '#777777'}});
                nodes.update({id: last_central_node, color: {background: 'transparent', hover: {background: 'transparent'}}});
                nodes.update({id: last_node, label: old_title});
                network.unselectAll();
                network.redraw();
                network.fit({
                    position: {x:0, y:0},
                    scale: 1.0,
                    animation: true
                });
                controle_z = true;
                last_central_node = '';
                last_node = '';
                old_title = '';
                new_title = '';
                nod = [];
                num_nos = 0;
                network.physics.physicsEnabled = true;
                network.physics.startSimulation();
            }
        });

        //// EXEMPLO DE LOOP PERIÓDICO EM JAVASCRIPT
        //let x = 0;
        //function loop () {
        //    console.log("hi");
        //    x++;
        //
        //    if (x<3) {
        //        setTimeout(loop, 5000);
        //    }
        //}
        //loop();



        // ################################################################
        // AJUSTA MASSA DOS NÓS PROPORCIONALMENTE AO TAMANHO - TECLA ESPAÇO
        // ################################################################

        // Usa variável global 't_ref'

        var pula_100 = true
        var n_bag = [];

        //var desativa_teclas_m_n = false
        document.addEventListener('keydown', (event) => {
            
            if (event.key == ' ') {

                //desativa_teclas_m_n = true;

                // EXECUTA APENAS UMA VEZ, FAZENDO t_ref = 100
                if (pula_100 == true) {
                    t_ref = 100
                    pula_100 = false 
                }             

                let all = nodes.getIds();

                let sel_e = []

                for (let i=0; i<all.length; i++) {

                    if (n_bag.includes(all[i])) {
                        continue;
                    }

                    
                    let size = network.body.nodes[all[i]].options.size;

                    // VERIFICA SE O TAMANHO DO NÓ ESTÁ NO INTERVALO ENTRE MAIOR OU IGUAL A t_ref E MENOR QUE t_ref + 5
                    if (size >= t_ref && size < (t_ref + 5)) {
                        sel_e.push(all[i])
                    }
                    
                    if (size >= t_ref) {
                        nodes.update({id: all[i], mass: (size/20) ** 3});
                        n_bag.push(all[i]);
                        n_bag = [...new Set(n_bag)];
                    }
                    
                }

                if (t_ref == 20) {
                    sel_e = [];
                }
                //console.log(sel_e)
                network.selectNodes(sel_e);

                tempAlert(`Tecla Espaço - Aumenta massa dos nós proporcionalmente ao tamanho >= ${t_ref} (${n_bag.length})`, 3000);

                t_ref = t_ref - 5;

                if (t_ref == 15) {
                    //t_ref = 100;
                    t_ref = 40;
                    pula_100 = true;
                    n_bag = [];
                    for (let i=0; i<all.length; i++) {
                        nodes.update({id: all[i], mass: 1});
                    }
                }

            }
        });



        // #####################################################################
        // ALTERNA SELEÇÃO DOS NÓS ADJACENTES QUE SÃO ORIGEM E DESTINO - TECLA ;
        // #####################################################################

        var toggle_pv = 0;
        var no_sel_pv = [];
        var ar_adj = [];
        var temp_to = [];
        var temp_from = [];

        document.addEventListener('keydown', (event) => {

            // O SCRIPT FOI INICIALMENTE CONCEBIDO PARA SER USADO COM MAIS DE UM NÓ SELECIONADO. DEPOIS AJUSTAMOS PARA APENAS UM NÓ SELECIONADO
            if (event.key == ";") {



                if (toggle_pv == 0 && network.getSelectedNodes().length != 1) {
                    tempAlert(`Tecla ${event.key} - Selecione um nó para aplicar a função`, 3000);
                    return;
                }



                // EXECUTA APENAS NA PRIMEIRA VEZ QUE A TECLA ; É PRESSIONADA OU QUANDO toggle_pv == 0
                if (toggle_pv == 0 && temp_to.length == 0 && temp_from.length == 0) {
                
                    // EXTRAI ARRAY COM NÓ SELECIONADO
                    no_sel_pv = network.getSelectedNodes();
                    
                    // EXTRAI TODAS AS ARESTAS ADJACENTES AOS NÓS SELECIONADOS
                    for (let i=0; i<no_sel_pv.length; i++) {
                        ar_adj.push(network.getConnectedEdges(no_sel_pv[i]));
                    }
            
                    // ELIMINA DUPLICIDADES NO ARRAY DE ARESTAS - atentar para o [0]
                    ar_adj = [...new Set(ar_adj[0])];

                    // EXTRAI NÓS DESTINO
                    for (let i=0; i<ar_adj.length; i++) {
                        if (network.body.edges[ar_adj[i]] != undefined) {
                            temp_to.push(network.body.edges[ar_adj[i]].toId);
                        }
                    }
                    temp_to = [...new Set(temp_to)];
                    temp_to = temp_to.filter(x => !no_sel_pv.includes(x));

                    // EXTRAI NÓS ORIGEM
                    for (let i=0; i<ar_adj.length; i++) {
                        if (network.body.edges[ar_adj[i]] != undefined) {
                            temp_from.push(network.body.edges[ar_adj[i]].fromId);
                        }
                    }
                    temp_from = [...new Set(temp_from)];
                    temp_from = temp_from.filter(x => !no_sel_pv.includes(x));

                    //console.log('atualização das variáveis');
                    //console.log('temp_from:', temp_from);
                    //console.log('temp_to:', temp_to);
                }

                

                // SELECIONA NÓS ADJACENTES QUE SÃO DESTINO DO NÓ SELECIONADO (PONTA DAS SETAS QUE SAEM -->)
                if (toggle_pv == 0 && no_sel_pv.length == 1 && temp_to.length > 0) {
                
                    network.unselectAll();
                    network.redraw();
                    network.selectNodes(temp_to);

                    let cn = network.getConnectedNodes(no_sel_pv[0]).length

                    // RETORNA COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
                    for (let i=0; i<no_sel_pv.length; i++) {
                        nodes.update({id: no_sel_pv[i], font: {color: '#777777'}});
                    }
                    
                    toggle_pv = 1;

                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes de DESTINO (-->) das arestas do selecionado (${temp_to.length}/${cn})`, 3000);



                } else if (toggle_pv == 0 && no_sel_pv.length == 1 && temp_to.length == 0) {

                    network.unselectAll();
                    network.redraw();
                    toggle_pv = 1;

                    let cn = network.getConnectedNodes(no_sel_pv[0]).length
                    
                    // RETORN COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
                    for (let i=0; i<no_sel_pv.length; i++) {
                        nodes.update({id: no_sel_pv[i], font: {color: '#777777'}});
                    }

                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes de DESTINO (-->) das arestas do selecionado (${temp_to.length}/${cn})`, 3000);

                    
                
                // SELECIONA NÓS ADJACENTES QUE CHEGAM NO NÓ-ALVO (ORIGEM)
                } else if (toggle_pv == 1 && no_sel_pv.length == 1 && temp_from.length > 0) {

                    // DESSELECIONA TODOS OS NÓS
                    network.unselectAll();
                    network.redraw();
                    network.selectNodes(temp_from);

                    let cn = network.getConnectedNodes(no_sel_pv[0]).length

                    // RETORN COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
                    for (let i=0; i<no_sel_pv.length; i++) {
                        nodes.update({id: no_sel_pv[i], font: {color: '#777777'}});
                    }
                
                    toggle_pv = 2;
                    
                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes de ORIGEM (<--) das arestas dos nó selecionado (${temp_from.length}/${cn})`, 3000);




                } else if (toggle_pv == 1 && no_sel_pv.length == 1 && temp_from.length == 0) {
                    
                    network.unselectAll();
                    network.redraw();
                    toggle_pv = 2;

                    let cn = network.getConnectedNodes(no_sel_pv[0]).length

                    // RETORN COR CINZA PARA OS RÓTULOS DOS NÓS SELECIONADOS
                    for (let i=0; i<no_sel_pv.length; i++) {
                        nodes.update({id: no_sel_pv[i], font: {color: '#777777'}});
                    }

                    tempAlert(`Tecla ${event.key} - Seleciona nós adjacentes de ORIGEM (<--) das arestas dos nó selecionado (${temp_from.length}/${cn})`, 3000);


            
                // VOLTA A SELECIONAR OS NÓS ORIGINAIS
                } else if (toggle_pv == 2 && no_sel_pv.length == 1) {
                
                    network.unselectAll();
                    network.selectNodes(no_sel_pv);
            
                    // REINICIALIZA VARIÁVEIS
                    toggle_pv = 0;
                    no_sel_pv = [];
                    ar_adj = [];
                    temp_to = [];
                    temp_from = [];
            
                    tempAlert(`Tecla ${event.key} - Retorna à seleção inicial (${network.getSelectedNodes().length})`, 3000);
                }

            }

            // USADO APENAS PARA TESTES
            //console.log(toggle_pv);
            //console.log(no_sel_pv);
            //console.log(network.getSelectedNodes().length);

        })



        // ##########################################################
        // SELECIONA NÓS ADICIONADOS NA REQUISIÇÃO ANTERIOR - TECLA /
        // ##########################################################

        var toggle_bar = 0;

        document.addEventListener('keydown', (event) => {

            if (event.key == "/" && toggle_bar == 0 && network.getSelectedNodes().length == 0) {

                let nn = nodes.getIds()
                let li = []
                for (let i=0; i<nn.length; i++) {
                    if (network.body.nodes[nn[i]].options.z_novo == '1') {
                        li.push(nn[i])
                    }
                }
                network.selectNodes(li);

                toggle_bar = 1;

                tempAlert(`Tecla ${event.key} - Seleciona nós adicionados na requisição anterior (${li.length}/${nn.length})`, 3000);
            
            } else if (event.key == "/" && toggle_bar == 1) {
                let nn = nodes.getIds()
                network.unselectAll();
                toggle_bar = 0;
                tempAlert(`Tecla ${event.key} - Desseleciona todos os nós (${nn.length})`, 3000);
            }

        });



        // ##################################################
        // SELECIONA NÓS POR TIPO DE IMAGEM - TECLA . (PONTO)
        // ##################################################

        // 15 itens (0 a 14)
        var img_names2 = ['pf_homem.png',
                        'pf_homem_com_bandeira.png',
                        'pf_mulher.png',
                        'pf_mulher_com_bandeira.png',
                        'pj_brasil_ativa.png',
                        'pj_brasil_ativa_com_bandeira.png', 
                        'pj_brasil_inativa.png',
                        'pj_brasil_inativa_com_bandeira.png',
                        'pj_exterior_ativa.png',
                        'pj_exterior_ativa_com_bandeira.png', 
                        'pj_exterior_inativa.png', 
                        'pj_exterior_inativa_com_bandeira.png', 
                        'endereco.png', 
                        'telefone.png',
                        'email.png'
                        ]

        var cont_t2 = 0
        var temp_t2 = nodes.getIds();

        document.addEventListener('keydown', (event) => {
            
            //if (event.key == '.' && network.physics.options.solver == 'barnesHut') {
            if (event.key == '.') {
                                    
                if (cont_t2 == 0) {
                    network.unselectAll();
                    let temp = [];
                    let nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[0]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona homens (${temp.length}/${nodes.getIds().length})`, 3000);
                    

                    
                } else if (cont_t2 == 1) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[1]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona homens com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 2) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[2]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona mulheres (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 3) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[3]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona mulheres com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 4) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[4]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs ativas domiciliadas no Brasil (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 5) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[5]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs ativas domiciliadas no Brasil com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 6) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[6]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs inativas domiciliadas no Brasil (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 7) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[7]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs inativas domiciliadas no Brasil com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);
                

                } else if (cont_t2 == 8) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[8]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs ativas domiciliadas no exterior (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 9) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[9]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs ativas domiciliadas no exterior com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 10) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[10]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs inativas domiciliadas no exterior (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 11) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[11]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona PJs inativas domiciliadas no exterior com bandeira vermelha (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 12) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[12]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona endereços (${temp.length}/${nodes.getIds().length})`, 3000);

                    
                } else if (cont_t2 == 13) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[13]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona telefones (${temp.length}/${nodes.getIds().length})`, 3000);


                } else if (cont_t2 == 14) {
                    network.unselectAll();
                    temp = [];
                    nos = [];
                    for (var i = 0; i < temp_t2.length; i++) {

                        // CASO O NÓ TENHA SIDO DELETADO
                        if (network.body.nodes[temp_t2[i]] == undefined) {
                            continue
                        }

                        if (network.body.nodes[temp_t2[i]].options.image == img_names2[14]) {
                            nos.push(temp_t2[i]);
                            temp.push(temp_t2[i]);
                        }
                    }
                    network.selectNodes(nos);
                    cont_t2 = cont_t2 + 1;
                    tempAlert(`Tecla ${event.key} - Seleciona e-mails (${temp.length}/${nodes.getIds().length})`, 3000);

                    
                } else if (cont_t2 == 15) {
                    network.unselectAll();
                    cont_t2 = 0;
                    tempAlert(`Tecla ${event.key} - Desseleciona todos os nós (${temp_t.length})`, 3000);
                }
            }
        });



        // ###################################################################
        // ABRE POPUP COM RÓTULOS DAS ARESTAS SELECIONADAS - TECLA , (VÍRGULA)
        // ###################################################################

        var popup;
        var nn;

        // EXECUTA FUNÇÃO createPopup SE HOUVER NÓ SELECIONADO, PASSANDO COMO PARÂMETRO O ARRAY DOS NÓS SELECIONADOS (nn)
        document.addEventListener("keydown", function(event) {
        if (event.key === ",") {
            nn = network.getSelectedNodes();
            if (nn.length > 0) {
                createPopup(nn);
                tempAlert(`Tecla ${event.key} - Abre popup com nós selecionados (${nn.length})`, 3000);
            } else if (nn.length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um ou mais nós para abrir popup contendo nós selecionados`, 3000);
            }
        }
        });


        // DECLARA FUNÇÃO QUE CRIA POPUP
        function createPopup(nn) {

        let lista_ids = [];

        // EXTRAI ARRAY DE RÓTULOS (labels)
        let labels = nn.map(function(nodeId) {
            let node = network.body.data.nodes.get(nodeId);
            lista_ids.push(node.id + '_' + node.label);
        });


        // ordena os labels conforme especificação
        lista_ids.sort(function(a, b) {
            let order = ["PF_", "PJ_", "PE_", "EN_", "TE_", "EM_"];
            let aIndex = order.findIndex(function(prefix) {
            return a.startsWith(prefix);
            });
            let bIndex = order.findIndex(function(prefix) {
            return b.startsWith(prefix);
            });
            return aIndex - bIndex;
        });

        let popupText = "";
        let prefix_atual = '';
            
        let lista_pf = [];
        let lista_pj = [];
        let lista_pe = [];
        let lista_en = [];
        let lista_te = [];
        let lista_em = [];
            
        for (let i=0; i<lista_ids.length; i++) {

            if (lista_ids[i].split('_')[0] == 'PF') {
                lista_pf.push(lista_ids[i].split('_')[2]);
            }

            if (lista_ids[i].split('_')[0] == 'PJ') {
                lista_pj.push(lista_ids[i].split('_')[2]);
            }

            if (lista_ids[i].split('_')[0] == 'PE') {
                lista_pe.push(lista_ids[i].split('_')[2]);
            }

            if (lista_ids[i].split('_')[0] == 'EN') {
                lista_en.push(lista_ids[i].split('_')[2]);
            }

            if (lista_ids[i].split('_')[0] == 'TE') {
                lista_te.push(lista_ids[i].split('_')[2]);
            }

            if (lista_ids[i].split('_')[0] == 'EM') {
                lista_em.push(lista_ids[i].split('_')[2]);
            }
        }

        lista_pf.sort();
        lista_pj.sort();
        lista_pe.sort();
        lista_en.sort();
        lista_te.sort();
        lista_em.sort();

        if (lista_pf.length > 0) {
            popupText += `<b>PF (${lista_pf.length})</b><br>${lista_pf.join("<br>")}<br><br>`;
        }

        if (lista_pj.length > 0) {
            popupText += `<b>PJ (${lista_pj.length})</b><br>${lista_pj.join("<br>")}<br><br>`;
        }

        if (lista_pe.length > 0) {
            popupText += `<b>PE (${lista_pe.length})</b><br>${lista_pe.join("<br>")}<br><br>`;
        }

        if (lista_en.length > 0) {
            popupText += `<b>EN (${lista_en.length})</b><br>${lista_en.join("<br>")}<br><br>`;
        }

        if (lista_te.length > 0) {
            popupText += `<b>TE (${lista_te.length})</b><br>${lista_te.join("<br>")}<br><br>`;
        }

        if (lista_em.length > 0) {
            popupText += `<b>EM (${lista_em.length})</b><br>${lista_em.join("<br>")}<br><br>`;
        }
            
        if (!popup || popup.closed) {
            popup = window.open('', target="_blank", windowFeature='left=50,top=50,width=700,height=400,scrollbars=1');
            popup.document.write("<html><head><title>Nós Selecionados</title><style>body{font-family: Arial, sans-serif; font-size: 14px; color: #333; line-height: 1.5; padding: 10px;} b{font-size: 16px; font-weight: bold;}</style></head><body></body></html>");
        }

        var popupBody = popup.document.body;
        popupBody.innerHTML = `<b>NÓS SELECIONADOS (${lista_ids.length}):</b><br><br>` + popupText;
        }

        network.on("click", function(params) {
        if (popup) {
            popup.close();
        }
        });



        // ##############################################
        // AUMENTA A MASSA DOS NÓS SELECIONADOS - TECLA ]
        // ##############################################

        var par_n2 = 1

        document.addEventListener('keydown', (event) => {

            if (event.key == ']' && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length > 0) {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}});
                }
                
                let nod = network.getSelectedNodes();

                for (let i=0; i<nod.length; i++) {

                    let size = network.body.nodes[nod[i]].options.size

                    // Aumenta a massa em 5 unidades
                    nodes.update({
                        id: nod[i],
                        mass: par_n2 + 5
                    });
                }

                // Atualiza variável
                par_n2 = par_n2 + 5

                network.redraw();
            
                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Aumenta massa dos nós selecionados em 5 unidades. Fator: ${par_n2} (${network.getSelectedNodes().length})`, 3000);
            
            } else if (event.key == ']' && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um ou mais nós para aumentar suas massas em 5 unidades`, 3000);
            }
        })



        // ###############################################
        // DIMINUI A MASSA DOS NÓS SELECIONADOS - TECLAS [
        // ###############################################

        document.addEventListener('keydown', (event) => {

            if (event.key == '[' && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length > 0) {

                if (network.physics.adaptiveTimestepEnabled == false) {
                    network.setOptions({"nodes": {"physics": true}});
                }

                //if ((par_n2 - 5) <= 1) {
                //    par_n2 = 1;
                //}

                par_n2 = par_n2 - 5

                if (par_n2 <= 1) {
                    par_n2 = 1
                }
                
                let nod = network.getSelectedNodes();

                for (let i=0; i<nod.length; i++) {

                    let size = network.body.nodes[nod[i]].options.size

                    // Diminui a massa em 5 unidades
                    nodes.update({
                        id: nod[i],
                        //mass: par_n2 - 5
                        mass: par_n2
                    });
                }

                // Atualiza a variável
                //par_n2 = par_n2 - 5

                //if (par_n2 <= 1) {
                //    par_n2 = 1
                //}

                // Exibe informações na tela
                tempAlert(`Tecla ${event.key} - Diminui massa dos nós selecionados em 5 unidades. Fator: ${par_n2} (${network.getSelectedNodes().length})`, 3000);

                network.redraw();

            } else if (event.key == '[' && network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length == 0) {
                tempAlert(`Tecla ${event.key} - Selecione um ou mais nós para diminuir suas massas em 5 unidades`, 3000);
            }

        })



        // ###############################
        // EXIBE ATALHOS NA TELA - TECLA ?
        // ###############################

        var el_3 = '';
        var toggle_int = 0;

        document.addEventListener('keydown', (event) => {

            if(event.key == "?" && toggle_int == 0){

                let ajuste_tela = "ATALHOS:<br>Ajuste da tela: F11 b<br>";
                let posicao = "Posição dos nós: F5 k K f p Clique+Arrasto Shift+(L/R) ] [ m M Espaço n N Shift+(U/D)<br>";
                let selecao = "Seleção dos nós: Clique Ctrl+Clique a A ç Ç I j J q Q . ; /<br>";
                let visualizacao = "Visualização de nós e arestas: Esc BotãoDeRolagem c e l i t T V z<br>";
                let cor = "Cor dos nós: x X w v<br>";
                let sites = "Sites externos: d D g G 1 2 4 5 6 7 8<br>";
                let delecao = "Deleção de nós: Del r R W<br>";
                let consulta = "Consulta à base de dados: o O s y<br>";
                let outras = "Outras funções: h ? L u U + - , |";

                el_3 = document.createElement("div");
                
                el_3.setAttribute("style","position:absolute; top:30px; left:21px; border:5px solid rgba(178, 255, 178, 0); border-radius: 5px; background-color:rgba(178, 255, 178, 0); font-size: 18px; color: rgba(59, 190, 30, 1); font-family: serif; background-color: rgba(59, 190, 30, 0.05); border-radius: 5px");
                
                el_3.innerHTML = ajuste_tela + posicao + selecao + visualizacao + cor + sites + delecao + consulta + outras;
                
                document.body.appendChild(el_3);

                el_3.style.opacity = "1.0";

                toggle_int = 1

                tempAlert(`Tecla ${event.key} - Exibe teclas de atalhos`, 3000);

            } else if (event.key == "?" && toggle_int == 1) {

                el_3.style.display = "none";

                toggle_int = 0

                tempAlert(`Tecla ${event.key} - Oculta teclas de atalhos`, 3000);

            }
        });

        

        // ################################
        // DESSELECIONA TODOS OS NÓS
        // ################################

        network.on('click', (event) => {

            let sel = network.getSelectedNodes();
            //let n = nodes.getIds().length;

            // Quando se clica no fundo, sel.length fica igual a 0 (truque)
            if (sel.length == 0) {
                network.unselectAll();
                
                //con_nodes_a = [];
                sel_anterior = [];
                
                // EXTRAÍDO DA TECLA A
                //temp_a = [];

                // VARIÁVEL DECLARADA NO SCRIPT DA TECLA q
                target_id = [];
                con = [];

                // Exibe informações na tela
                //if (sel_anterior.length > 0) {
                tempAlert(`Clique na área do grafo - Desseleciona todos os nós (${nodes.getIds().length})`, 3000);
                //}
            }

            // Exibe informações na tela
            //if (sel.length > 0) {
            //    tempAlert(`Clique na área do grafo - Desseleciona todos os nós`, 3000);
            //}

            // Limpa variável global 'sel_anterior' do script da TECLA 'a', para não continuar a sequência de seleção do ponto onde parou
            sel_anterior = [];

            // Limpa vairávels globais do script da TECLA ';', para não reiniciar a sequência de seleção do ponto onde parou
            toggle_pv = 0;
            no_sel_pv = [];
            ar_adj = [];
            temp_to = [];
            temp_from = [];

        });



        // ############################################################
        // DESSELECIONA ARESTAS DE NÓS DESSELECIONADOS COM CTRL + CLICK
        // ############################################################

        // AO SAIR DE CIMA DO NÓ E EVENTUALMENTE TER DESSELECIONADO O RESPECTIVO NÓ
        network.on('blurNode', function f(params) {

            // OBTÉM LISTA DE NÓS SELECIONADOS (O QUE ACABOU DE SER DESSELECIONADO NÃO APARECE NESTA LISTA)
            let ns = network.getSelectedNodes();

            // DESSELECIONA TUDO
            network.unselectAll();
            network.redraw();

            // SELECIONA APENAS NÓS MANTIDOS SELECIONADOS APÓS DESS
            //console.log(ns);
            network.selectNodes(ns);

        });



        // ###############################################################################################
        // AUMENTA E DIMINUI TAMANHO REFERNCIAL DOS NÓS - TECLAS SHIFT + 'ARROW UP' / SHIFT + 'ARROW DOWN'
        // ###############################################################################################

        var t_ref = 40

        document.addEventListener('keydown', (event) => {

            if (event.shiftKey && event.key == 'ArrowUp'){

                if (t_ref < 100) {
                    t_ref = t_ref + 5
                }

//                if (screen_height == 0) {
//                    let element = document.getElementById("mynetwork");
//                    let canvas = element.querySelector('canvas');
//                    screen_height = parseInt(canvas.getAttribute('height'));
//                }
//                element = document.getElementById("mynetwork");
//                screen_height = screen_height + 10;
//                element.style.height = screen_height.toString() + 'px';
                //console.log(screen_height);

                //alert(`Tamanho referencial do no para ativacao da tecla n: ${t_ref}`)
                //console.log(t_ref)
                network.redraw();

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Teclas Shift + ${event.key} - Aumenta tamanho referencial dos nós em 5 unidades (${t_ref})`, 3000);

            } else if (event.shiftKey && event.key == 'ArrowDown') {
                
                if (t_ref > 20) {
                    t_ref = t_ref - 5
                }
                
//                if (screen_height == 0) {
//                    let element = document.getElementById("mynetwork");
//                    let canvas = element.querySelector('canvas');
//                    screen_height = parseInt(canvas.getAttribute('height'));
//                }
//                element = document.getElementById("mynetwork");
//                screen_height = screen_height - 10;
//                element.style.height = screen_height.toString() + 'px';
                //console.log(screen_height);

                //alert(`Tamanho referencial do no para ativacao da tecla n: ${t_ref}`)
                //console.log(t_ref)
                network.redraw();

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Teclas Shift + ${event.key} - Diminui tamanho referencial dos nós em 5 unidades (${t_ref})`, 3000);

            }
        });



        // ###############################################
        // DESSELECIONA TODOS OS NÓS - TECLAS 'ARROW DOWN'
        // ###############################################

        //document.addEventListener('keydown', (event) => {

        //    if (event.key == "ArrowDown") {
        //        network.unselectAll();
        //        network.redraw();

        //    }

        //})



        // ###########################################################################################
        // AUMENTA E DIMINUI A TAMANHO DA ARESTA - TECLAS SHIFT + 'ARROW LEFT' / SHIFT + 'ARROW RIGHT'
        // ###########################################################################################

        var tamanho_aresta = 300
        document.addEventListener('keydown', (event) => {

            if (event.shiftKey && event.key == 'ArrowRight'){
                tamanho_aresta += 30;
                network.setOptions({"edges": {"length": tamanho_aresta}});

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Teclas Shift + ${event.key} - Aumenta tamanho das arestas em 30 unidades (${tamanho_aresta})`, 3000);
                
            } else if (event.shiftKey && event.key == 'ArrowLeft') {
                tamanho_aresta -= 30;
                network.setOptions({"edges": {"length": tamanho_aresta}});

                // Exibe informações na tela (não está exibindo)
                tempAlert(`Teclas Shift + ${event.key} - Diminui tamanho das arestas em 30 unidades (${tamanho_aresta})`, 3000);
            }
        });



        // ########################################################
        // MUDA COR DO TEXTO DO NÓ PARA VERMELHO QUANDO SELECIONADO
        // ########################################################

        network.on("select", function (params) {

            let selectedNodeId2 = params.nodes;
            //console.log('selected:', selectedNodeId2);
            for (id in params.nodes) {
                let node = network.body.nodes[params.nodes[id]];
                //console.log('node:', node);
                node.setOptions({
                    font: {
                    color: "#ff0000"
                    }
                });
            }

            // Exibe informações na tela
            let temp = network.getSelectedNodes();

            let tamanho = ''
            let massa2 = ''
            let grupo = ''
            if (network.getSelectedNodes().length == 1) {
                let temp_t = Math.round(network.body.nodes[params.nodes[id]].options.size * 100) / 100
                //tamanho = `. Tamanho: ${network.body.nodes[params.nodes[id]].options.size}`
                tamanho = `. Tamanho: ${temp_t}`
                
                massa2 = `. Massa: ${Math.round(network.body.nodes[params.nodes[id]].options.mass * 100) / 100}`
                //Math.round(num * 100) / 100

                grupo = `. Grupo: ${network.body.nodes[params.nodes[id]].options.z_group}/${network.body.nodes[params.nodes[id]].options.z_group_total}`
            }
            
            tempAlert(`Clique sobre nó - Seleciona nó (${temp.length})${tamanho}${massa2}${grupo}`, 3000);

        });

        network.on("deselectNode", function (params) {
            let deselectedNodeId2 = params.previousSelection.nodes; // ACRESCENTEI ['id']
            if (deselectedNodeId2.length > 100) { // DESSELECIONA TUDO DE UMA SÓ VEZ A PARTIR DE 100 NÓS SELECIONADOS
                network.unselectAll();
                return;
            }
            //console.log('deselected:', deselectedNodeId2);
            for (id in params.previousSelection.nodes) {
                let node = network.body.nodes[params.previousSelection.nodes[id]['id']];
                //console.log('node:', node);
                
                // INCLUÍDO POR DAR ERRO 'id' UNDEFINED
                if (node != undefined) { 
                    nodes.update({id: node['id'], // deselectedNodeId
                            font: {
                                color: '#777777'
                            },
                    });
                } 
                // INCLUIDO POR DAR ERRO 'id' UNDEFINED
            }
        });



        // ############################################################
        // APROXIMA NÓ SELECIONADO QUANDO CLICA SOBRE ELE - DUPLO CLICK
        // ############################################################

//        function esconde() {
//            network.setOptions({nodes: {scaling: {label: {enabled: true, drawThreshold: 100}}}, edges: {hidden: true}});
//        }
//        function exibe() {
//            network.setOptions({nodes: {scaling: {label: {enabled: true, drawThreshold: 8}}}, edges: {hidden: false}});
//        }

//        network.on('doubleClick', function() {

            // SÓ FUNCIONA NO LAYOUT BARNES-HUT E COM APENAS UM NÓ SELECIONADO
//            if (network.physics.options.solver == 'barnesHut' && network.getSelectedNodes().length == 1) {

//                // Evita compartamento errático no modo transparência (arestas são exibidas)
//                if (toggle_t == false) {
//                    return;
//                }            

//                let ss = network.getSelectedNodes();
//                if (ss.length == 1 ) {
//                    network.setOptions({physics: {enabled: false}});
//                    let p = '';
//                    let num = 0;
//                    num = 700 + nodes.length / 10;
//                    //console.log(num + 'ms');
//                    for (let i=0; i<ss.length; i++) {
//                        esconde();
//                        p = network.getPosition(ss[i]);
//                        network.moveTo({
//                            position: {x: p['x'], y: p['y']},
//                            scale: 1,
//                            offset: {x:0, y:0},
//                            animation: {
//                            duration: 500,
//                            easingFunction: 'linear'
//                            }
//                        });
//                        setTimeout(exibe, num);
//                        network.releaseNode();
//                    }
//                    network.setOptions({physics: {enabled: true}});

                // ###################################

//                let ss = network.getSelectedNodes();
                //if (ss.length == 1 ) {

                // DESABILITA ELEMENTOS PARA EVITAR TRAVAMENTO DA IMAGEM (MELHROA DESEMPENHO)
//                network.physics.physicsEnabled = false;
//                network.physics.stopSimulation();
                //network.setOptions({"edges": {"hidden": true}});
                //network.setOptions({nodes: {scaling: {label: {enabled: true, drawThreshold: 100}}}});
                
                // FÓRMULA QUE CALCULA O TEMPO DE TRANSIÇÃO. USADO TAMBÉM NA FUNÇÃO setTimeout
//                let time = 1000;
//                if (1000 + nodes.getIds().length - 500 < 1000) {
//                    time = 1000;
//                } else {
//                    time = 1000 + nodes.getIds().length - 1000;
//                }

                // PARTE PRINCIPAL
//                for (let i=0; i<ss.length; i++) {
//                    p = network.getPosition(ss[i]);
//                    network.moveTo({
//                        position: {x: p['x'], y: p['y']},
//                        scale: 1,
//                        offset: {x:0, y:0}
                        //animation: {
                        //duration: time,
                        //easingFunction: 'linear'
                        //}
//                    });
//                }
                //network.fit({nodes: con_nodes, maxZoomLevel: 1, animation: {duraton: time}});

                // FUNÇÃO EXECUTADA APÓS A PARTE PRINCIPAL
//                function f2() {

                    // REABILITA RECURSOS APÓS PARTE PRINCIPAL
                    //network.setOptions({physics: {enabled: true}});
//                    network.setOptions({"edges": {"hidden": false}});
//                    network.setOptions({nodes: {scaling: {label: {enabled: true, drawThreshold: 8}}}});
//                }
                
                //setTimeout(f2, time + 100);

                //setTimeout(network.physics.stopSimulation, time + 200);





//                }
//            }
//        });



        // ##################################
        // ALTERA FLATICON E TÍTULO DA PÁGINA
        // ##################################
        var link = document.querySelector("link[rel~='icon']");
        if (!link) {
            link = document.createElement('link');
            link.rel = 'icon';
            document.getElementsByTagName('head')[0].appendChild(link);
        }
        //link.href = 'https://cdn-icons-png.flaticon.com/512/4803/4803070.png';
        
        // Colocar logo na pasta do arquivo sinarc.py
        link.href = 'logo.png';
        

        // EXTRAI NÚMERO DE CAMADAS SOLICITADAS
        var no_pj = nodes.getIds();
        var n_cam = '';
        for (let i=0; i < no_pj.length; i++) {
            if (no_pj[i].slice(0, 3) == 'PJ_') {
                n_cam = network.body.nodes[no_pj[i]].options.camadas_solicitadas;
                //console.log(n_cam);
                break;
            };
        };


        // EXTRAI NÚMERO DE NÓS E FORMATA PARA STRING
        // https://stackoverflow.com/questions/9461621/format-a-number-as-2-5k-if-a-thousand-or-more-otherwise-900
        function nFormatter(num, digits) {
            const lookup = [
                { value: 1, symbol: "" },
                { value: 1e3, symbol: "k" },
                { value: 1e6, symbol: "M" },
                { value: 1e9, symbol: "G" },
                { value: 1e12, symbol: "T" },
                { value: 1e15, symbol: "P" },
                { value: 1e18, symbol: "E" }
            ];
            const rx = /\.0+$|(\.[0-9]*[1-9])0+$/;
            var item = lookup.slice().reverse().find(function(item) {
                return num >= item.value;
            });
            return item ? (num / item.value).toFixed(digits).replace(rx, "$1") + item.symbol : "0";
        }

        let temp = ''
        temp = nFormatter(nodes.getIds().length, 1);

        //let datt = new Date();
        //let hora = datt.getHours();
        //let minuto = datt.getMinutes();
        //document.title = n_cam + '/' + temp + ' SINARC ' + hora + ':' + minuto;

        document.title = n_cam + '/' + temp + ' SINARC';



        // ##############################################
        // AJUSTA ALTURA DA ÁREA DO GRAFO AUTOMATICAMENTE
        // ##############################################

        // Existe outra forma de ajustar a altura área do grafo: TECLA b

        function ajusta_altura() {

            // EXTRAI DIMENSÕES DA JANELA DO NAVEGADOR
            var w = window.innerWidth;
            let h = window.innerHeight;

            // AJUSTA ALTURA DA ÁREA DO GRAFO À ALTURA DA JANELA
            document.querySelector('#mynetwork').style.height = h - 20;

        };

        setTimeout(() => {ajusta_altura()}, 0);



    """

    # Substitui texto do arquivo HTML
    antes = '</script>\n    </body>'
    depois = code + '\n' + antes
    temp = temp.replace(antes, depois)

    # Atualiza arquivo HTML
    fout = open("grafo_final.html", "w")
    fout.write(temp)
    fout.close()



    ##########################
    # EXIBE GRAFO NO NAVEGADOR
    ##########################

    # Exibe grafo no navegador padrão
    webbrowser.open('grafo_final.html', new=1, autoraise=True)



    ########################################
    # CRIA ARQUIVO ZIP PARA COMPARTILHAMENTO
    ########################################

    # Cria arquivo ZIP para compartilhamento
    with ZipFile(f'{diretorio}/arquivo_sinarc.zip', 'w') as zip_object:

        # Adiciona arquivos ao arquivo ZIP criado
        zip_object.write('bandeira_vermelha.png')
        zip_object.write('pf_homem.png')
        zip_object.write('pf_homem_com_bandeira.png')
        zip_object.write('pf_mulher.png')
        zip_object.write('pf_mulher_com_bandeira.png')
        zip_object.write('pj_brasil_ativa.png')
        zip_object.write('pj_brasil_ativa_com_bandeira.png')
        zip_object.write('pj_brasil_inativa.png')
        zip_object.write('pj_brasil_inativa_com_bandeira.png')
        zip_object.write('pj_exterior_ativa.png')
        zip_object.write('pj_exterior_ativa_com_bandeira.png')
        zip_object.write('pj_exterior_inativa.png')
        zip_object.write('pj_exterior_inativa_com_bandeira.png')
        zip_object.write('telefone.png')
        zip_object.write('email.png')
        zip_object.write('endereco.png')
        zip_object.write('print_screen.png')
        zip_object.write('logo.png')
        zip_object.write('help.html')
        zip_object.write('grafo_final.html')

import traceback
def main():

    while True:
        try:
            # As demais funções são chamadas dentro desta função
            captura_cnpj()
        except Exception:
            print(traceback.format_exc())
            print('SINARC não localizou o servidor REDE CNPJ. Nova tentativa em 10 segundos...')
            time.sleep(10)


if __name__ == '__main__':

    try:
        main()

    # Caso ocorra erro na execução
    except:
        print()
        print('#' * 56)
        print('MENSAGEM DE ERRO:')
        print('#' * 56)
        print()

        # Atualiza arquivo com histórico de mensagens de erro
        with open('mensagens_de_erro.txt', 'a') as f:

            # Utiliza a função "format_exc()" do módulo "traceback" do Python para formatar a exceção em uma string e capturá-la, armazenando-a na variável "erro".
            erro = format_exc().split('\n\n')[-1]
            print(erro)

            # Extrai hora atual
            now = datetime.now()
            agora = now.strftime(r"%A %d/%m/%Y %H:%M:%S")

            # Atribui à variável o texto armazenado na área de transferência
            texto_area_transferencia = pyperclip.paste()

            # imprime resumo do erro
            temp = '=' * 50 + '\n' + agora + '\n\n' + erro + '\n\n' + texto_area_transferencia + '\n\n'
            f.write(temp)
